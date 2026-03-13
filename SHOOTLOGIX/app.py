"""
app.py — ShootLogix
Flask REST API for the full production logistics SPA.
Run: python3 app.py  →  http://localhost:5002
"""
import csv
import hashlib
import io
import json
import os
import tempfile
import threading
from flask import Flask, jsonify, request, render_template, abort, Response, g, make_response

from db_compat import get_table_names as _compat_get_table_names, get_backend_info
from database import (
    init_db, get_db,
    get_productions, get_production, create_production,
    get_departments, seed_departments,
    get_shooting_days, create_shooting_day, update_shooting_day,
    delete_shooting_day, get_shooting_day,
    get_events_for_day, create_event, update_event, delete_event, delete_events_for_day,
    get_boats, create_boat, update_boat, delete_boat,
    get_boat_functions, create_boat_function, update_boat_function,
    delete_boat_function, delete_boat_assignment_by_function,
    get_boat_assignments, create_boat_assignment, update_boat_assignment, delete_boat_assignment,
    get_picture_boats, create_picture_boat, update_picture_boat, delete_picture_boat,
    get_picture_boat_assignments, create_picture_boat_assignment,
    update_picture_boat_assignment, delete_picture_boat_assignment,
    delete_picture_boat_assignment_by_function,
    get_transport_vehicles, create_transport_vehicle, update_transport_vehicle, delete_transport_vehicle,
    get_transport_assignments, create_transport_assignment, update_transport_assignment,
    delete_transport_assignment, delete_transport_assignment_by_function,
    get_fuel_entries, upsert_fuel_entry, delete_fuel_entry, delete_fuel_entries_for_assignment,
    get_fuel_machinery, create_fuel_machinery, update_fuel_machinery, delete_fuel_machinery,
    get_fuel_locked_prices, set_fuel_locked_price, delete_fuel_locked_price,
    get_helpers, create_helper, update_helper, delete_helper,
    get_helper_assignments, create_helper_assignment, update_helper_assignment,
    delete_helper_assignment, delete_helper_assignment_by_function,
    get_helper_schedules,
    get_security_boats, create_security_boat, update_security_boat, delete_security_boat,
    get_security_boat_assignments, create_security_boat_assignment,
    update_security_boat_assignment, delete_security_boat_assignment,
    delete_security_boat_assignment_by_function,
    get_fnb_services,
    get_fuel_logs,
    get_transport_schedules,
    get_guard_schedules,
    get_budget, get_daily_budget,
    get_history, get_activity_feed, undo_last_boat_assignment, undo_history_entry,
    get_setting, set_setting,
    # Budget snapshots & price log (AXE 6.3)
    create_budget_snapshot, get_budget_snapshots, get_budget_snapshot,
    compare_budget_snapshots, delete_budget_snapshot,
    log_price_change, get_price_change_log,
    working_days,
    # New modules
    get_location_schedules, upsert_location_schedule,
    delete_location_schedule, delete_location_schedule_by_id,
    lock_location_schedules, auto_fill_locations_from_pdt,
    sync_pdt_day_to_locations, remove_pdt_film_days_for_date,
    get_guard_location_schedules, upsert_guard_location_schedule,
    delete_guard_location_schedule, lock_guard_location_schedules,
    sync_guard_location_from_locations, update_guard_location_nb_guards,
    get_fnb_tracking, upsert_fnb_tracking, delete_fnb_tracking, get_fnb_summary,
    # FNB v2
    get_fnb_categories, create_fnb_category, update_fnb_category, delete_fnb_category,
    get_fnb_items, create_fnb_item, update_fnb_item, delete_fnb_item,
    get_fnb_entries, upsert_fnb_entry, delete_fnb_entry, get_fnb_budget_data,
    # Location sites CRUD
    get_location_sites, create_location_site, update_location_site,
    delete_location_site, rename_location_in_schedules,
    # Guard posts CRUD
    get_guard_posts, create_guard_post, update_guard_post,
    delete_guard_post, rename_guard_post_in_schedules,
    # Guard Camp (Base Camp guards)
    get_guard_camp_workers, create_guard_camp_worker, update_guard_camp_worker,
    delete_guard_camp_worker,
    get_guard_camp_assignments, create_guard_camp_assignment,
    update_guard_camp_assignment, delete_guard_camp_assignment,
    delete_guard_camp_assignment_by_function,
    # PDT cascade (AXE 7.2)
    cascade_preview, cascade_apply,
    # Export preferences (AXE 2.2)
    get_export_preference, save_export_preference, get_module_date_range,
    # Comments & Notifications (AXE 9)
    get_comments, get_comment_counts, create_comment, delete_comment,
    get_notifications, get_unread_notification_count, create_notification,
    create_notifications_for_production, mark_notification_read, mark_all_notifications_read,
    # AXE 10.2 — Duplication
    duplicate_assignment, duplicate_fnb_category,
    # P2.3 — Physical Vessels
    get_physical_vessels, create_physical_vessel, update_physical_vessel,
    delete_physical_vessel, check_vessel_cross_module_conflict,
    # P2.4 — Soft delete restore
    restore_entity, restore_fnb_category,
    # P6.4 — Exchange rates
    get_exchange_rates, upsert_exchange_rate, get_latest_rate,
    # Daily checklists
    generate_daily_checklist, get_daily_checklist, check_checklist_item,
)

from validation import (ValidationError, validate_assignment, validate_assignment_dates,
    validate_fuel_entry, validate_shooting_day, validate_date_range, validate_positive_number,
    validate_required, validate_guard_schedule, validate_assignment_overlap,
    validate_required_fields, validate_numeric_fields, validate_entity_name)

app = Flask(__name__)

# ─── Background Export System ─────────────────────────────────────────────────
import uuid
import time as _time

_export_jobs = {}  # { job_id: { status, path, filename, created_at, error } }
_EXPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'exports')
os.makedirs(_EXPORT_DIR, exist_ok=True)


# ─── Price Change Logger Helper (AXE 6.3) ────────────────────────────────────
_PRICE_FIELDS = {
    'daily_rate_estimate', 'daily_rate_actual', 'rate_estimate', 'rate_actual',
    'price_override', 'price_p', 'price_f', 'price_w', 'global_deal',
    'unit_price', 'unit_price_estimate', 'unit_price_actual',
}

def _log_price_changes(prod_id, entity_type, entity_id, entity_name, old_data, new_data):
    """Compare old and new data dicts; log any price field changes."""
    user_id = getattr(g, 'user_id', None)
    nickname = getattr(g, 'nickname', None)
    override_reason = new_data.get('override_reason')
    for field in _PRICE_FIELDS:
        old_val = old_data.get(field)
        new_val = new_data.get(field)
        if new_val is not None and old_val != new_val:
            try:
                log_price_change(
                    prod_id, entity_type, entity_id, entity_name,
                    field, float(old_val) if old_val is not None else None,
                    float(new_val), user_id, nickname,
                    override_reason=override_reason if field == 'price_override' else None
                )
            except (ValueError, TypeError):
                pass


def _validate_price_override(data, old_data):
    """If price_override is being set or changed, require override_reason."""
    new_po = data.get('price_override')
    if new_po is None:
        return None
    old_po = old_data.get('price_override') if old_data else None
    # Convert for comparison
    try:
        new_po_f = float(new_po) if new_po is not None else None
        old_po_f = float(old_po) if old_po is not None else None
    except (ValueError, TypeError):
        new_po_f, old_po_f = new_po, old_po
    if new_po_f != old_po_f and new_po_f is not None and new_po_f != 0:
        reason = (data.get('override_reason') or '').strip()
        if not reason:
            return jsonify({"error": "override_reason is required when changing price_override"}), 400
    return None


def _cleanup_old_exports():
    """Remove exports older than 1 hour."""
    now = _time.time()
    to_remove = []
    for job_id, job in _export_jobs.items():
        if now - job.get("created_at", 0) > 3600:
            to_remove.append(job_id)
            if job.get("path") and os.path.exists(job["path"]):
                try:
                    os.unlink(job["path"])
                except OSError:
                    pass
    for jid in to_remove:
        del _export_jobs[jid]


# ─── Export Date Filtering Helpers (AXE 2.1 / 2.3) ────────────────────────────

def _export_date_params():
    """Extract from/to date range from query params."""
    return request.args.get("from"), request.args.get("to")


def _filter_assignments_by_date(rows, date_from, date_to):
    """Filter assignment rows whose date range overlaps [date_from, date_to].
    Assignments with no dates are always included."""
    if not date_from and not date_to:
        return rows
    filtered = []
    for r in rows:
        start = (r.get("start_date") or "")[:10]
        end = (r.get("end_date") or "")[:10]
        if not start and not end:
            filtered.append(r)
            continue
        if date_from and end and end < date_from:
            continue
        if date_to and start and start > date_to:
            continue
        filtered.append(r)
    return filtered


def _filter_entries_by_date(entries, date_from, date_to):
    """Filter entries (fuel, FNB) by their date field."""
    if not date_from and not date_to:
        return entries
    filtered = []
    for e in entries:
        d = (e.get("date") or "")[:10]
        if not d:
            filtered.append(e)
            continue
        if date_from and d < date_from:
            continue
        if date_to and d > date_to:
            continue
        filtered.append(e)
    return filtered


def _export_fname(prod_name, module, date_from, date_to, ext):
    """Build export filename with optional date range suffix."""
    from datetime import datetime as dt
    date_str = dt.now().strftime('%y%m%d')
    if date_from and date_to:
        range_suffix = f"_{date_from}_{date_to}".replace("-", "")
    else:
        range_suffix = ""
    return f"{prod_name}_{module}_{date_str}{range_suffix}.{ext}"


@app.route("/api/exports/<job_id>", methods=["GET"])
def api_export_status(job_id):
    """Check export job status. Returns download link when ready."""
    job = _export_jobs.get(job_id)
    if not job:
        return jsonify({"error": "Export not found"}), 404
    if job["status"] == "done":
        return jsonify({
            "status": "done",
            "download_url": f"/api/exports/{job_id}/download",
            "filename": job["filename"],
        })
    elif job["status"] == "error":
        return jsonify({"status": "error", "error": job.get("error", "Unknown error")})
    return jsonify({"status": "processing"})


@app.route("/api/exports/<job_id>/download", methods=["GET"])
def api_export_download(job_id):
    """Download completed export file."""
    job = _export_jobs.get(job_id)
    if not job or job["status"] != "done" or not job.get("path"):
        return jsonify({"error": "Export not ready"}), 404
    from flask import send_file
    return send_file(
        job["path"],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=job["filename"],
    )


@app.errorhandler(ValidationError)
def handle_validation_error(e):
    return jsonify({"error": "Validation failed", "fields": e.errors}), 422


def jsonify_cached(data):
    """Return a JSON response with ETag. If client sends matching If-None-Match, return 304."""
    body = json.dumps(data, separators=(',', ':'), sort_keys=True)
    etag = '"' + hashlib.md5(body.encode()).hexdigest() + '"'
    if_none_match = request.headers.get('If-None-Match')
    if if_none_match == etag:
        return Response(status=304)
    resp = make_response(body)
    resp.headers['Content-Type'] = 'application/json'
    resp.headers['ETag'] = etag
    resp.headers['Cache-Control'] = 'private, no-cache'
    return resp

# ─── Auth: Register blueprint & protect all /api/ routes ─────────────────────
from auth.routes import auth_bp
from auth.admin_routes import admin_bp
from auth.tokens import decode_access_token
from auth.rbac import check_role_access, check_permission_access, get_user_allowed_tabs
from auth.models import (
    get_membership, ensure_user_permissions, get_user_global_permissions,
    user_has_entity_restrictions, get_user_allowed_entity_ids,
)

app.register_blueprint(auth_bp)
app.register_blueprint(admin_bp)

# Routes that do NOT require authentication
AUTH_EXEMPT_PREFIXES = (
    "/api/auth/",    # Login, refresh, logout
    "/api/health",   # Health check
)


@app.before_request
def enforce_auth():
    """
    Global auth middleware: require a valid JWT for all /api/ routes
    except auth endpoints and health check. Static files and HTML pages
    are not affected.
    """
    path = request.path

    # Only protect API routes
    if not path.startswith("/api/"):
        return None

    # Skip auth-exempt routes
    for prefix in AUTH_EXEMPT_PREFIXES:
        if path.startswith(prefix):
            return None

    # Extract and validate token
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return jsonify({"error": "Authentication required", "code": "NO_TOKEN"}), 401

    token = auth_header[7:]
    payload = decode_access_token(token)
    if payload is None:
        return jsonify({"error": "Invalid or expired token", "code": "INVALID_TOKEN"}), 401

    # Attach user info to Flask request context
    g.user_id = payload.get("user_id", int(payload["sub"]))
    g.nickname = payload["nickname"]
    g.is_admin = payload.get("is_admin", False)
    g.role = "ADMIN" if g.is_admin else None

    # RBAC: determine user's role on the current project
    # Extract production_id from URL if present
    import re
    prod_match = re.search(r'/api/productions/(\d+)', path)
    if prod_match and not g.is_admin:
        prod_id = int(prod_match.group(1))
        membership = get_membership(g.user_id, prod_id)
        if membership is None:
            return jsonify({"error": "You are not a member of this project", "code": "NOT_MEMBER"}), 403
        g.role = membership["role"]
    elif not g.is_admin:
        # For routes without prod_id (e.g., /api/boats/<id>), get role from
        # X-Project-Id header or default to their first membership
        project_header = request.headers.get("X-Project-Id")
        if project_header:
            try:
                membership = get_membership(g.user_id, int(project_header))
                if membership:
                    g.role = membership["role"]
            except (ValueError, TypeError):
                pass
        # If still no role, check if they have any membership at all
        if g.role is None:
            from auth.models import get_user_memberships
            memberships = get_user_memberships(g.user_id)
            if memberships:
                g.role = memberships[0]["role"]
            else:
                return jsonify({"error": "You are not assigned to any project", "code": "NO_PROJECT"}), 403

    # RBAC V2: load granular permissions for non-admin users
    if g.is_admin:
        g.permissions = None  # ADMIN has full access, no per-module check needed
        g.global_permissions = {"can_lock_unlock": True, "can_view_history": True}
    else:
        # Determine production_id for permission lookup
        _perm_prod_id = None
        import re as _re
        _pm = _re.search(r'/api/productions/(\d+)', path)
        if _pm:
            _perm_prod_id = int(_pm.group(1))
        else:
            _ph = request.headers.get("X-Project-Id")
            if _ph:
                try:
                    _perm_prod_id = int(_ph)
                except (ValueError, TypeError):
                    pass

        if _perm_prod_id and g.role:
            g.permissions = ensure_user_permissions(g.user_id, _perm_prod_id, g.role)
            g.global_permissions = get_user_global_permissions(g.user_id, _perm_prod_id)
        else:
            g.permissions = {}
            g.global_permissions = {"can_lock_unlock": False, "can_view_history": False}

    # P6.15: Entity-level restrictions flag
    # If user has entity permissions, routes can use g.has_entity_restrictions
    # and call get_user_allowed_entity_ids() to filter results
    if g.is_admin:
        g.has_entity_restrictions = False
    else:
        g.has_entity_restrictions = user_has_entity_restrictions(g.user_id)

    # RBAC: check access using V2 permissions (or V1 fallback for admin)
    if g.is_admin:
        pass  # Full access
    elif g.permissions:
        allowed, reason = check_permission_access(
            g.permissions, g.global_permissions, path, request.method, is_admin=False
        )
        if not allowed:
            return jsonify({"error": reason, "code": "FORBIDDEN"}), 403
    elif g.role:
        # Fallback to V1 role check if no permissions loaded
        allowed, reason = check_role_access(g.role, path, request.method)
        if not allowed:
            return jsonify({"error": reason, "code": "FORBIDDEN"}), 403

    return None  # Continue to the route handler


# ─── P6.14: Access logging middleware ─────────────────────────────────────────

_STATIC_EXTENSIONS = ('.css', '.js', '.png', '.jpg', '.jpeg', '.gif', '.ico', '.svg', '.woff', '.woff2', '.ttf', '.map')

@app.after_request
def log_access(response):
    """Log authenticated API requests for audit (P6.14)."""
    path = request.path
    # Skip static files
    if not path.startswith("/api/") or path.lower().endswith(_STATIC_EXTENSIONS):
        return response
    # Skip auth endpoints (login/refresh generate too much noise)
    for prefix in AUTH_EXEMPT_PREFIXES:
        if path.startswith(prefix):
            return response
    # Only log if user was authenticated
    user_id = getattr(g, 'user_id', None)
    if user_id is None:
        return response
    try:
        with get_db() as conn:
            conn.execute(
                "INSERT INTO access_logs (user_id, endpoint, method, status_code, ip_address, user_agent) VALUES (?, ?, ?, ?, ?, ?)",
                (user_id, path, request.method, response.status_code,
                 request.remote_addr, str(request.user_agent)[:500])
            )
    except Exception:
        pass  # Never break the response for logging failures
    return response


# ─── Helpers ──────────────────────────────────────────────────────────────────

def prod_or_404(prod_id):
    p = get_production(prod_id)
    if not p:
        abort(404, description=f"Production {prod_id} not found")
    return p


def _row_or_404(conn, table, id_):
    r = conn.execute(f"SELECT * FROM {table} WHERE id=?", (id_,)).fetchone()
    if not r:
        abort(404)
    return dict(r)


def _validate_dates_for_prod(prod_id, data):
    """Validate assignment dates against production date range."""
    start = data.get("start_date")
    end = data.get("end_date")
    if start or end:
        prod = get_production(prod_id)
        if prod:
            validate_assignment_dates(start, end, prod.get("start_date"), prod.get("end_date"))


def _get_prod_id_from_function(boat_function_id):
    """Get production_id from a boat_function_id."""
    if not boat_function_id:
        return None
    with get_db() as conn:
        bf = conn.execute("SELECT production_id FROM boat_functions WHERE id=?", (boat_function_id,)).fetchone()
    return bf["production_id"] if bf else None


# ─── UI ───────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/login")
def login_page():
    return render_template("login.html")


# ─── Health ───────────────────────────────────────────────────────────────────

@app.route("/api/health")
def health():
    with get_db() as conn:
        tables = _compat_get_table_names(conn)
    info = get_backend_info()
    return jsonify({"status": "ok", "tables": tables, "table_count": len(tables), "backend": info["backend"]})


# ─── Productions ──────────────────────────────────────────────────────────────

@app.route("/api/productions", methods=["GET"])
def api_productions():
    return jsonify(get_productions())


@app.route("/api/productions", methods=["POST"])
def api_create_production():
    data = request.json or {}
    if not data.get("name"):
        return jsonify({"error": "name required"}), 400
    prod_id = create_production(data)
    seed_departments(prod_id)
    return jsonify(get_production(prod_id)), 201


@app.route("/api/productions/<int:prod_id>", methods=["GET"])
def api_get_production(prod_id):
    return jsonify(prod_or_404(prod_id))


# ─── Departments ──────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/departments")
def api_departments(prod_id):
    prod_or_404(prod_id)
    return jsonify(get_departments(prod_id))


# ─── Shooting days ────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/shooting-days", methods=["GET"])
def api_shooting_days(prod_id):
    prod_or_404(prod_id)
    return jsonify_cached(get_shooting_days(prod_id))


@app.route("/api/productions/<int:prod_id>/shooting-days", methods=["POST"])
def api_create_shooting_day(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data["production_id"] = prod_id
    if not data.get("date"):
        return jsonify({"error": "date required"}), 400
    validate_shooting_day(data)
    day_id = create_shooting_day(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM shooting_days WHERE id=?", (day_id,)).fetchone()
    _notify_pdt_change(prod_id, 'create', data.get('date', ''))
    return jsonify(dict(row)), 201


@app.route("/api/productions/<int:prod_id>/shooting-days/<int:day_id>", methods=["GET"])
def api_get_shooting_day(prod_id, day_id):
    prod_or_404(prod_id)
    day = get_shooting_day(day_id)
    if not day:
        abort(404)
    return jsonify(day)


@app.route("/api/productions/<int:prod_id>/shooting-days/<int:day_id>", methods=["PUT"])
def api_update_shooting_day(prod_id, day_id):
    prod_or_404(prod_id)
    data = request.json or {}
    validate_shooting_day(data)
    update_shooting_day(day_id, data)
    day = get_shooting_day(day_id)
    if not day:
        abort(404)
    _notify_pdt_change(prod_id, 'update', data.get('date', day.get('date', '')))
    return jsonify(day)


@app.route("/api/productions/<int:prod_id>/shooting-days/<int:day_id>", methods=["DELETE"])
def api_delete_shooting_day(prod_id, day_id):
    prod_or_404(prod_id)
    day = get_shooting_day(day_id)
    delete_shooting_day(day_id)
    _notify_pdt_change(prod_id, 'delete', day.get('date', '') if day else '')
    return jsonify({"deleted": day_id})


# ─── Tides (P6.3) ────────────────────────────────────────────────────────────

@app.route("/api/tides", methods=["GET"])
def api_tides():
    """Return tide predictions for a location and date range."""
    from tides import get_tides
    lat = request.args.get("lat", type=float)
    lng = request.args.get("lng", type=float)
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    if not start or not end:
        return jsonify({"error": "start and end required"}), 400
    data = get_tides(lat, lng, start, end)
    return jsonify(data)


# ─── PDT Cascade (AXE 7.2) ──────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/shooting-days/<int:day_id>/cascade-preview", methods=["POST"])
def api_cascade_preview(prod_id, day_id):
    """Preview cascade effects when moving a shooting day to a new date."""
    prod_or_404(prod_id)
    data = request.json or {}
    old_date = data.get("old_date")
    new_date = data.get("new_date")
    if not old_date or not new_date:
        return jsonify({"error": "old_date and new_date required"}), 400
    if old_date == new_date:
        return jsonify({"assignments": [], "fuel_entries": [], "location_schedules": [],
                        "summary": {"assignments": 0, "fuel_entries": 0, "location_schedules": 0}})
    preview = cascade_preview(prod_id, day_id, old_date, new_date)
    return jsonify(preview)


@app.route("/api/productions/<int:prod_id>/shooting-days/<int:day_id>/cascade-apply", methods=["POST"])
def api_cascade_apply(prod_id, day_id):
    """Apply cascade: move all date-keyed data from old_date to new_date."""
    prod_or_404(prod_id)
    data = request.json or {}
    old_date = data.get("old_date")
    new_date = data.get("new_date")
    if not old_date or not new_date:
        return jsonify({"error": "old_date and new_date required"}), 400
    if old_date == new_date:
        return jsonify({"applied": {"assignments": 0, "fuel_entries": 0, "location_schedules": 0}})
    applied = cascade_apply(prod_id, day_id, old_date, new_date)
    return jsonify({"applied": applied})


# ─── Shooting day events ─────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/shooting-days/<int:day_id>/events", methods=["GET"])
def api_get_day_events(prod_id, day_id):
    prod_or_404(prod_id)
    return jsonify(get_events_for_day(day_id))


@app.route("/api/productions/<int:prod_id>/shooting-days/<int:day_id>/events", methods=["POST"])
def api_create_event(prod_id, day_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data["shooting_day_id"] = day_id
    if not data.get("event_type"):
        return jsonify({"error": "event_type required"}), 400
    event_id = create_event(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM shooting_day_events WHERE id=?", (event_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/events/<int:event_id>", methods=["PUT"])
def api_update_event(event_id):
    data = request.json or {}
    validate_entity_name(data, field="event_type")
    update_event(event_id, data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM shooting_day_events WHERE id=?", (event_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/events/<int:event_id>", methods=["DELETE"])
def api_delete_event(event_id):
    delete_event(event_id)
    return jsonify({"deleted": event_id})


# ─── Parse PDT PDF ────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/parse-pdt", methods=["POST"])
def api_parse_pdt(prod_id):
    prod_or_404(prod_id)
    from pdf_parser import parse_pdt_pdf

    # Check existing days
    existing = get_shooting_days(prod_id)
    if existing and not (request.json or {}).get("force"):
        return jsonify({
            "message": "Shooting days already exist. Pass force=true to overwrite.",
            "existing_count": len(existing),
        }), 409

    # Delete existing if force
    if existing:
        with get_db() as conn:
            conn.execute(
                "DELETE FROM shooting_days WHERE production_id=?", (prod_id,)
            )

    days = parse_pdt_pdf()
    created = []
    for d in days:
        events = d.pop("events", [])
        d["production_id"] = prod_id
        day_id = create_shooting_day(d)
        created.append(day_id)
        for ev in events:
            ev["shooting_day_id"] = day_id
            create_event(ev)

    return jsonify({
        "created": len(created),
        "message": f"{len(created)} shooting days imported from PDT PDF",
    }), 201


# ─── Upload PDT PDF (browser file picker + smart merge) ─────────────────────

@app.route("/api/productions/<int:prod_id>/upload-pdt", methods=["POST"])
def api_upload_pdt(prod_id):
    """
    Receive a PDF file upload, parse it, and smart-merge with existing days.
    Merge logic:
      - Match by day_number
      - If existing day has status='modifie' -> skip (preserve manual edits)
      - If existing day has other status -> update fields from PDF
      - If day doesn't exist -> create it
      - Days in DB but not in PDF -> keep them (no delete)
    """
    prod_or_404(prod_id)
    from pdf_parser import parse_pdt_pdf

    if 'pdf' not in request.files:
        return jsonify({"error": "No PDF file provided"}), 400

    pdf_file = request.files['pdf']
    if not pdf_file.filename:
        return jsonify({"error": "Empty filename"}), 400

    # Save to temp file, parse, then delete
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.pdf')
    try:
        pdf_file.save(tmp_path)
        parsed_days = parse_pdt_pdf(pdf_path=tmp_path)
    finally:
        os.close(tmp_fd)
        os.unlink(tmp_path)

    # Build lookup of existing days by day_number
    existing = get_shooting_days(prod_id)
    existing_by_dn = {}
    for d in existing:
        if d.get('day_number'):
            existing_by_dn[d['day_number']] = d

    stats = {"created": 0, "updated": 0, "skipped": 0}

    for parsed in parsed_days:
        events = parsed.pop("events", [])
        dn = parsed.get("day_number")
        if not dn:
            continue

        if dn in existing_by_dn:
            ex = existing_by_dn[dn]
            # Skip days manually edited by user
            if ex.get('status') == 'modifié':
                stats["skipped"] += 1
                continue
            # Update day fields from PDF
            update_shooting_day(ex['id'], parsed)
            # Replace events: delete old, insert new
            delete_events_for_day(ex['id'])
            for ev in events:
                ev["shooting_day_id"] = ex['id']
                create_event(ev)
            stats["updated"] += 1
        else:
            # Create new day
            parsed["production_id"] = prod_id
            day_id = create_shooting_day(parsed)
            for ev in events:
                ev["shooting_day_id"] = day_id
                create_event(ev)
            stats["created"] += 1

    return jsonify(stats), 200


# ─── Boats ────────────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/boats", methods=["GET"])
def api_boats(prod_id):
    prod_or_404(prod_id)
    include_deleted = request.args.get('include_deleted', '').lower() == 'true'
    return jsonify_cached(get_boats(prod_id, include_deleted=include_deleted))


@app.route("/api/productions/<int:prod_id>/boats", methods=["POST"])
def api_create_boat(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data["production_id"] = prod_id
    if not data.get("name"):
        return jsonify({"error": "name required"}), 400
    boat_id = create_boat(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM boats WHERE id=?", (boat_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/boats/<int:boat_id>", methods=["GET"])
def api_get_boat(boat_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM boats WHERE id=?", (boat_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/boats/<int:boat_id>", methods=["PUT"])
def api_update_boat(boat_id):
    data = request.json or {}
    validate_entity_name(data)
    validate_numeric_fields(data, ['daily_rate_estimate', 'daily_rate_actual', 'boat_nr'])
    # AXE 6.3: capture old data for price change logging
    with get_db() as conn:
        old = conn.execute("SELECT * FROM boats WHERE id=?", (boat_id,)).fetchone()
    if old:
        old = dict(old)
        prod_id = old.get('production_id')
    ok = update_boat(boat_id, data)
    if ok is False:
        return jsonify({"error": "This record was modified by another user. Please refresh."}), 409
    with get_db() as conn:
        row = conn.execute("SELECT * FROM boats WHERE id=?", (boat_id,)).fetchone()
    if row and old:
        _log_price_changes(prod_id, 'boat', boat_id, old.get('name', ''), old, data)
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/boats/<int:boat_id>", methods=["DELETE"])
def api_delete_boat(boat_id):
    delete_boat(boat_id)
    return jsonify({"deleted": boat_id})


@app.route("/api/boats/<int:boat_id>/duplicate", methods=["POST"])
def api_duplicate_boat(boat_id):
    """Duplicate a boat (all properties, name + ' (copy)', no assignments)."""
    with get_db() as conn:
        row = conn.execute("SELECT * FROM boats WHERE id=?", (boat_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    data = dict(row)
    del data["id"]
    data["name"] = data["name"] + " (copy)"
    data.pop("image_path", None)
    new_id = create_boat(data)
    with get_db() as conn:
        new_row = conn.execute("SELECT * FROM boats WHERE id=?", (new_id,)).fetchone()
    return jsonify(dict(new_row)), 201


@app.route("/api/boats/<int:boat_id>/upload-image", methods=["POST"])
def api_upload_boat_image(boat_id):
    import os
    f = request.files.get('image')
    if not f:
        return jsonify({"error": "No file provided"}), 400
    upload_dir = os.path.join(os.path.dirname(__file__), 'static', 'uploads', 'boats')
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(f.filename or '')[1].lower() or '.jpg'
    filename = f"boat_{boat_id}{ext}"
    filepath = os.path.join(upload_dir, filename)
    f.save(filepath)
    rel_path = f"static/uploads/boats/{filename}"
    update_boat(boat_id, {"image_path": rel_path})
    with get_db() as conn:
        row = conn.execute("SELECT * FROM boats WHERE id=?", (boat_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


# ─── Boat functions ───────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/boat-functions", methods=["GET"])
def api_boat_functions(prod_id):
    prod_or_404(prod_id)
    context = request.args.get('context')
    include_deleted = request.args.get('include_deleted', '').lower() == 'true'
    return jsonify_cached(get_boat_functions(prod_id, context=context, include_deleted=include_deleted))


@app.route("/api/productions/<int:prod_id>/boat-functions", methods=["POST"])
def api_create_boat_function(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data["production_id"] = prod_id
    if not data.get("name"):
        return jsonify({"error": "name required"}), 400
    func_id = create_boat_function(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM boat_functions WHERE id=?", (func_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/boat-functions/<int:func_id>", methods=["PUT"])
def api_update_boat_function(func_id):
    data = request.json or {}
    validate_entity_name(data)
    update_boat_function(func_id, data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM boat_functions WHERE id=?", (func_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/boat-functions/<int:func_id>", methods=["DELETE"])

def api_delete_boat_function(func_id):
    delete_boat_function(func_id)
    return jsonify({"deleted": func_id})


@app.route("/api/productions/<int:prod_id>/boat-functions/reorder", methods=["POST"])
def api_reorder_boat_functions(prod_id):
    """Batch update sort_order and optionally function_group for functions."""
    prod_or_404(prod_id)
    items = (request.json or {}).get("items", [])
    if not isinstance(items, list):
        return jsonify({"error": "items must be a list"}), 400
    for item in items:
        if not isinstance(item, dict) or not item.get("id"):
            return jsonify({"error": "Each item must have an id"}), 400
    with get_db() as conn:
        for item in items:
            fid = item.get("id")
            conn.execute(
                "UPDATE boat_functions SET sort_order=?, function_group=? WHERE id=? AND production_id=?",
                (item.get("sort_order", 0), item.get("function_group"), fid, prod_id),
            )
        conn.commit()
    return jsonify({"ok": True})


# ─── Boat assignments ─────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/assignments", methods=["GET"])
def api_assignments(prod_id):
    prod_or_404(prod_id)
    context = request.args.get('context')
    return jsonify_cached(get_boat_assignments(prod_id, context=context))


@app.route("/api/productions/<int:prod_id>/assignments", methods=["POST"])
def api_create_assignment(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    if not data.get("boat_function_id"):
        return jsonify({"error": "boat_function_id required"}), 400
    validate_assignment(data)
    _validate_dates_for_prod(prod_id, data)
    if data.get("boat_id") and data.get("start_date") and data.get("end_date"):
        validate_assignment_overlap('boat_assignments', 'boat_id', data['boat_id'],
                                    data['start_date'], data['end_date'])
    assignment_id = create_boat_assignment(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM boat_assignments WHERE id=?", (assignment_id,)).fetchone()
        bf = conn.execute("SELECT production_id FROM boat_functions WHERE id=?", (data['boat_function_id'],)).fetchone()
    if bf:
        _notify_assignment_change(bf['production_id'], 'create', 'boat assignment',
                                  data.get('boat_name_override', f'#{assignment_id}'))
    result = dict(row)
    # P2.3: cross-module vessel conflict warning
    if data.get("boat_id") and data.get("start_date") and data.get("end_date"):
        with get_db() as conn:
            boat = conn.execute("SELECT physical_vessel_id FROM boats WHERE id=?", (data["boat_id"],)).fetchone()
        if boat and boat["physical_vessel_id"]:
            warnings = check_vessel_cross_module_conflict(
                boat["physical_vessel_id"], data["start_date"], data["end_date"],
                exclude_module="boat_assignments", exclude_id=assignment_id)
            if warnings:
                result["_warnings"] = warnings
    return jsonify(result), 201


@app.route("/api/assignments/<int:assignment_id>", methods=["PUT"])
def api_update_assignment(assignment_id):
    data = request.json or {}
    validate_assignment(data)
    prod_id = _get_prod_id_from_function(data.get("boat_function_id"))
    if prod_id:
        _validate_dates_for_prod(prod_id, data)
    if data.get("boat_id") and data.get("start_date") and data.get("end_date"):
        validate_assignment_overlap('boat_assignments', 'boat_id', data['boat_id'],
                                    data['start_date'], data['end_date'], exclude_id=assignment_id)
    # AXE 6.3: capture old data for price change logging
    with get_db() as conn:
        old = conn.execute("SELECT ba.*, bf.production_id FROM boat_assignments ba LEFT JOIN boat_functions bf ON ba.boat_function_id=bf.id WHERE ba.id=?", (assignment_id,)).fetchone()
    old_d = dict(old) if old else {}
    # P5.6: validate override_reason when price_override changes
    err = _validate_price_override(data, old_d)
    if err:
        return err
    update_boat_assignment(assignment_id, data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM boat_assignments WHERE id=?", (assignment_id,)).fetchone()
    if old_d.get('production_id'):
        _log_price_changes(old_d['production_id'], 'boat_assignment', assignment_id,
                          old_d.get('boat_name_override', f'Assignment #{assignment_id}'), old_d, data)
        _notify_assignment_change(old_d['production_id'], 'update', 'boat assignment',
                                  old_d.get('boat_name_override', f'#{assignment_id}'))
        if data.get('assignment_status') == 'breakdown' and old_d.get('assignment_status') != 'breakdown':
            _notify_vessel_breakdown(old_d['production_id'],
                                     old_d.get('boat_name_override', f'Boat #{assignment_id}'), 'Boats')
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/assignments/<int:assignment_id>", methods=["DELETE"])
def api_delete_assignment(assignment_id):
    delete_boat_assignment(assignment_id)
    return jsonify({"deleted": assignment_id})


@app.route("/api/productions/<int:prod_id>/assignments/function/<int:func_id>", methods=["DELETE"])
def api_delete_assignment_by_function(prod_id, func_id):
    prod_or_404(prod_id)
    delete_boat_assignment_by_function(func_id)
    return jsonify({"deleted_for_function": func_id})


# ─── Picture Boats ────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/picture-boats", methods=["GET"])
def api_picture_boats(prod_id):
    prod_or_404(prod_id)
    include_deleted = request.args.get('include_deleted', '').lower() == 'true'
    return jsonify_cached(get_picture_boats(prod_id, include_deleted=include_deleted))


@app.route("/api/productions/<int:prod_id>/picture-boats", methods=["POST"])
def api_create_picture_boat(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data["production_id"] = prod_id
    if not data.get("name"):
        return jsonify({"error": "name required"}), 400
    pb_id = create_picture_boat(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM picture_boats WHERE id=?", (pb_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/picture-boats/<int:pb_id>", methods=["GET"])
def api_get_picture_boat(pb_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM picture_boats WHERE id=?", (pb_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/picture-boats/<int:pb_id>", methods=["PUT"])
def api_update_picture_boat(pb_id):
    data = request.json or {}
    validate_entity_name(data)
    validate_numeric_fields(data, ['daily_rate_estimate', 'daily_rate_actual', 'boat_nr'])
    ok = update_picture_boat(pb_id, data)
    if ok is False:
        return jsonify({"error": "This record was modified by another user. Please refresh."}), 409
    with get_db() as conn:
        row = conn.execute("SELECT * FROM picture_boats WHERE id=?", (pb_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/picture-boats/<int:pb_id>", methods=["DELETE"])
def api_delete_picture_boat(pb_id):
    delete_picture_boat(pb_id)
    return jsonify({"deleted": pb_id})


@app.route("/api/picture-boats/<int:pb_id>/duplicate", methods=["POST"])
def api_duplicate_picture_boat(pb_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM picture_boats WHERE id=?", (pb_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    data = dict(row)
    del data["id"]
    data["name"] = data["name"] + " (copy)"
    data.pop("image_path", None)
    new_id = create_picture_boat(data)
    with get_db() as conn:
        new_row = conn.execute("SELECT * FROM picture_boats WHERE id=?", (new_id,)).fetchone()
    return jsonify(dict(new_row)), 201


@app.route("/api/picture-boats/<int:pb_id>/upload-image", methods=["POST"])
def api_upload_picture_boat_image(pb_id):
    import os
    f = request.files.get('image')
    if not f:
        return jsonify({"error": "No file provided"}), 400
    upload_dir = os.path.join(os.path.dirname(__file__), 'static', 'uploads', 'picture-boats')
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(f.filename or '')[1].lower() or '.jpg'
    filename = f"pb_{pb_id}{ext}"
    filepath = os.path.join(upload_dir, filename)
    f.save(filepath)
    rel_path = f"static/uploads/picture-boats/{filename}"
    update_picture_boat(pb_id, {"image_path": rel_path})
    with get_db() as conn:
        row = conn.execute("SELECT * FROM picture_boats WHERE id=?", (pb_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


# ─── Picture Boat Assignments ─────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/picture-boat-assignments", methods=["GET"])
def api_picture_boat_assignments(prod_id):
    prod_or_404(prod_id)
    return jsonify_cached(get_picture_boat_assignments(prod_id))


@app.route("/api/productions/<int:prod_id>/picture-boat-assignments", methods=["POST"])
def api_create_picture_boat_assignment(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    if not data.get("boat_function_id"):
        return jsonify({"error": "boat_function_id required"}), 400
    validate_assignment(data)
    _validate_dates_for_prod(prod_id, data)
    if data.get("picture_boat_id") and data.get("start_date") and data.get("end_date"):
        validate_assignment_overlap('picture_boat_assignments', 'picture_boat_id', data['picture_boat_id'],
                                    data['start_date'], data['end_date'])
    assignment_id = create_picture_boat_assignment(data)
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM picture_boat_assignments WHERE id=?", (assignment_id,)
        ).fetchone()
    result = dict(row)
    # P2.3: cross-module vessel conflict warning
    if data.get("picture_boat_id") and data.get("start_date") and data.get("end_date"):
        with get_db() as conn:
            pb = conn.execute("SELECT physical_vessel_id FROM picture_boats WHERE id=?", (data["picture_boat_id"],)).fetchone()
        if pb and pb["physical_vessel_id"]:
            warnings = check_vessel_cross_module_conflict(
                pb["physical_vessel_id"], data["start_date"], data["end_date"],
                exclude_module="picture_boat_assignments", exclude_id=assignment_id)
            if warnings:
                result["_warnings"] = warnings
    return jsonify(result), 201


@app.route("/api/picture-boat-assignments/<int:assignment_id>", methods=["PUT"])
def api_update_picture_boat_assignment(assignment_id):
    data = request.json or {}
    validate_assignment(data)
    prod_id = _get_prod_id_from_function(data.get("boat_function_id"))
    if prod_id:
        _validate_dates_for_prod(prod_id, data)
    if data.get("picture_boat_id") and data.get("start_date") and data.get("end_date"):
        validate_assignment_overlap('picture_boat_assignments', 'picture_boat_id', data['picture_boat_id'],
                                    data['start_date'], data['end_date'], exclude_id=assignment_id)
    old_status = None
    with get_db() as conn:
        old_row = conn.execute("SELECT * FROM picture_boat_assignments WHERE id=?", (assignment_id,)).fetchone()
    old_d = dict(old_row) if old_row else {}
    old_status = old_d.get('assignment_status')
    old_name = old_d.get('boat_name_override', f'PB #{assignment_id}')
    # P5.6: validate override_reason when price_override changes
    err = _validate_price_override(data, old_d)
    if err:
        return err
    update_picture_boat_assignment(assignment_id, data)
    # P5.6: log price changes
    if prod_id:
        _log_price_changes(prod_id, 'picture_boat_assignment', assignment_id, old_name, old_d, data)
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM picture_boat_assignments WHERE id=?", (assignment_id,)
        ).fetchone()
    if data.get('assignment_status') == 'breakdown' and old_status != 'breakdown' and prod_id:
        _notify_vessel_breakdown(prod_id, old_name, 'Picture Boats')
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/picture-boat-assignments/<int:assignment_id>", methods=["DELETE"])
def api_delete_picture_boat_assignment(assignment_id):
    delete_picture_boat_assignment(assignment_id)
    return jsonify({"deleted": assignment_id})


@app.route("/api/productions/<int:prod_id>/picture-boat-assignments/function/<int:func_id>",
           methods=["DELETE"])
def api_delete_picture_boat_assignment_by_function(prod_id, func_id):
    prod_or_404(prod_id)
    delete_picture_boat_assignment_by_function(func_id)
    return jsonify({"deleted_for_function": func_id})


# ─── Helpers ──────────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/helpers", methods=["GET"])
def api_helpers(prod_id):
    prod_or_404(prod_id)
    include_deleted = request.args.get('include_deleted', '').lower() == 'true'
    return jsonify_cached(get_helpers(prod_id, include_deleted=include_deleted))


@app.route("/api/productions/<int:prod_id>/helpers", methods=["POST"])
def api_create_helper(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data["production_id"] = prod_id
    if not data.get("name"):
        return jsonify({"error": "name required"}), 400
    hid = create_helper(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM helpers WHERE id=?", (hid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/helpers/<int:helper_id>", methods=["GET"])
def api_get_helper(helper_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM helpers WHERE id=?", (helper_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/helpers/<int:helper_id>", methods=["PUT"])
def api_update_helper(helper_id):
    data = request.json or {}
    validate_entity_name(data)
    validate_numeric_fields(data, ['daily_rate_estimate', 'daily_rate_actual'])
    with get_db() as conn:
        old = conn.execute("SELECT * FROM helpers WHERE id=?", (helper_id,)).fetchone()
    old_d = dict(old) if old else {}
    ok = update_helper(helper_id, data)
    if ok is False:
        return jsonify({"error": "This record was modified by another user. Please refresh."}), 409
    with get_db() as conn:
        row = conn.execute("SELECT * FROM helpers WHERE id=?", (helper_id,)).fetchone()
    if old_d.get('production_id'):
        _log_price_changes(old_d['production_id'], 'helper', helper_id, old_d.get('name', ''), old_d, data)
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/helpers/<int:helper_id>", methods=["DELETE"])
def api_delete_helper(helper_id):
    delete_helper(helper_id)
    return jsonify({"deleted": helper_id})


@app.route("/api/helpers/<int:helper_id>/duplicate", methods=["POST"])
def api_duplicate_helper(helper_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM helpers WHERE id=?", (helper_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    data = dict(row)
    del data["id"]
    data["name"] = data["name"] + " (copy)"
    data.pop("image_path", None)
    new_id = create_helper(data)
    with get_db() as conn:
        new_row = conn.execute("SELECT * FROM helpers WHERE id=?", (new_id,)).fetchone()
    return jsonify(dict(new_row)), 201


@app.route("/api/productions/<int:prod_id>/helpers/bulk", methods=["POST"])
def api_bulk_create_helpers(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    try:
        count = int(data.get("count", 0))
    except (ValueError, TypeError):
        return jsonify({"error": "count must be an integer"}), 400
    prefix = data.get("prefix", "Helper")
    if count < 1 or count > 200:
        return jsonify({"error": "count must be 1-200"}), 400
    shared = {
        "production_id": prod_id,
        "role": data.get("role"),
        "group_name": data.get("group_name", "GENERAL"),
        "daily_rate_estimate": data.get("daily_rate_estimate", 45),
        "notes": data.get("notes"),
    }
    func_id = data.get("boat_function_id")
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    created = []
    for i in range(1, count + 1):
        rec = dict(shared)
        rec["name"] = f"{prefix} {i}"
        hid = create_helper(rec)
        if func_id and hid:
            assign_data = {"boat_function_id": func_id, "helper_id": hid,
                           "start_date": start_date, "end_date": end_date}
            create_helper_assignment(assign_data)
        created.append(hid)
    return jsonify({"created": len(created), "ids": created}), 201


@app.route("/api/productions/<int:prod_id>/helpers/import-csv", methods=["POST"])
def api_import_helpers_csv(prod_id):
    import csv, io
    prod_or_404(prod_id)
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400
    content = f.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    created = []
    for row in reader:
        name = (row.get("name") or "").strip()
        if not name:
            continue
        rec = {
            "production_id": prod_id,
            "name": name,
            "role": (row.get("role") or "").strip() or None,
            "group_name": (row.get("group") or row.get("group_name") or "GENERAL").strip(),
            "daily_rate_estimate": float(row.get("rate") or row.get("daily_rate_estimate") or 45),
            "notes": (row.get("notes") or "").strip() or None,
        }
        hid = create_helper(rec)
        created.append(hid)
    return jsonify({"created": len(created), "ids": created}), 201


@app.route("/api/helpers/<int:helper_id>/upload-image", methods=["POST"])
def api_upload_helper_image(helper_id):
    import os
    f = request.files.get('image')
    if not f:
        return jsonify({"error": "No file provided"}), 400
    upload_dir = os.path.join(os.path.dirname(__file__), 'static', 'uploads', 'labour')
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(f.filename or '')[1].lower() or '.jpg'
    filename = f"worker_{helper_id}{ext}"
    filepath = os.path.join(upload_dir, filename)
    f.save(filepath)
    rel_path = f"static/uploads/labour/{filename}"
    update_helper(helper_id, {"image_path": rel_path})
    with get_db() as conn:
        row = conn.execute("SELECT * FROM helpers WHERE id=?", (helper_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/productions/<int:prod_id>/helper-assignments", methods=["GET"])
def api_helper_assignments(prod_id):
    prod_or_404(prod_id)
    return jsonify_cached(get_helper_assignments(prod_id))


@app.route("/api/productions/<int:prod_id>/helper-assignments", methods=["POST"])
def api_create_helper_assignment(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    if not data.get("boat_function_id"):
        return jsonify({"error": "boat_function_id required"}), 400
    validate_assignment(data)
    _validate_dates_for_prod(prod_id, data)
    if data.get("helper_id") and data.get("start_date") and data.get("end_date"):
        validate_assignment_overlap('helper_assignments', 'helper_id', data['helper_id'],
                                    data['start_date'], data['end_date'])
    aid = create_helper_assignment(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM helper_assignments WHERE id=?", (aid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/helper-assignments/<int:assignment_id>", methods=["PUT"])
def api_update_helper_assignment(assignment_id):
    data = request.json or {}
    validate_assignment(data)
    prod_id = _get_prod_id_from_function(data.get("boat_function_id"))
    if prod_id:
        _validate_dates_for_prod(prod_id, data)
    if data.get("helper_id") and data.get("start_date") and data.get("end_date"):
        validate_assignment_overlap('helper_assignments', 'helper_id', data['helper_id'],
                                    data['start_date'], data['end_date'], exclude_id=assignment_id)
    # P5.6: validate override_reason when price_override changes
    with get_db() as conn:
        old_row = conn.execute("SELECT ha.*, bf.production_id FROM helper_assignments ha LEFT JOIN boat_functions bf ON ha.boat_function_id=bf.id WHERE ha.id=?", (assignment_id,)).fetchone()
    old_d = dict(old_row) if old_row else {}
    err = _validate_price_override(data, old_d)
    if err:
        return err
    update_helper_assignment(assignment_id, data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM helper_assignments WHERE id=?", (assignment_id,)).fetchone()
    if old_d.get('production_id'):
        _log_price_changes(old_d['production_id'], 'helper_assignment', assignment_id,
                          old_d.get('helper_name_override', f'Assignment #{assignment_id}'), old_d, data)
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/helper-assignments/<int:assignment_id>", methods=["DELETE"])
def api_delete_helper_assignment(assignment_id):
    delete_helper_assignment(assignment_id)
    return jsonify({"deleted": assignment_id})


@app.route("/api/productions/<int:prod_id>/helper-assignments/function/<int:func_id>",
           methods=["DELETE"])
def api_delete_helper_assignment_by_function(prod_id, func_id):
    prod_or_404(prod_id)
    delete_helper_assignment_by_function(func_id)
    return jsonify({"deleted_for_function": func_id})


@app.route("/api/productions/<int:prod_id>/helpers/schedules", methods=["GET"])
def api_helper_schedules(prod_id):
    prod_or_404(prod_id)
    return jsonify(get_helper_schedules(prod_id))


@app.route("/api/productions/<int:prod_id>/export/labour/csv")
def api_export_labour_csv(prod_id):
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    rows = [r for r in get_helper_assignments(prod_id) if r.get("working_days")]
    rows = _filter_assignments_by_date(rows, date_from, date_to)
    from collections import OrderedDict
    by_group = OrderedDict()
    for r in rows:
        g = r.get("function_group") or r.get("helper_group") or "GENERAL"
        if g not in by_group:
            by_group[g] = []
        by_group[g].append(r)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Group", "Function", "Worker", "Role", "Contact",
                     "Start", "End", "Working Days", "Rate/day",
                     "Total Estimate", "Total Actual"])
    grand_est = 0
    grand_act = 0
    for group_name, group_rows in by_group.items():
        group_est = 0
        group_act = 0
        for r in group_rows:
            est = r.get("amount_estimate") or 0
            act = r.get("amount_actual") or 0
            group_est += est
            group_act += act
            writer.writerow([
                group_name,
                r.get("function_name") or "",
                r.get("helper_name_override") or r.get("helper_name") or "",
                r.get("helper_role") or "",
                r.get("helper_contact") or "",
                r.get("start_date") or "", r.get("end_date") or "",
                r.get("working_days") or "",
                r.get("price_override") or r.get("helper_daily_rate_estimate") or "",
                est,
                act if act else "",
            ])
        writer.writerow(["", "", "", "", "", "", f"SUB-TOTAL {group_name}", "", "", group_est, group_act if group_act else ""])
        writer.writerow([])
        grand_est += group_est
        grand_act += group_act
    writer.writerow(["", "", "", "", "", "", "GRAND TOTAL", "", "", grand_est, grand_act if grand_act else ""])
    output.seek(0)
    prod_name = prod.get("name", "PRODUCTION")
    fname = _export_fname(prod_name, "LABOUR", date_from, date_to, "csv")
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# Keep old endpoint for backward compatibility
@app.route("/api/productions/<int:prod_id>/export/helpers/csv")
def api_export_helpers_csv(prod_id):
    return api_export_labour_csv(prod_id)


# ─── Security Boats ──────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/security-boats", methods=["GET"])
def api_security_boats(prod_id):
    prod_or_404(prod_id)
    include_deleted = request.args.get('include_deleted', '').lower() == 'true'
    return jsonify_cached(get_security_boats(prod_id, include_deleted=include_deleted))


@app.route("/api/productions/<int:prod_id>/security-boats", methods=["POST"])
def api_create_security_boat(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data["production_id"] = prod_id
    if not data.get("name"):
        return jsonify({"error": "name required"}), 400
    sb_id = create_security_boat(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM security_boats WHERE id=?", (sb_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/security-boats/<int:sb_id>", methods=["GET"])
def api_get_security_boat(sb_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM security_boats WHERE id=?", (sb_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/security-boats/<int:sb_id>", methods=["PUT"])
def api_update_security_boat(sb_id):
    data = request.json or {}
    validate_entity_name(data)
    validate_numeric_fields(data, ['daily_rate_estimate', 'daily_rate_actual', 'boat_nr'])
    ok = update_security_boat(sb_id, data)
    if ok is False:
        return jsonify({"error": "This record was modified by another user. Please refresh."}), 409
    with get_db() as conn:
        row = conn.execute("SELECT * FROM security_boats WHERE id=?", (sb_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/security-boats/<int:sb_id>", methods=["DELETE"])
def api_delete_security_boat(sb_id):
    delete_security_boat(sb_id)
    return jsonify({"deleted": sb_id})


@app.route("/api/security-boats/<int:sb_id>/duplicate", methods=["POST"])
def api_duplicate_security_boat(sb_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM security_boats WHERE id=?", (sb_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    data = dict(row)
    del data["id"]
    data["name"] = data["name"] + " (copy)"
    data.pop("image_path", None)
    new_id = create_security_boat(data)
    with get_db() as conn:
        new_row = conn.execute("SELECT * FROM security_boats WHERE id=?", (new_id,)).fetchone()
    return jsonify(dict(new_row)), 201


@app.route("/api/security-boats/<int:sb_id>/upload-image", methods=["POST"])
def api_upload_security_boat_image(sb_id):
    import os
    f = request.files.get('image')
    if not f:
        return jsonify({"error": "No file provided"}), 400
    upload_dir = os.path.join(os.path.dirname(__file__), 'static', 'uploads', 'security-boats')
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(f.filename or '')[1].lower() or '.jpg'
    filename = f"sb_{sb_id}{ext}"
    filepath = os.path.join(upload_dir, filename)
    f.save(filepath)
    rel_path = f"static/uploads/security-boats/{filename}"
    update_security_boat(sb_id, {"image_path": rel_path})
    with get_db() as conn:
        row = conn.execute("SELECT * FROM security_boats WHERE id=?", (sb_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/productions/<int:prod_id>/security-boat-assignments", methods=["GET"])
def api_security_boat_assignments(prod_id):
    prod_or_404(prod_id)
    return jsonify_cached(get_security_boat_assignments(prod_id))


@app.route("/api/productions/<int:prod_id>/security-boat-assignments", methods=["POST"])
def api_create_security_boat_assignment(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    if not data.get("boat_function_id"):
        return jsonify({"error": "boat_function_id required"}), 400
    validate_assignment(data)
    _validate_dates_for_prod(prod_id, data)
    if data.get("security_boat_id") and data.get("start_date") and data.get("end_date"):
        validate_assignment_overlap('security_boat_assignments', 'security_boat_id', data['security_boat_id'],
                                    data['start_date'], data['end_date'])
    aid = create_security_boat_assignment(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM security_boat_assignments WHERE id=?", (aid,)).fetchone()
    result = dict(row)
    # P2.3: cross-module vessel conflict warning
    if data.get("security_boat_id") and data.get("start_date") and data.get("end_date"):
        with get_db() as conn:
            sb = conn.execute("SELECT physical_vessel_id FROM security_boats WHERE id=?", (data["security_boat_id"],)).fetchone()
        if sb and sb["physical_vessel_id"]:
            warnings = check_vessel_cross_module_conflict(
                sb["physical_vessel_id"], data["start_date"], data["end_date"],
                exclude_module="security_boat_assignments", exclude_id=aid)
            if warnings:
                result["_warnings"] = warnings
    return jsonify(result), 201


@app.route("/api/security-boat-assignments/<int:assignment_id>", methods=["PUT"])
def api_update_security_boat_assignment(assignment_id):
    data = request.json or {}
    validate_assignment(data)
    prod_id = _get_prod_id_from_function(data.get("boat_function_id"))
    if prod_id:
        _validate_dates_for_prod(prod_id, data)
    if data.get("security_boat_id") and data.get("start_date") and data.get("end_date"):
        validate_assignment_overlap('security_boat_assignments', 'security_boat_id', data['security_boat_id'],
                                    data['start_date'], data['end_date'], exclude_id=assignment_id)
    old_status = None
    with get_db() as conn:
        old_row = conn.execute("SELECT * FROM security_boat_assignments WHERE id=?", (assignment_id,)).fetchone()
    old_d = dict(old_row) if old_row else {}
    old_status = old_d.get('assignment_status')
    old_name = old_d.get('boat_name_override', f'SB #{assignment_id}')
    # P5.6: validate override_reason when price_override changes
    err = _validate_price_override(data, old_d)
    if err:
        return err
    update_security_boat_assignment(assignment_id, data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM security_boat_assignments WHERE id=?", (assignment_id,)).fetchone()
    if prod_id:
        _log_price_changes(prod_id, 'security_boat_assignment', assignment_id, old_name, old_d, data)
    if data.get('assignment_status') == 'breakdown' and old_status != 'breakdown' and prod_id:
        _notify_vessel_breakdown(prod_id, old_name, 'Security Boats')
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/security-boat-assignments/<int:assignment_id>", methods=["DELETE"])
def api_delete_security_boat_assignment(assignment_id):
    delete_security_boat_assignment(assignment_id)
    return jsonify({"deleted": assignment_id})


@app.route("/api/productions/<int:prod_id>/security-boat-assignments/function/<int:func_id>",
           methods=["DELETE"])
def api_delete_security_boat_assignment_by_function(prod_id, func_id):
    prod_or_404(prod_id)
    delete_security_boat_assignment_by_function(func_id)
    return jsonify({"deleted_for_function": func_id})


@app.route("/api/productions/<int:prod_id>/export/security-boats/csv")
def api_export_security_boats_csv(prod_id):
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    rows = [r for r in get_security_boat_assignments(prod_id) if r.get("working_days")]
    rows = _filter_assignments_by_date(rows, date_from, date_to)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Function", "Group", "Boat", "Captain", "Vendor",
                     "Start", "End", "Working Days", "Rate/day (est.)",
                     "Total Estimate", "Total Actual"])
    for r in rows:
        writer.writerow([
            r.get("function_name") or "",
            r.get("function_group") or "",
            r.get("boat_name_override") or r.get("boat_name") or "",
            r.get("captain") or "",
            r.get("vendor") or "",
            r.get("start_date") or "", r.get("end_date") or "",
            r.get("working_days") or "",
            r.get("price_override") or r.get("boat_daily_rate_estimate") or "",
            r.get("amount_estimate") or "",
            r.get("amount_actual") or "",
        ])
    grand_est = sum(r.get("amount_estimate") or 0 for r in rows)
    grand_act = sum(r.get("amount_actual") or 0 for r in rows)
    writer.writerow([])
    writer.writerow(["", "", "", "", "", "", "GRAND TOTAL", "", "", grand_est, grand_act])
    output.seek(0)
    prod_name = prod.get("name", "PRODUCTION")
    fname = _export_fname(prod_name, "SECURITY-BOATS", date_from, date_to, "csv")
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@app.route("/api/productions/<int:prod_id>/export/security-boats/json")
def api_export_security_boats_json(prod_id):
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    assignments = [a for a in get_security_boat_assignments(prod_id) if a.get("working_days")]
    assignments = _filter_assignments_by_date(assignments, date_from, date_to)
    data = {
        "production": get_production(prod_id),
        "security_boats": get_security_boats(prod_id),
        "assignments": assignments,
        "date_range": {"from": date_from, "to": date_to},
    }
    prod_name = prod.get("name", "PRODUCTION")
    fname = _export_fname(prod_name, "SECURITY-BOATS", date_from, date_to, "json")
    return Response(
        json.dumps(data, indent=2, ensure_ascii=False),
        mimetype="application/json",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ─── Incidents (P6.7) ─────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/incidents", methods=["GET"])
def api_get_incidents(prod_id):
    prod_or_404(prod_id)
    status_filter = request.args.get('status')
    with get_db() as conn:
        if status_filter:
            rows = conn.execute(
                "SELECT * FROM incidents WHERE production_id=? AND status=? ORDER BY date DESC",
                (prod_id, status_filter)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM incidents WHERE production_id=? ORDER BY date DESC",
                (prod_id,)
            ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/productions/<int:prod_id>/incidents", methods=["POST"])
def api_create_incident(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    required = ['entity_type', 'entity_id', 'incident_type', 'date']
    for f in required:
        if not data.get(f):
            return jsonify({"error": f"{f} required"}), 400
    valid_types = ('engine_failure', 'accident', 'delay', 'weather', 'other')
    if data['incident_type'] not in valid_types:
        return jsonify({"error": f"incident_type must be one of {valid_types}"}), 400
    with get_db() as conn:
        cur = conn.execute(
            """INSERT INTO incidents (production_id, entity_type, entity_id, incident_type,
               date, description, status, schedule_impact)
               VALUES (?, ?, ?, ?, ?, ?, 'open', ?)""",
            (prod_id, data['entity_type'], data['entity_id'], data['incident_type'],
             data['date'], data.get('description', ''), data.get('schedule_impact', ''))
        )
        iid = cur.lastrowid
        row = conn.execute("SELECT * FROM incidents WHERE id=?", (iid,)).fetchone()
    # Auto-set breakdown on matching boat assignments for today
    _apply_breakdown_for_incident(prod_id, data['entity_type'], data['entity_id'], data['date'])
    return jsonify(dict(row)), 201


@app.route("/api/productions/<int:prod_id>/incidents/<int:iid>", methods=["PUT"])
def api_update_incident(prod_id, iid):
    prod_or_404(prod_id)
    data = request.json or {}
    with get_db() as conn:
        old = conn.execute("SELECT * FROM incidents WHERE id=? AND production_id=?", (iid, prod_id)).fetchone()
        if not old:
            return jsonify({"error": "not found"}), 404
        old_d = dict(old)
        sets, vals = [], []
        for col in ('entity_type', 'entity_id', 'incident_type', 'date', 'description',
                     'status', 'schedule_impact'):
            if col in data:
                sets.append(f"{col}=?")
                vals.append(data[col])
        if data.get('status') == 'resolved' and old_d['status'] != 'resolved':
            sets.append("resolved_at=CURRENT_TIMESTAMP")
        if not sets:
            return jsonify(dict(old_d))
        vals.append(iid)
        conn.execute(f"UPDATE incidents SET {', '.join(sets)} WHERE id=?", vals)
        row = conn.execute("SELECT * FROM incidents WHERE id=?", (iid,)).fetchone()
    # If resolved, clear breakdown on matching assignments
    if data.get('status') == 'resolved' and old_d['status'] != 'resolved':
        _clear_breakdown_for_incident(prod_id, old_d['entity_type'], old_d['entity_id'], old_d['date'])
    return jsonify(dict(row))


@app.route("/api/productions/<int:prod_id>/incidents/<int:iid>", methods=["DELETE"])
def api_delete_incident(prod_id, iid):
    prod_or_404(prod_id)
    with get_db() as conn:
        old = conn.execute("SELECT * FROM incidents WHERE id=? AND production_id=?", (iid, prod_id)).fetchone()
        if old:
            old_d = dict(old)
            conn.execute("DELETE FROM incidents WHERE id=?", (iid,))
            # Clear breakdown if the incident was open
            if old_d['status'] == 'open':
                _clear_breakdown_for_incident(prod_id, old_d['entity_type'], old_d['entity_id'], old_d['date'])
    return jsonify({"deleted": iid})


def _apply_breakdown_for_incident(prod_id, entity_type, entity_id, date):
    """Set assignment_status='breakdown' on matching boat assignments active on incident date."""
    table_map = {
        'boat': ('boat_assignments', 'boat_id', 'boat_functions'),
        'picture': ('picture_boat_assignments', 'picture_boat_id', 'boat_functions'),
        'security': ('security_boat_assignments', 'security_boat_id', 'boat_functions'),
    }
    info = table_map.get(entity_type)
    if not info:
        return
    tbl, id_col, func_tbl = info
    with get_db() as conn:
        conn.execute(
            f"""UPDATE {tbl} SET assignment_status='breakdown'
                WHERE {id_col}=? AND start_date<=? AND end_date>=?
                AND assignment_status != 'breakdown'
                AND boat_function_id IN (SELECT id FROM {func_tbl} WHERE production_id=?)""",
            (entity_id, date, date, prod_id)
        )


def _clear_breakdown_for_incident(prod_id, entity_type, entity_id, date):
    """Remove breakdown status when incident is resolved (revert to confirmed)."""
    table_map = {
        'boat': ('boat_assignments', 'boat_id', 'boat_functions'),
        'picture': ('picture_boat_assignments', 'picture_boat_id', 'boat_functions'),
        'security': ('security_boat_assignments', 'security_boat_id', 'boat_functions'),
    }
    info = table_map.get(entity_type)
    if not info:
        return
    tbl, id_col, func_tbl = info
    # Only clear if no other open incident exists for same entity on same date
    with get_db() as conn:
        other = conn.execute(
            "SELECT COUNT(*) as cnt FROM incidents WHERE production_id=? AND entity_type=? AND entity_id=? AND date=? AND status='open'",
            (prod_id, entity_type, entity_id, date)
        ).fetchone()
        if other and other['cnt'] > 0:
            return
        conn.execute(
            f"""UPDATE {tbl} SET assignment_status='confirmed'
                WHERE {id_col}=? AND start_date<=? AND end_date>=?
                AND assignment_status='breakdown'
                AND boat_function_id IN (SELECT id FROM {func_tbl} WHERE production_id=?)""",
            (entity_id, date, date, prod_id)
        )


# ─── Physical Vessels (P2.3) ──────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/physical-vessels", methods=["GET"])
def api_physical_vessels(prod_id):
    prod_or_404(prod_id)
    include_deleted = request.args.get('include_deleted', '').lower() == 'true'
    return jsonify(get_physical_vessels(prod_id, include_deleted=include_deleted))


@app.route("/api/productions/<int:prod_id>/physical-vessels", methods=["POST"])
def api_create_physical_vessel(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data["production_id"] = prod_id
    if not data.get("vessel_name"):
        return jsonify({"error": "vessel_name required"}), 400
    vessel_id = create_physical_vessel(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM physical_vessels WHERE id=?", (vessel_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/physical-vessels/<int:vessel_id>", methods=["PUT"])
def api_update_physical_vessel(vessel_id):
    data = request.json or {}
    validate_entity_name(data, field="vessel_name")
    update_physical_vessel(vessel_id, data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM physical_vessels WHERE id=?", (vessel_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/physical-vessels/<int:vessel_id>", methods=["DELETE"])
def api_delete_physical_vessel(vessel_id):
    delete_physical_vessel(vessel_id)
    return jsonify({"deleted": vessel_id})


# ─── FNB ──────────────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/fnb", methods=["GET"])
def api_fnb(prod_id):
    prod_or_404(prod_id)
    return jsonify(get_fnb_services(prod_id))


# ─── Fuel ─────────────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/fuel", methods=["GET"])
def api_fuel(prod_id):
    prod_or_404(prod_id)
    return jsonify(get_fuel_logs(prod_id))


# ─── Transport ────────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/transport", methods=["GET"])
def api_transport(prod_id):
    prod_or_404(prod_id)
    return jsonify(get_transport_schedules(prod_id))


# ─── Guards ───────────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/guards", methods=["GET"])
def api_guards(prod_id):
    prod_or_404(prod_id)
    return jsonify(get_guard_schedules(prod_id))


# ─── Budget ───────────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/budget", methods=["GET"])
def api_budget(prod_id):
    prod_or_404(prod_id)
    ref_cur = request.args.get("currency", "USD").upper()
    return jsonify(get_budget(prod_id, ref_currency=ref_cur))


# ─── Exchange Rates (P6.4) ───────────────────────────────────────────────────

@app.route("/api/exchange-rates", methods=["GET"])
def api_exchange_rates_list():
    return jsonify(get_exchange_rates())


@app.route("/api/exchange-rates", methods=["POST"])
def api_exchange_rates_create():
    d = request.get_json(force=True)
    date = d.get("date")
    from_cur = d.get("from_currency", d.get("from", ""))
    to_cur = d.get("to_currency", d.get("to", ""))
    rate = d.get("rate")
    if not date or not from_cur or not to_cur or rate is None:
        return jsonify({"error": "date, from_currency, to_currency, rate required"}), 400
    try:
        rate = float(rate)
    except (ValueError, TypeError):
        return jsonify({"error": "rate must be a number"}), 400
    if rate <= 0:
        return jsonify({"error": "rate must be positive"}), 400
    result = upsert_exchange_rate(date, from_cur, to_cur, rate)
    return jsonify(result), 201


# ─── History / Undo ───────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/activity", methods=["GET"])
def api_activity(prod_id):
    """Enriched activity feed with date grouping, module mapping, and change details."""
    prod_or_404(prod_id)
    limit = int(request.args.get("limit", 100))
    module = request.args.get("module")
    user_id = request.args.get("user_id")
    date = request.args.get("date")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    return jsonify(get_activity_feed(
        prod_id, limit=limit, module=module, user_id=user_id,
        date=date, date_from=date_from, date_to=date_to
    ))


@app.route("/api/productions/<int:prod_id>/history", methods=["GET"])
def api_history(prod_id):
    prod_or_404(prod_id)
    limit = int(request.args.get("limit", 50))
    entity_type = request.args.get("entity_type")
    entity_id = request.args.get("entity_id")
    user_id = request.args.get("user_id")
    action_type = request.args.get("action_type")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    if entity_id:
        entity_id = int(entity_id)
    if user_id:
        user_id = int(user_id)
    return jsonify(get_history(
        prod_id, limit,
        entity_type=entity_type, entity_id=entity_id,
        user_id=user_id, action_type=action_type,
        date_from=date_from, date_to=date_to
    ))


@app.route("/api/productions/<int:prod_id>/undo", methods=["POST"])
def api_undo(prod_id):
    prod_or_404(prod_id)
    return jsonify(undo_last_boat_assignment(prod_id))


@app.route("/api/history/<int:history_id>/undo", methods=["POST"])
def api_undo_history(history_id):
    return jsonify(undo_history_entry(history_id))


@app.route("/api/productions/<int:prod_id>/history/<int:history_id>/undo", methods=["POST"])
def api_undo_history_scoped(prod_id, history_id):
    """Production-scoped undo endpoint."""
    prod_or_404(prod_id)
    return jsonify(undo_history_entry(history_id))


@app.route("/api/productions/<int:prod_id>/history/undoable", methods=["GET"])
def api_undoable_history(prod_id):
    """Return the last 50 undoable history entries for the production."""
    prod_or_404(prod_id)
    limit = int(request.args.get("limit", 50))
    since = request.args.get("since")  # ISO timestamp for session filtering
    from database import get_db
    with get_db() as conn:
        query = """SELECT id, table_name, record_id, action, human_description,
                          user_nickname, created_at, undone_at
                   FROM history
                   WHERE (production_id = ? OR production_id IS NULL)
                     AND action IN ('create', 'update', 'delete')
                     AND undone_at IS NULL"""
        params = [prod_id]
        if since:
            query += " AND created_at >= ?"
            params.append(since)
        query += " ORDER BY id DESC LIMIT ?"
        params.append(limit)
        rows = conn.execute(query, params).fetchall()
        return jsonify([dict(r) for r in rows])


# ─── Delete Impact (cascade preview) ─────────────────────────────────────────

@app.route("/api/boats/<int:boat_id>/impact")
def api_boat_impact(boat_id):
    with get_db() as conn:
        assignments = conn.execute(
            "SELECT COUNT(*) AS c FROM boat_assignments WHERE boat_id=? AND deleted_at IS NULL", (boat_id,)
        ).fetchone()["c"]
        fuel = conn.execute(
            "SELECT COUNT(*) AS c FROM fuel_entries WHERE source_type='boats' AND source_id=?", (boat_id,)
        ).fetchone()["c"]
    return jsonify({"assignments": assignments, "fuel_entries": fuel})


@app.route("/api/picture-boats/<int:pb_id>/impact")
def api_picture_boat_impact(pb_id):
    with get_db() as conn:
        assignments = conn.execute(
            "SELECT COUNT(*) AS c FROM picture_boat_assignments WHERE picture_boat_id=? AND deleted_at IS NULL", (pb_id,)
        ).fetchone()["c"]
        fuel = conn.execute(
            "SELECT COUNT(*) AS c FROM fuel_entries WHERE source_type='picture_boats' AND source_id=?", (pb_id,)
        ).fetchone()["c"]
    return jsonify({"assignments": assignments, "fuel_entries": fuel})


@app.route("/api/security-boats/<int:sb_id>/impact")
def api_security_boat_impact(sb_id):
    with get_db() as conn:
        assignments = conn.execute(
            "SELECT COUNT(*) AS c FROM security_boat_assignments WHERE security_boat_id=? AND deleted_at IS NULL", (sb_id,)
        ).fetchone()["c"]
        fuel = conn.execute(
            "SELECT COUNT(*) AS c FROM fuel_entries WHERE source_type='security_boats' AND source_id=?", (sb_id,)
        ).fetchone()["c"]
    return jsonify({"assignments": assignments, "fuel_entries": fuel})


@app.route("/api/transport-vehicles/<int:vehicle_id>/impact")
def api_transport_vehicle_impact(vehicle_id):
    with get_db() as conn:
        assignments = conn.execute(
            "SELECT COUNT(*) AS c FROM transport_assignments WHERE vehicle_id=? AND deleted_at IS NULL", (vehicle_id,)
        ).fetchone()["c"]
        fuel = conn.execute(
            "SELECT COUNT(*) AS c FROM fuel_entries WHERE source_type='transport' AND source_id=?", (vehicle_id,)
        ).fetchone()["c"]
    return jsonify({"assignments": assignments, "fuel_entries": fuel})


@app.route("/api/helpers/<int:helper_id>/impact")
def api_helper_impact(helper_id):
    with get_db() as conn:
        assignments = conn.execute(
            "SELECT COUNT(*) AS c FROM helper_assignments WHERE helper_id=? AND deleted_at IS NULL", (helper_id,)
        ).fetchone()["c"]
    return jsonify({"assignments": assignments})


@app.route("/api/guard-camp-workers/<int:worker_id>/impact")
def api_guard_camp_worker_impact(worker_id):
    with get_db() as conn:
        assignments = conn.execute(
            "SELECT COUNT(*) AS c FROM guard_camp_assignments WHERE helper_id=? AND deleted_at IS NULL", (worker_id,)
        ).fetchone()["c"]
    return jsonify({"assignments": assignments})


@app.route("/api/locations/<int:loc_id>/impact")
def api_location_impact(loc_id):
    with get_db() as conn:
        loc = conn.execute("SELECT * FROM locations WHERE id=?", (loc_id,)).fetchone()
        schedules = 0
        guard_schedules = 0
        if loc:
            schedules = conn.execute(
                "SELECT COUNT(*) AS c FROM location_schedules WHERE location_id=?", (loc_id,)
            ).fetchone()["c"]
            guard_schedules = conn.execute(
                "SELECT COUNT(*) AS c FROM guard_location_schedules WHERE location_name=? AND production_id=?",
                (loc["name"], loc["production_id"])
            ).fetchone()["c"]
    return jsonify({"schedules": schedules, "guard_schedules": guard_schedules})


@app.route("/api/guard-posts/<int:post_id>/impact")
def api_guard_post_impact(post_id):
    with get_db() as conn:
        post = conn.execute("SELECT * FROM guard_posts WHERE id=?", (post_id,)).fetchone()
        guard_schedules = 0
        if post:
            guard_schedules = conn.execute(
                "SELECT COUNT(*) AS c FROM guard_location_schedules WHERE location_name=? AND production_id=?",
                (post["name"], post["production_id"])
            ).fetchone()["c"]
    return jsonify({"guard_schedules": guard_schedules})


@app.route("/api/fuel-machinery/<int:machinery_id>/impact")
def api_fuel_machinery_impact(machinery_id):
    with get_db() as conn:
        fuel = conn.execute(
            "SELECT COUNT(*) AS c FROM fuel_entries WHERE source_type='machinery' AND source_id=?", (machinery_id,)
        ).fetchone()["c"]
    return jsonify({"fuel_entries": fuel})


@app.route("/api/fnb-categories/<int:cat_id>/impact")
def api_fnb_category_impact(cat_id):
    with get_db() as conn:
        items = conn.execute(
            "SELECT COUNT(*) AS c FROM fnb_items WHERE category_id=? AND deleted_at IS NULL", (cat_id,)
        ).fetchone()["c"]
        entries = conn.execute(
            "SELECT COUNT(*) AS c FROM fnb_entries WHERE item_id IN (SELECT id FROM fnb_items WHERE category_id=?)", (cat_id,)
        ).fetchone()["c"]
    return jsonify({"items": items, "entries": entries})


@app.route("/api/fnb-items/<int:item_id>/impact")
def api_fnb_item_impact(item_id):
    with get_db() as conn:
        entries = conn.execute(
            "SELECT COUNT(*) AS c FROM fnb_entries WHERE item_id=?", (item_id,)
        ).fetchone()["c"]
    return jsonify({"entries": entries})


# ─── Working days util ────────────────────────────────────────────────────────

@app.route("/api/working-days")
def api_working_days():
    start = request.args.get("start")
    end = request.args.get("end")
    return jsonify({"working_days": working_days(start, end)})


# ─── Export ───────────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/export/csv")
def api_export_csv(prod_id):
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    budget = get_budget(prod_id)
    rows = _filter_assignments_by_date(budget["rows"], date_from, date_to)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Department", "Function", "Boat", "Vendor", "Start", "End",
                     "Working Days", "Rate/day", "Total Estimate", "Total Actual"])
    for r in rows:
        writer.writerow([
            r.get("department") or "BOATS",
            r.get("name") or "",
            r.get("boat") or r.get("boat_name") or r.get("boat_name_override") or "",
            r.get("vendor") or "",
            r.get("start_date") or "", r.get("end_date") or "",
            r.get("working_days") or "",
            r.get("unit_price_estimate") or "",
            r.get("amount_estimate") or "", r.get("amount_actual") or "",
        ])
    grand_est = sum(r.get("amount_estimate") or 0 for r in rows)
    grand_act = sum(r.get("amount_actual") or 0 for r in rows)
    writer.writerow([])
    writer.writerow(["", "", "", "", "", "", "GRAND TOTAL", "", grand_est, grand_act])
    output.seek(0)
    prod_name = prod.get("name", "PRODUCTION")
    fname = _export_fname(prod_name, "BOATS", date_from, date_to, "csv")
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@app.route("/api/productions/<int:prod_id>/export/json")
def api_export_json(prod_id):
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    assignments = [a for a in get_boat_assignments(prod_id, context='boats') if a.get("working_days")]
    assignments = _filter_assignments_by_date(assignments, date_from, date_to)
    data = {
        "production": get_production(prod_id),
        "shooting_days": get_shooting_days(prod_id),
        "boats": get_boats(prod_id),
        "boat_functions": get_boat_functions(prod_id),
        "assignments": assignments,
        "budget": get_budget(prod_id),
        "date_range": {"from": date_from, "to": date_to},
    }
    prod_name = prod.get("name", "PRODUCTION")
    fname = _export_fname(prod_name, "BOATS", date_from, date_to, "json")
    return Response(
        json.dumps(data, indent=2, ensure_ascii=False),
        mimetype="application/json",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ─── Picture Boats export ─────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/export/picture-boats/csv")
def api_export_pb_csv(prod_id):
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    rows = [r for r in get_picture_boat_assignments(prod_id) if r.get("working_days")]
    rows = _filter_assignments_by_date(rows, date_from, date_to)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Function", "Group", "Boat", "Captain", "Vendor",
                     "Start", "End", "Working Days", "Rate/day (est.)",
                     "Total Estimate", "Total Actual"])
    for r in rows:
        writer.writerow([
            r.get("function_name") or "",
            r.get("function_group") or "",
            r.get("boat_name_override") or r.get("boat_name") or "",
            r.get("captain") or "",
            r.get("vendor") or "",
            r.get("start_date") or "", r.get("end_date") or "",
            r.get("working_days") or "",
            r.get("price_override") or r.get("boat_daily_rate_estimate") or "",
            r.get("amount_estimate") or "",
            r.get("amount_actual") or "",
        ])
    grand_est = sum(r.get("amount_estimate") or 0 for r in rows)
    grand_act = sum(r.get("amount_actual") or 0 for r in rows)
    writer.writerow([])
    writer.writerow(["", "", "", "", "", "", "GRAND TOTAL", "", "", grand_est, grand_act])
    output.seek(0)
    prod_name = prod.get("name", "PRODUCTION")
    fname = _export_fname(prod_name, "PICTURE-BOATS", date_from, date_to, "csv")
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@app.route("/api/productions/<int:prod_id>/export/picture-boats/json")
def api_export_pb_json(prod_id):
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    assignments = [a for a in get_picture_boat_assignments(prod_id) if a.get("working_days")]
    assignments = _filter_assignments_by_date(assignments, date_from, date_to)
    data = {
        "production": get_production(prod_id),
        "picture_boats": get_picture_boats(prod_id),
        "assignments": assignments,
        "date_range": {"from": date_from, "to": date_to},
    }
    prod_name = prod.get("name", "PRODUCTION")
    fname = _export_fname(prod_name, "PICTURE-BOATS", date_from, date_to, "json")
    return Response(
        json.dumps(data, indent=2, ensure_ascii=False),
        mimetype="application/json",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ─── Transport ────────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/transport-vehicles", methods=["GET"])
def api_transport_vehicles(prod_id):
    prod_or_404(prod_id)
    include_deleted = request.args.get('include_deleted', '').lower() == 'true'
    return jsonify_cached(get_transport_vehicles(prod_id, include_deleted=include_deleted))


@app.route("/api/productions/<int:prod_id>/transport-vehicles", methods=["POST"])
def api_create_transport_vehicle(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data["production_id"] = prod_id
    if not data.get("name"):
        return jsonify({"error": "name required"}), 400
    vid = create_transport_vehicle(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM transport_vehicles WHERE id=?", (vid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/transport-vehicles/<int:vehicle_id>", methods=["GET"])
def api_get_transport_vehicle(vehicle_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM transport_vehicles WHERE id=?", (vehicle_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/transport-vehicles/<int:vehicle_id>", methods=["PUT"])
def api_update_transport_vehicle(vehicle_id):
    data = request.json or {}
    validate_entity_name(data)
    validate_numeric_fields(data, ['daily_rate_estimate', 'daily_rate_actual'])
    with get_db() as conn:
        old = conn.execute("SELECT * FROM transport_vehicles WHERE id=?", (vehicle_id,)).fetchone()
    old_d = dict(old) if old else {}
    ok = update_transport_vehicle(vehicle_id, data)
    if ok is False:
        return jsonify({"error": "This record was modified by another user. Please refresh."}), 409
    with get_db() as conn:
        row = conn.execute("SELECT * FROM transport_vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if old_d.get('production_id'):
        _log_price_changes(old_d['production_id'], 'vehicle', vehicle_id, old_d.get('name', ''), old_d, data)
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/transport-vehicles/<int:vehicle_id>", methods=["DELETE"])
def api_delete_transport_vehicle(vehicle_id):
    delete_transport_vehicle(vehicle_id)
    return jsonify({"deleted": vehicle_id})


@app.route("/api/transport-vehicles/<int:vehicle_id>/duplicate", methods=["POST"])
def api_duplicate_transport_vehicle(vehicle_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM transport_vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    data = dict(row)
    del data["id"]
    data["name"] = data["name"] + " (copy)"
    data.pop("image_path", None)
    new_id = create_transport_vehicle(data)
    with get_db() as conn:
        new_row = conn.execute("SELECT * FROM transport_vehicles WHERE id=?", (new_id,)).fetchone()
    return jsonify(dict(new_row)), 201


@app.route("/api/transport-vehicles/<int:vehicle_id>/upload-image", methods=["POST"])
def api_upload_transport_vehicle_image(vehicle_id):
    import os
    f = request.files.get('image')
    if not f:
        return jsonify({"error": "No file provided"}), 400
    upload_dir = os.path.join(os.path.dirname(__file__), 'static', 'uploads', 'transport')
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(f.filename or '')[1].lower() or '.jpg'
    filename = f"tv_{vehicle_id}{ext}"
    filepath = os.path.join(upload_dir, filename)
    f.save(filepath)
    rel_path = f"static/uploads/transport/{filename}"
    update_transport_vehicle(vehicle_id, {"image_path": rel_path})
    with get_db() as conn:
        row = conn.execute("SELECT * FROM transport_vehicles WHERE id=?", (vehicle_id,)).fetchone()
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/productions/<int:prod_id>/transport-assignments", methods=["GET"])
def api_transport_assignments(prod_id):
    prod_or_404(prod_id)
    return jsonify_cached(get_transport_assignments(prod_id))


@app.route("/api/productions/<int:prod_id>/transport-assignments", methods=["POST"])
def api_create_transport_assignment(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    if not data.get("boat_function_id"):
        return jsonify({"error": "boat_function_id required"}), 400
    validate_assignment(data)
    _validate_dates_for_prod(prod_id, data)
    if data.get("vehicle_id") and data.get("start_date") and data.get("end_date"):
        validate_assignment_overlap('transport_assignments', 'vehicle_id', data['vehicle_id'],
                                    data['start_date'], data['end_date'])
    aid = create_transport_assignment(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM transport_assignments WHERE id=?", (aid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/transport-assignments/<int:assignment_id>", methods=["PUT"])
def api_update_transport_assignment(assignment_id):
    data = request.json or {}
    validate_assignment(data)
    prod_id = _get_prod_id_from_function(data.get("boat_function_id"))
    if prod_id:
        _validate_dates_for_prod(prod_id, data)
    if data.get("vehicle_id") and data.get("start_date") and data.get("end_date"):
        validate_assignment_overlap('transport_assignments', 'vehicle_id', data['vehicle_id'],
                                    data['start_date'], data['end_date'], exclude_id=assignment_id)
    # P5.6: validate override_reason when price_override changes
    with get_db() as conn:
        old_row = conn.execute("SELECT ta.*, bf.production_id FROM transport_assignments ta LEFT JOIN boat_functions bf ON ta.boat_function_id=bf.id WHERE ta.id=?", (assignment_id,)).fetchone()
    old_d = dict(old_row) if old_row else {}
    err = _validate_price_override(data, old_d)
    if err:
        return err
    update_transport_assignment(assignment_id, data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM transport_assignments WHERE id=?", (assignment_id,)).fetchone()
    if old_d.get('production_id'):
        _log_price_changes(old_d['production_id'], 'transport_assignment', assignment_id,
                          old_d.get('vehicle_name_override', f'Assignment #{assignment_id}'), old_d, data)
    return jsonify(dict(row)) if row else ("", 404)


@app.route("/api/transport-assignments/<int:assignment_id>", methods=["DELETE"])
def api_delete_transport_assignment(assignment_id):
    delete_transport_assignment(assignment_id)
    return jsonify({"deleted": assignment_id})


@app.route("/api/productions/<int:prod_id>/transport-assignments/function/<int:func_id>",
           methods=["DELETE"])
def api_delete_transport_assignment_by_function(prod_id, func_id):
    prod_or_404(prod_id)
    delete_transport_assignment_by_function(func_id)
    return jsonify({"deleted_for_function": func_id})


@app.route("/api/productions/<int:prod_id>/export/transport/csv")
def api_export_transport_csv(prod_id):
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    rows = [r for r in get_transport_assignments(prod_id) if r.get("working_days")]
    rows = _filter_assignments_by_date(rows, date_from, date_to)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Function", "Group", "Vehicle", "Type", "Driver", "Vendor",
                     "Start", "End", "Working Days", "Rate/day (est.)",
                     "Total Estimate", "Total Actual"])
    for r in rows:
        writer.writerow([
            r.get("function_name") or "",
            r.get("function_group") or "",
            r.get("vehicle_name_override") or r.get("vehicle_name") or "",
            r.get("vehicle_type") or "",
            r.get("driver") or "",
            r.get("vendor") or "",
            r.get("start_date") or "", r.get("end_date") or "",
            r.get("working_days") or "",
            r.get("price_override") or r.get("vehicle_daily_rate_estimate") or "",
            r.get("amount_estimate") or "",
            r.get("amount_actual") or "",
        ])
    grand_est = sum(r.get("amount_estimate") or 0 for r in rows)
    grand_act = sum(r.get("amount_actual") or 0 for r in rows)
    writer.writerow([])
    writer.writerow(["", "", "", "", "", "", "GRAND TOTAL", "", "", "", grand_est, grand_act])
    output.seek(0)
    prod_name = prod.get("name", "PRODUCTION")
    fname = _export_fname(prod_name, "TRANSPORT", date_from, date_to, "csv")
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@app.route("/api/productions/<int:prod_id>/export/transport/json")
def api_export_transport_json(prod_id):
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    assignments = [a for a in get_transport_assignments(prod_id) if a.get("working_days")]
    assignments = _filter_assignments_by_date(assignments, date_from, date_to)
    data = {
        "production": get_production(prod_id),
        "transport_vehicles": get_transport_vehicles(prod_id),
        "assignments": assignments,
        "date_range": {"from": date_from, "to": date_to},
    }
    prod_name = prod.get("name", "PRODUCTION")
    fname = _export_fname(prod_name, "TRANSPORT", date_from, date_to, "json")
    return Response(
        json.dumps(data, indent=2, ensure_ascii=False),
        mimetype="application/json",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ─── Fuel overview ────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/fuel/overview", methods=["GET"])
def api_fuel_overview(prod_id):
    prod_or_404(prod_id)
    from database import get_db
    with get_db() as conn:
        # All fuel entries for this production
        entries = conn.execute(
            "SELECT fe.source_type, fe.assignment_id, fe.date, fe.liters, fe.fuel_type "
            "FROM fuel_entries fe WHERE fe.production_id=? ORDER BY fe.date, fe.source_type",
            (prod_id,)
        ).fetchall()

        # Get source names from assignments / machinery
        boat_names = {}
        for tbl, ctx in [('boat_assignments','boats'), ('picture_boat_assignments','picture_boats'),
                         ('security_boat_assignments','security_boats'), ('transport_assignments','transport')]:
            name_col = 'vehicle_name_override' if ctx == 'transport' else 'boat_name_override'
            fallback_col = 'vehicle_name' if ctx == 'transport' else 'boat_name'
            try:
                rows = conn.execute(
                    f"SELECT a.id, COALESCE(a.{name_col}, a.{fallback_col}, '?') as name "
                    f"FROM {tbl} a WHERE a.production_id=?", (prod_id,)
                ).fetchall()
                for r in rows:
                    boat_names[(ctx, r['id'])] = r['name']
            except Exception:
                pass

        mach_rows = conn.execute(
            "SELECT id, name FROM fuel_machinery WHERE production_id=?", (prod_id,)
        ).fetchall()
        for r in mach_rows:
            boat_names[('machinery', r['id'])] = r['name']

    # Build overview data
    dates_set = set()
    sources = {}  # key = "source_type:assignment_id"
    for e in entries:
        dk = e['date']
        dates_set.add(dk)
        skey = f"{e['source_type']}:{e['assignment_id']}"
        if skey not in sources:
            sources[skey] = {
                'key': skey,
                'source_type': e['source_type'],
                'assignment_id': e['assignment_id'],
                'name': boat_names.get((e['source_type'], e['assignment_id']), '?'),
                'by_date': {},
                'total': 0,
                'diesel': 0,
                'petrol': 0,
            }
        src = sources[skey]
        liters = e['liters'] or 0
        ft = e['fuel_type'] or 'DIESEL'
        if dk not in src['by_date']:
            src['by_date'][dk] = 0
        src['by_date'][dk] += liters
        src['total'] += liters
        if ft == 'PETROL':
            src['petrol'] += liters
        else:
            src['diesel'] += liters

    dates = sorted(dates_set)

    # Daily totals
    daily_totals = {}
    for dk in dates:
        daily_totals[dk] = 0
    for src in sources.values():
        for dk, l in src['by_date'].items():
            daily_totals[dk] = daily_totals.get(dk, 0) + l

    # Category labels
    cat_labels = {
        'boats': 'BOAT', 'picture_boats': 'PB',
        'security_boats': 'SB', 'transport': 'TRANSPO', 'machinery': 'MACH'
    }
    source_list = sorted(sources.values(), key=lambda s: (s['source_type'], s['name']))
    for s in source_list:
        s['category'] = cat_labels.get(s['source_type'], s['source_type'].upper())

    return jsonify({
        'dates': dates,
        'sources': source_list,
        'daily_totals': daily_totals,
        'grand_total': sum(daily_totals.values()),
    })


# ─── Fuel entries ────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/fuel-entries", methods=["GET"])
def api_get_fuel_entries(prod_id):
    prod_or_404(prod_id)
    source = request.args.get('source')
    return jsonify_cached(get_fuel_entries(prod_id, source_type=source or None))


@app.route("/api/productions/<int:prod_id>/fuel-entries", methods=["POST"])
def api_upsert_fuel_entry(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data['production_id'] = prod_id
    validate_fuel_entry(data)
    entry = upsert_fuel_entry(data)
    return jsonify(entry or {}), 200


@app.route("/api/fuel-entries/<int:entry_id>", methods=["DELETE"])
def api_delete_fuel_entry(entry_id):
    delete_fuel_entry(entry_id)
    return jsonify({"ok": True})


@app.route("/api/productions/<int:prod_id>/fuel-entries/assignment/<source_type>/<int:assignment_id>", methods=["DELETE"])
def api_delete_fuel_entries_for_assignment(prod_id, source_type, assignment_id):
    prod_or_404(prod_id)
    delete_fuel_entries_for_assignment(source_type, assignment_id)
    return jsonify({"ok": True})


# ─── Fuel machinery ───────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/fuel-machinery", methods=["GET"])
def api_get_fuel_machinery(prod_id):
    prod_or_404(prod_id)
    include_deleted = request.args.get('include_deleted', '').lower() == 'true'
    return jsonify(get_fuel_machinery(prod_id, include_deleted=include_deleted))


@app.route("/api/productions/<int:prod_id>/fuel-machinery", methods=["POST"])
def api_create_fuel_machinery(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data['production_id'] = prod_id
    if not data.get("name"):
        return jsonify({"error": "name required"}), 400
    validate_numeric_fields(data, ['liters_per_day'])
    mid = create_fuel_machinery(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM fuel_machinery WHERE id=?", (mid,)).fetchone()
    return jsonify(dict(row) if row else {}), 201


@app.route("/api/fuel-machinery/<int:machinery_id>", methods=["PUT"])
def api_update_fuel_machinery(machinery_id):
    data = request.json or {}
    validate_entity_name(data)
    validate_numeric_fields(data, ['liters_per_day'])
    update_fuel_machinery(machinery_id, data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM fuel_machinery WHERE id=?", (machinery_id,)).fetchone()
    return jsonify(dict(row) if row else {})


@app.route("/api/fuel-machinery/<int:machinery_id>", methods=["DELETE"])
def api_delete_fuel_machinery(machinery_id):
    delete_fuel_machinery(machinery_id)
    return jsonify({"ok": True})


# ─── Fuel prices (global settings) ───────────────────────────────────────────

@app.route("/api/fuel-prices", methods=["GET"])
def api_get_fuel_prices():
    """Get global fuel prices (diesel/petrol in USD per litre)."""
    diesel = get_setting("fuel_price_diesel", "0")
    petrol = get_setting("fuel_price_petrol", "0")
    return jsonify({"diesel": float(diesel), "petrol": float(petrol)})


@app.route("/api/fuel-prices", methods=["PUT"])
def api_set_fuel_prices():
    """Set global fuel prices."""
    data = request.json or {}
    validate_numeric_fields(data, ['diesel', 'petrol'])
    if "diesel" in data:
        set_setting("fuel_price_diesel", str(float(data["diesel"])))
    if "petrol" in data:
        set_setting("fuel_price_petrol", str(float(data["petrol"])))
    diesel = get_setting("fuel_price_diesel", "0")
    petrol = get_setting("fuel_price_petrol", "0")
    return jsonify({"diesel": float(diesel), "petrol": float(petrol)})


# ─── Fuel locked prices (day snapshots) ─────────────────────────────────────

@app.route("/api/fuel-locked-prices", methods=["GET"])
def api_get_fuel_locked_prices():
    """Get all locked day price snapshots."""
    return jsonify(get_fuel_locked_prices())


@app.route("/api/fuel-locked-prices", methods=["POST"])
def api_lock_fuel_day():
    """Lock a day with current fuel prices snapshot."""
    data = request.json or {}
    date = data.get("date")
    if not date:
        return jsonify({"error": "date is required"}), 400
    validate_numeric_fields(data, ['diesel_price', 'petrol_price'])
    diesel_price = float(data.get("diesel_price", 0))
    petrol_price = float(data.get("petrol_price", 0))
    locked_by = getattr(g, 'nickname', None)
    set_fuel_locked_price(date, diesel_price, petrol_price, locked_by=locked_by)
    # AXE 6.3: auto-snapshot on fuel lock
    # Need prod_id — try X-Project-Id header
    prod_id_header = request.headers.get("X-Project-Id")
    if prod_id_header:
        try:
            create_budget_snapshot(
                int(prod_id_header), trigger_type='lock',
                trigger_detail=f"Fuel lock: {date}",
                user_id=getattr(g, 'user_id', None),
                user_nickname=getattr(g, 'nickname', None)
            )
        except Exception:
            pass
    return jsonify({"ok": True, "date": date, "diesel_price": diesel_price, "petrol_price": petrol_price})


@app.route("/api/fuel-locked-prices/<date>", methods=["DELETE"])
def api_unlock_fuel_day(date):
    """Unlock a day (remove the price snapshot)."""
    delete_fuel_locked_price(date)
    return jsonify({"ok": True})


# ─── Fuel budget export (from BUDGET tab) ────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/export/fuel-budget/csv")
def api_export_fuel_budget_csv(prod_id):
    """Export fuel budget breakdown by consumer: total litres + total price.
    Filename: KLAS7_FUEL_YYMMDD
    """
    from datetime import datetime as dt
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    entries = get_fuel_entries(prod_id)
    entries = _filter_entries_by_date(entries, date_from, date_to)
    machinery = get_fuel_machinery(prod_id)
    locked_prices = get_fuel_locked_prices()

    # Current global prices
    cur_diesel = float(get_setting("fuel_price_diesel", "0"))
    cur_petrol = float(get_setting("fuel_price_petrol", "0"))

    # Build consumer breakdown: group by source_type + entity name
    # We need assignment data to resolve consumer names
    from database import (get_boat_assignments, get_picture_boat_assignments,
                          get_transport_assignments, get_security_boat_assignments)
    asgn_map = {}
    for ctx, fetcher in [
        ('boats', lambda: get_boat_assignments(prod_id, context='boats')),
        ('picture_boats', lambda: get_picture_boat_assignments(prod_id)),
        ('security_boats', lambda: get_security_boat_assignments(prod_id)),
        ('transport', lambda: get_transport_assignments(prod_id)),
    ]:
        for a in fetcher():
            asgn_map[(ctx, a['id'])] = (
                a.get('boat_name_override') or a.get('boat_name') or
                a.get('vehicle_name_override') or a.get('vehicle_name') or '?',
                a.get('function_name') or '?'
            )

    # Group entries by consumer (skip machinery — handled separately below)
    consumers = {}
    for e in entries:
        if e['source_type'] == 'machinery':
            continue
        key = (e['source_type'], e['assignment_id'])
        name_info = asgn_map.get(key, (f"#{e['assignment_id']}", '?'))
        consumer_key = f"{e['source_type'].upper()} | {name_info[0]} | {name_info[1]}"
        if consumer_key not in consumers:
            consumers[consumer_key] = {'diesel_l': 0, 'petrol_l': 0, 'cost_up_to_date': 0, 'cost_estimate': 0}
        ft = e.get('fuel_type', 'DIESEL')
        liters = e.get('liters', 0) or 0
        date = e.get('date', '')
        # Price: use locked price if day is locked, else current price
        if date in locked_prices:
            price = locked_prices[date]['diesel_price'] if ft == 'DIESEL' else locked_prices[date]['petrol_price']
            consumers[consumer_key]['cost_up_to_date'] += liters * price
        else:
            price = cur_diesel if ft == 'DIESEL' else cur_petrol
            consumers[consumer_key]['cost_estimate'] += liters * price
        if ft == 'PETROL':
            consumers[consumer_key]['petrol_l'] += liters
        else:
            consumers[consumer_key]['diesel_l'] += liters

    # Machinery — computed from actual fuel_entries (source_type='machinery')
    # Build a name lookup from machinery records
    machinery_names = {m['id']: m['name'] for m in machinery}
    for e in entries:
        if e['source_type'] != 'machinery':
            continue
        m_name = machinery_names.get(e['assignment_id'], f"Machine #{e['assignment_id']}")
        consumer_key = f"MACHINERY | {m_name}"
        if consumer_key not in consumers:
            consumers[consumer_key] = {'diesel_l': 0, 'petrol_l': 0, 'cost_up_to_date': 0, 'cost_estimate': 0}
        ft = e.get('fuel_type', 'DIESEL')
        liters = e.get('liters', 0) or 0
        date = e.get('date', '')
        if date in locked_prices:
            price = locked_prices[date]['diesel_price'] if ft == 'DIESEL' else locked_prices[date]['petrol_price']
            consumers[consumer_key]['cost_up_to_date'] += liters * price
        else:
            price = cur_diesel if ft == 'DIESEL' else cur_petrol
            consumers[consumer_key]['cost_estimate'] += liters * price
        if ft == 'PETROL':
            consumers[consumer_key]['petrol_l'] += liters
        else:
            consumers[consumer_key]['diesel_l'] += liters

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Consumer", "Diesel (L)", "Petrol (L)", "Total (L)",
                "Cost Up to Date ($)", "Cost Estimate ($)", "Total Cost ($)"])
    grand_diesel = 0
    grand_petrol = 0
    grand_cost_utd = 0
    grand_cost_est = 0
    for name, data in sorted(consumers.items()):
        total_l = data['diesel_l'] + data['petrol_l']
        total_cost = data['cost_up_to_date'] + data['cost_estimate']
        w.writerow([name, round(data['diesel_l'], 1), round(data['petrol_l'], 1),
                    round(total_l, 1), round(data['cost_up_to_date'], 2),
                    round(data['cost_estimate'], 2), round(total_cost, 2)])
        grand_diesel += data['diesel_l']
        grand_petrol += data['petrol_l']
        grand_cost_utd += data['cost_up_to_date']
        grand_cost_est += data['cost_estimate']
    w.writerow([])
    grand_total_l = grand_diesel + grand_petrol
    grand_total_cost = grand_cost_utd + grand_cost_est
    # Compute average price per fuel type (cost / litres for each type separately)
    diesel_cost_total = 0
    petrol_cost_total = 0
    for e in entries:
        ft = e.get('fuel_type', 'DIESEL')
        liters = e.get('liters', 0) or 0
        date = e.get('date', '')
        if date in locked_prices:
            price = locked_prices[date]['diesel_price'] if ft == 'DIESEL' else locked_prices[date]['petrol_price']
        else:
            price = cur_diesel if ft == 'DIESEL' else cur_petrol
        if ft == 'PETROL':
            petrol_cost_total += liters * price
        else:
            diesel_cost_total += liters * price
    avg_diesel = diesel_cost_total / grand_diesel if grand_diesel > 0 else 0
    avg_petrol = petrol_cost_total / grand_petrol if grand_petrol > 0 else 0
    w.writerow(["GRAND TOTAL", round(grand_diesel, 1), round(grand_petrol, 1),
                round(grand_total_l, 1), round(grand_cost_utd, 2),
                round(grand_cost_est, 2), round(grand_total_cost, 2)])
    w.writerow(["AVG PRICE PER LITRE — DIESEL", "", "", "", "", "", round(avg_diesel, 4)])
    w.writerow(["AVG PRICE PER LITRE — PETROL", "", "", "", "", "", round(avg_petrol, 4)])
    w.writerow([])
    w.writerow([f"Current Diesel price: ${cur_diesel}/L"])
    w.writerow([f"Current Petrol price: ${cur_petrol}/L"])
    out.seek(0)
    prod_name = prod.get("name", "PRODUCTION")
    fname = _export_fname(prod_name, "FUEL-BUDGET", date_from, date_to, "csv")
    return Response(out.read(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


# ─── Fuel exports ─────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/export/fuel/csv")
def api_export_fuel_csv(prod_id):
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    entries = get_fuel_entries(prod_id)
    entries = _filter_entries_by_date(entries, date_from, date_to)
    machinery = get_fuel_machinery(prod_id)
    import csv, io
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Category", "Name / Function", "Date", "Liters", "Fuel Type"])
    totals = {"DIESEL": 0, "PETROL": 0}
    machinery_names = {m['id']: m['name'] for m in machinery}
    for e in entries:
        src = e.get("source_type", "")
        name = e.get("assignment_id", "")
        if src == "machinery":
            name = machinery_names.get(e.get("assignment_id"), name)
        w.writerow([src, name,
                    e.get("date", ""), e.get("liters", 0), e.get("fuel_type", "")])
        ft = e.get("fuel_type", "DIESEL")
        totals[ft] = totals.get(ft, 0) + (e.get("liters") or 0)
    w.writerow([])
    w.writerow(["GRAND TOTAL DIESEL", "", "", totals.get("DIESEL", 0), "DIESEL"])
    w.writerow(["GRAND TOTAL PETROL", "", "", totals.get("PETROL", 0), "PETROL"])
    out.seek(0)
    prod_name = prod.get("name", "PRODUCTION")
    fname = _export_fname(prod_name, "FUEL", date_from, date_to, "csv")
    return Response(out.read(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.route("/api/productions/<int:prod_id>/export/fuel/json")
def api_export_fuel_json(prod_id):
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    entries = get_fuel_entries(prod_id)
    entries = _filter_entries_by_date(entries, date_from, date_to)
    data = {
        "production": get_production(prod_id),
        "fuel_entries": entries,
        "fuel_machinery": get_fuel_machinery(prod_id),
        "date_range": {"from": date_from, "to": date_to},
    }
    prod_name = prod.get("name", "PRODUCTION")
    fname = _export_fname(prod_name, "FUEL", date_from, date_to, "json")
    return Response(
        json.dumps(data, indent=2, ensure_ascii=False),
        mimetype="application/json",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ─── Boat photos: auto-match from static folder + BATEAUX migration ──────────

def _normalize_name(name):
    """Normalize a boat name for fuzzy matching: lowercase, strip accents,
    remove special chars, collapse whitespace."""
    import re, unicodedata
    if not name:
        return ''
    # NFD decompose then strip combining marks (accents)
    s = unicodedata.normalize('NFD', name)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.lower()
    s = re.sub(r'[^a-z0-9]', '', s)
    return s


def _ensure_boat_images_symlink():
    """Create a symlink from SHOOTLOGIX/static/boat_images -> BATEAUX/static/boat_images
    so that existing image_path values (static/boat_images/...) resolve correctly."""
    import os
    src = os.path.join(os.path.dirname(__file__), '..', 'BATEAUX', 'static', 'boat_images')
    dst = os.path.join(os.path.dirname(__file__), 'static', 'boat_images')
    if os.path.isdir(src) and not os.path.exists(dst):
        try:
            os.symlink(os.path.abspath(src), dst)
        except OSError:
            pass  # symlink may fail on some systems; images will still be served from uploads


def _auto_match_boat_photos():
    """Scan available image files and update boats/picture_boats/security_boats
    that are missing a valid image_path.

    Priority order for each boat:
    1. Already-uploaded photo in static/uploads/boats/boat_{id}.{ext}
    2. BATEAUX source image matched by boat_nr (BOAT__{nr}__*.jpg pattern)
    3. BATEAUX source image matched by fuzzy boat name
    4. SVG placeholder matched by fuzzy boat name
    """
    import os, re, shutil

    app_dir = os.path.dirname(os.path.abspath(__file__))
    bateaux_src = os.path.join(app_dir, '..', 'BATEAUX', 'static', 'boat_images')
    uploads_boats = os.path.join(app_dir, 'static', 'uploads', 'boats')
    uploads_pb = os.path.join(app_dir, 'static', 'uploads', 'picture-boats')
    uploads_sb = os.path.join(app_dir, 'static', 'uploads', 'security-boats')
    os.makedirs(uploads_boats, exist_ok=True)
    os.makedirs(uploads_pb, exist_ok=True)
    os.makedirs(uploads_sb, exist_ok=True)

    # Build index of BATEAUX source images
    bateaux_by_nr = {}    # boat_nr -> filename (JPG only, from BOAT__{nr}__ pattern)
    bateaux_by_name = {}  # normalized_name -> filename (all files)
    if os.path.isdir(bateaux_src):
        for fname in os.listdir(bateaux_src):
            fpath = os.path.join(bateaux_src, fname)
            if not os.path.isfile(fpath):
                continue
            ext = os.path.splitext(fname)[1].lower()
            if ext not in ('.jpg', '.jpeg', '.png', '.webp', '.svg'):
                continue
            # Try boat_nr pattern: BOAT__{nr}__*.jpg
            m = re.match(r'BOAT_+(\d+)_+', fname)
            if m and ext in ('.jpg', '.jpeg', '.png', '.webp'):
                bateaux_by_nr[int(m.group(1))] = fname
            # Name-based: strip extension, normalize
            base = os.path.splitext(fname)[0]
            # Remove BOAT__{nr}__ prefix for name matching
            base_clean = re.sub(r'^BOAT_+\d+_+', '', base)
            norm = _normalize_name(base_clean)
            if norm:
                # Prefer JPG over SVG
                existing = bateaux_by_name.get(norm)
                if existing is None:
                    bateaux_by_name[norm] = fname
                elif ext in ('.jpg', '.jpeg', '.png', '.webp') and existing.lower().endswith('.svg'):
                    bateaux_by_name[norm] = fname  # upgrade SVG to real photo

    # Build index of already-uploaded files
    def _uploaded_files(directory):
        result = {}
        if os.path.isdir(directory):
            for f in os.listdir(directory):
                ext = os.path.splitext(f)[1].lower()
                if ext in ('.jpg', '.jpeg', '.png', '.webp'):
                    # Extract ID from filename: boat_123.jpg or pb_5.png etc
                    m2 = re.match(r'(?:boat|pb|sb)_(\d+)', f)
                    if m2:
                        result[int(m2.group(1))] = f
        return result

    uploaded_boats = _uploaded_files(uploads_boats)
    uploaded_pbs = _uploaded_files(uploads_pb)
    uploaded_sbs = _uploaded_files(uploads_sb)

    matched = 0

    with get_db() as conn:
        # --- BOATS ---
        boats = conn.execute("SELECT id, name, boat_nr, image_path FROM boats").fetchall()
        for b in boats:
            bid = b['id']
            current = b['image_path'] or ''

            # Priority 1: uploaded photo always wins (even over symlinked BATEAUX images)
            if bid in uploaded_boats:
                rel = f"static/uploads/boats/{uploaded_boats[bid]}"
                if current != rel:
                    conn.execute("UPDATE boats SET image_path=? WHERE id=?", (rel, bid))
                    matched += 1
                continue

            # Check if current image_path actually resolves to a file
            if current:
                full = os.path.join(app_dir, current)
                if os.path.isfile(full):
                    continue  # already valid, skip

            # Priority 2: BATEAUX by boat_nr
            nr = b['boat_nr']
            if nr and nr in bateaux_by_nr:
                src_file = bateaux_by_nr[nr]
                ext = os.path.splitext(src_file)[1].lower()
                dst_name = f"boat_{bid}{ext}"
                shutil.copy2(os.path.join(bateaux_src, src_file),
                             os.path.join(uploads_boats, dst_name))
                rel = f"static/uploads/boats/{dst_name}"
                conn.execute("UPDATE boats SET image_path=? WHERE id=?", (rel, bid))
                uploaded_boats[bid] = dst_name
                matched += 1
                continue

            # Priority 3: BATEAUX by fuzzy name (prefer real photo over SVG)
            norm = _normalize_name(b['name'])
            if norm and norm in bateaux_by_name:
                src_file = bateaux_by_name[norm]
                ext = os.path.splitext(src_file)[1].lower()
                if ext in ('.jpg', '.jpeg', '.png', '.webp'):
                    dst_name = f"boat_{bid}{ext}"
                    shutil.copy2(os.path.join(bateaux_src, src_file),
                                 os.path.join(uploads_boats, dst_name))
                    rel = f"static/uploads/boats/{dst_name}"
                else:
                    # SVG: reference via boat_images symlink
                    rel = f"static/boat_images/{src_file}"
                conn.execute("UPDATE boats SET image_path=? WHERE id=?", (rel, bid))
                matched += 1
                continue

        # --- PICTURE BOATS ---
        pbs = conn.execute("SELECT id, name, image_path FROM picture_boats").fetchall()
        for pb in pbs:
            pid = pb['id']
            current = pb['image_path'] or ''
            # Uploaded photo always wins
            if pid in uploaded_pbs:
                rel = f"static/uploads/picture-boats/{uploaded_pbs[pid]}"
                if current != rel:
                    conn.execute("UPDATE picture_boats SET image_path=? WHERE id=?", (rel, pid))
                    matched += 1
                continue
            if current:
                full = os.path.join(app_dir, current)
                if os.path.isfile(full):
                    continue
            # Try fuzzy name match against BATEAUX images
            norm = _normalize_name(pb['name'])
            if norm and norm in bateaux_by_name:
                src_file = bateaux_by_name[norm]
                ext = os.path.splitext(src_file)[1].lower()
                if ext in ('.jpg', '.jpeg', '.png', '.webp'):
                    dst_name = f"pb_{pid}{ext}"
                    shutil.copy2(os.path.join(bateaux_src, src_file),
                                 os.path.join(uploads_pb, dst_name))
                    rel = f"static/uploads/picture-boats/{dst_name}"
                else:
                    rel = f"static/boat_images/{src_file}"
                conn.execute("UPDATE picture_boats SET image_path=? WHERE id=?", (rel, pid))
                matched += 1

        # --- SECURITY BOATS ---
        sbs = conn.execute("SELECT id, name, image_path FROM security_boats").fetchall()
        for sb in sbs:
            sid = sb['id']
            current = sb['image_path'] or ''
            # Uploaded photo always wins
            if sid in uploaded_sbs:
                rel = f"static/uploads/security-boats/{uploaded_sbs[sid]}"
                if current != rel:
                    conn.execute("UPDATE security_boats SET image_path=? WHERE id=?", (rel, sid))
                    matched += 1
                continue
            if current:
                full = os.path.join(app_dir, current)
                if os.path.isfile(full):
                    continue
            norm = _normalize_name(sb['name'])
            if norm and norm in bateaux_by_name:
                src_file = bateaux_by_name[norm]
                ext = os.path.splitext(src_file)[1].lower()
                if ext in ('.jpg', '.jpeg', '.png', '.webp'):
                    dst_name = f"sb_{sid}{ext}"
                    shutil.copy2(os.path.join(bateaux_src, src_file),
                                 os.path.join(uploads_sb, dst_name))
                    rel = f"static/uploads/security-boats/{dst_name}"
                else:
                    rel = f"static/boat_images/{src_file}"
                conn.execute("UPDATE security_boats SET image_path=? WHERE id=?", (rel, sid))
                matched += 1

    return matched


@app.route("/api/migrate-boat-photos", methods=["POST"])
def api_migrate_boat_photos():
    """Re-run the auto-match logic and return how many boats were updated."""
    matched = _auto_match_boat_photos()
    return jsonify({"matched": matched})


@app.route("/api/auto-match-photos", methods=["POST"])
def api_auto_match_photos():
    """Scan static folders and auto-match photos to boats, picture boats,
    and security boats by name/number. Returns count of newly matched."""
    matched = _auto_match_boat_photos()
    return jsonify({"matched": matched})


# ─── Data reload ──────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/reload", methods=["POST"])
def api_reload(prod_id):
    prod_or_404(prod_id)
    from data_loader import migrate_from_bateaux
    result = migrate_from_bateaux(prod_id, force=True)
    return jsonify(result)


# ─── Location Sites CRUD ─────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/locations", methods=["GET"])
def api_get_locations(prod_id):
    prod_or_404(prod_id)
    include_deleted = request.args.get('include_deleted', '').lower() == 'true'
    return jsonify(get_location_sites(prod_id, include_deleted=include_deleted))


@app.route("/api/productions/<int:prod_id>/locations", methods=["POST"])
def api_create_location(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data['production_id'] = prod_id
    if not data.get('name'):
        return jsonify({"error": "name is required"}), 400
    result = create_location_site(data)
    return jsonify(result), 201


@app.route("/api/locations/<int:loc_id>", methods=["PUT"])
def api_update_location(loc_id):
    data = request.json or {}
    validate_entity_name(data)
    validate_numeric_fields(data, ['price_p', 'price_f', 'price_w', 'global_deal'])
    # Get old name for schedule cascade
    from database import get_db
    with get_db() as conn:
        old = conn.execute("SELECT * FROM locations WHERE id=?", (loc_id,)).fetchone()
    if not old:
        return jsonify({"error": "not found"}), 404
    old = dict(old)
    old_name = old['name']
    result = update_location_site(loc_id, data)
    if result is False:
        return jsonify({"error": "This record was modified by another user. Please refresh."}), 409
    # Cascade rename in schedules
    if 'name' in data and data['name'] != old_name:
        rename_location_in_schedules(old['production_id'], old_name, data['name'])
    # AXE 6.3: log price changes for locations
    _log_price_changes(old['production_id'], 'location', loc_id, old_name, old, data)
    return jsonify(result or {})


@app.route("/api/locations/<int:loc_id>", methods=["DELETE"])
def api_delete_location(loc_id):
    delete_location_site(loc_id)
    return jsonify({"deleted": loc_id})



# ─── Guard Posts CRUD ────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/guard-posts", methods=["GET"])
def api_get_guard_posts(prod_id):
    prod_or_404(prod_id)
    include_deleted = request.args.get('include_deleted', '').lower() == 'true'
    return jsonify(get_guard_posts(prod_id, include_deleted=include_deleted))


@app.route("/api/productions/<int:prod_id>/guard-posts", methods=["POST"])
def api_create_guard_post(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data['production_id'] = prod_id
    if not data.get('name'):
        return jsonify({"error": "name is required"}), 400
    result = create_guard_post(data)
    return jsonify(result), 201


@app.route("/api/guard-posts/<int:post_id>", methods=["PUT"])
def api_update_guard_post(post_id):
    data = request.json or {}
    validate_entity_name(data)
    from database import get_db
    with get_db() as conn:
        old = conn.execute("SELECT * FROM guard_posts WHERE id=?", (post_id,)).fetchone()
    if not old:
        return jsonify({"error": "not found"}), 404
    old = dict(old)
    old_name = old['name']
    result = update_guard_post(post_id, data)
    # Cascade rename in guard schedules
    if 'name' in data and data['name'] != old_name:
        rename_guard_post_in_schedules(old['production_id'], old_name, data['name'])
    return jsonify(result or {})


@app.route("/api/guard-posts/<int:post_id>", methods=["DELETE"])
def api_delete_guard_post(post_id):
    delete_guard_post(post_id)
    return jsonify({"deleted": post_id})


# ─── Location Schedules ──────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/location-schedules", methods=["GET"])
def api_get_location_schedules(prod_id):
    prod_or_404(prod_id)
    return jsonify(get_location_schedules(prod_id))


@app.route("/api/productions/<int:prod_id>/location-schedules", methods=["POST"])
def api_upsert_location_schedule(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data['production_id'] = prod_id
    if (not data.get('location_name') and not data.get('location_id')) or not data.get('date') or not data.get('status'):
        return jsonify({"error": "location_name or location_id, date, status required"}), 400
    result = upsert_location_schedule(data)
    return jsonify(result or {}), 200


@app.route("/api/productions/<int:prod_id>/location-schedules/delete", methods=["POST"])
def api_delete_location_schedule(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    if (not data.get('location_name') and not data.get('location_id')) or not data.get('date'):
        return jsonify({"error": "location_name or location_id, and date required"}), 400
    delete_location_schedule(prod_id, data.get('location_name'), data['date'], location_id=data.get('location_id'))
    return jsonify({"ok": True})


@app.route("/api/productions/<int:prod_id>/location-schedules/<int:schedule_id>", methods=["DELETE"])
def api_delete_location_schedule_by_id(prod_id, schedule_id):
    prod_or_404(prod_id)
    delete_location_schedule_by_id(schedule_id)
    return jsonify({"deleted": schedule_id})


@app.route("/api/productions/<int:prod_id>/location-schedules/lock", methods=["PUT"])
def api_lock_location_schedules(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    dates = data.get('dates', [])
    if not isinstance(dates, list):
        return jsonify({"error": "dates must be a list"}), 400
    locked = data.get('locked', True)
    lock_location_schedules(prod_id, dates, locked)
    # AXE 6.3: auto-snapshot on lock
    if locked and dates:
        try:
            create_budget_snapshot(
                prod_id, trigger_type='lock',
                trigger_detail=f"Location lock: {', '.join(dates[:5])}{'...' if len(dates) > 5 else ''}",
                user_id=getattr(g, 'user_id', None),
                user_nickname=getattr(g, 'nickname', None)
            )
        except Exception:
            pass  # snapshot failure should not block lock
    return jsonify({"ok": True})


@app.route("/api/productions/<int:prod_id>/location-schedules/auto-fill", methods=["POST"])
def api_auto_fill_locations(prod_id):
    prod_or_404(prod_id)
    created = auto_fill_locations_from_pdt(prod_id)
    return jsonify({"created": created})


@app.route("/api/productions/<int:prod_id>/sync-pdt-locations", methods=["POST"])
def api_sync_pdt_locations(prod_id):
    """Sync a single PDT day's locations to the Locations schedule.
    Body: { date: "YYYY-MM-DD", locations: ["LOC1", "LOC2", ...] }
    For deletions: { date: "YYYY-MM-DD", locations: [], deleted: true }
    """
    prod_or_404(prod_id)
    data = request.json or {}
    day_date = data.get('date')
    if not day_date:
        return jsonify({"error": "date required"}), 400

    if data.get('deleted'):
        remove_pdt_film_days_for_date(prod_id, day_date)
    else:
        locations = data.get('locations', [])
        sync_pdt_day_to_locations(prod_id, day_date, locations)
    return jsonify({"ok": True})


@app.route("/api/productions/<int:prod_id>/resync-pdt-locations", methods=["POST"])
def api_resync_all_pdt_locations(prod_id):
    """Full resync: iterate all shooting days and sync their locations."""
    prod_or_404(prod_id)
    days = get_shooting_days(prod_id)
    total_log = {'created': [], 'matched': [], 'ignored': []}
    for day in days:
        day_date = day.get('date')
        if not day_date:
            continue
        locations = []
        if day.get('location'):
            locations.append(day['location'])
        events = get_events_for_day(day['id'])
        for ev in events:
            if ev.get('location'):
                locations.append(ev['location'])
        if locations:
            log = sync_pdt_day_to_locations(prod_id, day_date, locations)
            if log:
                total_log['created'].extend(log.get('created', []))
                total_log['matched'].extend(log.get('matched', []))
                total_log['ignored'].extend(log.get('ignored', []))
    # Deduplicate
    total_log['created'] = list(set(total_log['created']))
    total_log['matched'] = list(set(total_log['matched']))
    return jsonify(total_log)


# ─── Guard Location Schedules ───────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/guard-schedules", methods=["GET"])
def api_get_guard_location_schedules(prod_id):
    prod_or_404(prod_id)
    return jsonify(get_guard_location_schedules(prod_id))


@app.route("/api/productions/<int:prod_id>/guard-schedules", methods=["POST"])
def api_upsert_guard_location_schedule(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data['production_id'] = prod_id
    if (not data.get('location_name') and not data.get('location_id')) or not data.get('date') or not data.get('status'):
        return jsonify({"error": "location_name or location_id, date, status required"}), 400
    result = upsert_guard_location_schedule(data)
    return jsonify(result or {}), 200


@app.route("/api/productions/<int:prod_id>/guard-schedules/delete", methods=["POST"])
def api_delete_guard_location_schedule(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    if (not data.get('location_name') and not data.get('location_id')) or not data.get('date'):
        return jsonify({"error": "location_name or location_id, and date required"}), 400
    delete_guard_location_schedule(prod_id, data.get('location_name'), data['date'], location_id=data.get('location_id'))
    return jsonify({"ok": True})


@app.route("/api/productions/<int:prod_id>/guard-schedules/lock", methods=["PUT"])
def api_lock_guard_location_schedules(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    dates = data.get('dates', [])
    if not isinstance(dates, list):
        return jsonify({"error": "dates must be a list"}), 400
    locked = data.get('locked', True)
    lock_guard_location_schedules(prod_id, dates, locked)
    # AXE 6.3: auto-snapshot on lock
    if locked and dates:
        try:
            create_budget_snapshot(
                prod_id, trigger_type='lock',
                trigger_detail=f"Guard lock: {', '.join(dates[:5])}{'...' if len(dates) > 5 else ''}",
                user_id=getattr(g, 'user_id', None),
                user_nickname=getattr(g, 'nickname', None)
            )
        except Exception:
            pass
    return jsonify({"ok": True})


@app.route("/api/productions/<int:prod_id>/guard-schedules/sync", methods=["POST"])
def api_sync_guard_location_schedules(prod_id):
    """Sync guard_location_schedules from location_schedules (auto-populate defaults)."""
    prod_or_404(prod_id)
    result = sync_guard_location_from_locations(prod_id)
    return jsonify(result)


@app.route("/api/productions/<int:prod_id>/guard-schedules/update-guards", methods=["POST"])
def api_update_guard_location_nb_guards(prod_id):
    """Update nb_guards for a specific location/date."""
    prod_or_404(prod_id)
    data = request.json or {}
    if (not data.get('location_name') and not data.get('location_id')) or not data.get('date'):
        return jsonify({"error": "location_name or location_id, and date required"}), 400
    validate_guard_schedule(data)
    nb_guards = int(data.get('nb_guards', 0))
    result = update_guard_location_nb_guards(prod_id, data.get('location_name'), data['date'], nb_guards, location_id=data.get('location_id'))
    return jsonify(result or {})


# ─── Guard Camp (Base Camp) Workers & Assignments ─────────────────────────

@app.route("/api/productions/<int:prod_id>/guard-camp-workers", methods=["GET"])
def api_guard_camp_workers(prod_id):
    prod_or_404(prod_id)
    include_deleted = request.args.get('include_deleted', '').lower() == 'true'
    return jsonify(get_guard_camp_workers(prod_id, include_deleted=include_deleted))


@app.route("/api/productions/<int:prod_id>/guard-camp-workers", methods=["POST"])
def api_create_guard_camp_worker(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data["production_id"] = prod_id
    if not data.get("name"):
        return jsonify({"error": "name required"}), 400
    worker_id = create_guard_camp_worker(data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM guard_camp_workers WHERE id=?", (worker_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/productions/<int:prod_id>/guard-camp-workers/bulk", methods=["POST"])
def api_bulk_create_guard_camp_workers(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    try:
        count = int(data.get("count", 0))
    except (ValueError, TypeError):
        return jsonify({"error": "count must be an integer"}), 400
    prefix = data.get("prefix", "Guard")
    if count < 1 or count > 200:
        return jsonify({"error": "count must be 1-200"}), 400
    shared = {
        "production_id": prod_id,
        "role": data.get("role"),
        "group_name": data.get("group_name", "GENERAL"),
        "daily_rate_estimate": data.get("daily_rate_estimate", 45),
        "notes": data.get("notes"),
    }
    func_id = data.get("boat_function_id")
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    created = []
    for i in range(1, count + 1):
        rec = dict(shared)
        rec["name"] = f"{prefix} {i}"
        wid = create_guard_camp_worker(rec)
        if func_id and wid:
            assign_data = {"boat_function_id": func_id, "helper_id": wid,
                           "start_date": start_date, "end_date": end_date}
            create_guard_camp_assignment(assign_data)
        created.append(wid)
    return jsonify({"created": len(created), "ids": created}), 201


@app.route("/api/productions/<int:prod_id>/guard-camp-workers/import-csv", methods=["POST"])
def api_import_guard_camp_workers_csv(prod_id):
    import csv, io
    prod_or_404(prod_id)
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400
    content = f.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    created = []
    for row in reader:
        name = (row.get("name") or "").strip()
        if not name:
            continue
        rec = {
            "production_id": prod_id,
            "name": name,
            "role": (row.get("role") or "").strip() or None,
            "group_name": (row.get("group") or row.get("group_name") or "GENERAL").strip(),
            "daily_rate_estimate": float(row.get("rate") or row.get("daily_rate_estimate") or 45),
            "notes": (row.get("notes") or "").strip() or None,
        }
        wid = create_guard_camp_worker(rec)
        created.append(wid)
    return jsonify({"created": len(created), "ids": created}), 201


@app.route("/api/guard-camp-workers/<int:worker_id>", methods=["PUT"])
def api_update_guard_camp_worker(worker_id):
    data = request.json or {}
    validate_entity_name(data)
    validate_numeric_fields(data, ['daily_rate_estimate', 'daily_rate_actual'])
    ok = update_guard_camp_worker(worker_id, data)
    if ok is False:
        return jsonify({"error": "This record was modified by another user. Please refresh."}), 409
    with get_db() as conn:
        row = conn.execute("SELECT * FROM guard_camp_workers WHERE id=?", (worker_id,)).fetchone()
    if not row:
        abort(404)
    return jsonify(dict(row))


@app.route("/api/guard-camp-workers/<int:worker_id>", methods=["DELETE"])
def api_delete_guard_camp_worker(worker_id):
    delete_guard_camp_worker(worker_id)
    return jsonify({"deleted": worker_id})


@app.route("/api/guard-camp-workers/<int:worker_id>/duplicate", methods=["POST"])
def api_duplicate_guard_camp_worker(worker_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM guard_camp_workers WHERE id=?", (worker_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    data = dict(row)
    del data["id"]
    data["name"] = data["name"] + " (copy)"
    data.pop("image_path", None)
    new_id = create_guard_camp_worker(data)
    with get_db() as conn:
        new_row = conn.execute("SELECT * FROM guard_camp_workers WHERE id=?", (new_id,)).fetchone()
    return jsonify(dict(new_row)), 201


# ─── Assignment Duplication (AXE 10.2) ────────────────────────────────────────

_ASSIGNMENT_TABLES = {
    "boat": "boat_assignments",
    "picture_boat": "picture_boat_assignments",
    "security_boat": "security_boat_assignments",
    "transport": "transport_assignments",
    "helper": "helper_assignments",
    "guard_camp": "guard_camp_assignments",
}


@app.route("/api/assignments/<string:atype>/<int:assignment_id>/duplicate", methods=["POST"])
def api_duplicate_assignment(atype, assignment_id):
    """Duplicate an assignment with dates shifted +7 days (configurable).
    Body: { offset_days?: 7 }"""
    table = _ASSIGNMENT_TABLES.get(atype)
    if not table:
        return jsonify({"error": f"Unknown assignment type: {atype}"}), 400
    data = request.json or {}
    offset = int(data.get("offset_days", 7))
    result = duplicate_assignment(table, assignment_id, offset)
    if not result:
        return jsonify({"error": "Assignment not found"}), 404
    return jsonify(result), 201


# ─── Bulk Operations (AXE 10.3) ───────────────────────────────────────────────

_ENTITY_TABLES = {
    "boats": ("boats", "name", update_boat, delete_boat),
    "picture_boats": ("picture_boats", "name", update_picture_boat, delete_picture_boat),
    "security_boats": ("security_boats", "name", update_security_boat, delete_security_boat),
    "transport": ("transport_vehicles", "name", update_transport_vehicle, delete_transport_vehicle),
    "helpers": ("helpers", "name", update_helper, delete_helper),
    "guard_camp": ("guard_camp_workers", "name", update_guard_camp_worker, delete_guard_camp_worker),
}


@app.route("/api/productions/<int:prod_id>/bulk-update", methods=["POST"])
def api_bulk_update(prod_id):
    """Bulk update entities. Body: { entity_type, ids: [...], updates: {field: value} }"""
    prod_or_404(prod_id)
    data = request.json or {}
    etype = data.get("entity_type")
    ids = data.get("ids", [])
    updates = data.get("updates", {})
    if not etype or not ids or not updates:
        return jsonify({"error": "entity_type, ids, and updates required"}), 400

    entry = _ENTITY_TABLES.get(etype)
    if not entry:
        return jsonify({"error": f"Unknown entity type: {etype}"}), 400
    table_name, name_field, update_fn, _ = entry

    # Whitelist safe fields
    safe_fields = {"group_name", "daily_rate_estimate", "daily_rate_actual", "rate_estimated",
                   "rate_actual", "vendor", "notes", "role", "sort_order", "pricing_type",
                   "include_sunday", "function_group", "currency"}
    filtered = {k: v for k, v in updates.items() if k in safe_fields}
    if not filtered:
        return jsonify({"error": "No valid fields to update"}), 400

    updated = 0
    for eid in ids:
        try:
            update_fn(eid, filtered)
            updated += 1
        except Exception:
            pass
    return jsonify({"updated": updated})


@app.route("/api/productions/<int:prod_id>/bulk-delete", methods=["POST"])
def api_bulk_delete(prod_id):
    """Bulk delete entities. Body: { entity_type, ids: [...] }"""
    prod_or_404(prod_id)
    data = request.json or {}
    etype = data.get("entity_type")
    ids = data.get("ids", [])
    if not etype or not ids:
        return jsonify({"error": "entity_type and ids required"}), 400

    entry = _ENTITY_TABLES.get(etype)
    if not entry:
        return jsonify({"error": f"Unknown entity type: {etype}"}), 400
    _, _, _, delete_fn = entry

    deleted = 0
    for eid in ids:
        try:
            delete_fn(eid)
            deleted += 1
        except Exception:
            pass
    return jsonify({"deleted": deleted})


@app.route("/api/guard-camp-workers/<int:worker_id>/upload-image", methods=["POST"])
def api_upload_guard_camp_worker_image(worker_id):
    if 'image' not in request.files:
        return jsonify({"error": "No image file"}), 400
    file = request.files['image']
    upload_dir = os.path.join(app.static_folder, 'uploads', 'guard_camp')
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(file.filename)[1] or '.jpg'
    fname = f"gc_{worker_id}{ext}"
    filepath = os.path.join(upload_dir, fname)
    file.save(filepath)
    rel_path = f"static/uploads/guard_camp/{fname}"
    update_guard_camp_worker(worker_id, {"image_path": rel_path})
    with get_db() as conn:
        row = conn.execute("SELECT * FROM guard_camp_workers WHERE id=?", (worker_id,)).fetchone()
    return jsonify(dict(row))


@app.route("/api/productions/<int:prod_id>/guard-camp-assignments", methods=["GET"])
def api_guard_camp_assignments(prod_id):
    prod_or_404(prod_id)
    return jsonify(get_guard_camp_assignments(prod_id))


@app.route("/api/productions/<int:prod_id>/guard-camp-assignments", methods=["POST"])
def api_create_guard_camp_assignment(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    if not data.get("boat_function_id"):
        return jsonify({"error": "boat_function_id required"}), 400
    validate_assignment(data)
    _validate_dates_for_prod(prod_id, data)
    if data.get("helper_id") and data.get("start_date") and data.get("end_date"):
        validate_assignment_overlap('guard_camp_assignments', 'helper_id', data['helper_id'],
                                    data['start_date'], data['end_date'])
    assignment_id = create_guard_camp_assignment(data)
    assignments = get_guard_camp_assignments(prod_id)
    asgn = next((a for a in assignments if a["id"] == assignment_id), None)
    return jsonify(asgn), 201


@app.route("/api/guard-camp-assignments/<int:assignment_id>", methods=["PUT"])
def api_update_guard_camp_assignment(assignment_id):
    data = request.json or {}
    validate_assignment(data)
    prod_id = _get_prod_id_from_function(data.get("boat_function_id"))
    if prod_id:
        _validate_dates_for_prod(prod_id, data)
    if data.get("helper_id") and data.get("start_date") and data.get("end_date"):
        validate_assignment_overlap('guard_camp_assignments', 'helper_id', data['helper_id'],
                                    data['start_date'], data['end_date'], exclude_id=assignment_id)
    # P5.6: validate override_reason when price_override changes
    with get_db() as conn:
        old_row = conn.execute("SELECT gca.*, bf.production_id FROM guard_camp_assignments gca LEFT JOIN boat_functions bf ON gca.boat_function_id=bf.id WHERE gca.id=?", (assignment_id,)).fetchone()
    old_d = dict(old_row) if old_row else {}
    err = _validate_price_override(data, old_d)
    if err:
        return err
    update_guard_camp_assignment(assignment_id, data)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM guard_camp_assignments WHERE id=?", (assignment_id,)).fetchone()
    if old_d.get('production_id'):
        _log_price_changes(old_d['production_id'], 'guard_camp_assignment', assignment_id,
                          old_d.get('helper_name_override', f'Assignment #{assignment_id}'), old_d, data)
    if not row:
        abort(404)
    return jsonify(dict(row))


@app.route("/api/guard-camp-assignments/<int:assignment_id>", methods=["DELETE"])
def api_delete_guard_camp_assignment(assignment_id):
    delete_guard_camp_assignment(assignment_id)
    return jsonify({"deleted": assignment_id})


@app.route("/api/productions/<int:prod_id>/guard-camp-assignments/function/<int:func_id>", methods=["DELETE"])
def api_delete_guard_camp_assignment_by_function(prod_id, func_id):
    prod_or_404(prod_id)
    delete_guard_camp_assignment_by_function(func_id)
    return jsonify({"ok": True})


@app.route("/api/productions/<int:prod_id>/export/guard-camp/csv")
def api_export_guard_camp_csv(prod_id):
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    rows = [r for r in get_guard_camp_assignments(prod_id) if r.get("working_days")]
    rows = _filter_assignments_by_date(rows, date_from, date_to)
    from collections import OrderedDict
    by_group = OrderedDict()
    for r in rows:
        g = r.get("function_group") or r.get("helper_group") or "GENERAL"
        if g not in by_group:
            by_group[g] = []
        by_group[g].append(r)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Group", "Function", "Guard", "Role", "Contact",
                     "Start", "End", "Working Days", "Rate/day",
                     "Total Estimate", "Total Actual"])
    grand_est = 0
    grand_act = 0
    for group_name, group_rows in by_group.items():
        group_est = 0
        group_act = 0
        for r in group_rows:
            est = r.get("amount_estimate") or 0
            act = r.get("amount_actual") or 0
            group_est += est
            group_act += act
            writer.writerow([
                group_name,
                r.get("function_name") or "",
                r.get("helper_name_override") or r.get("helper_name") or "",
                r.get("helper_role") or "",
                r.get("helper_contact") or "",
                r.get("start_date") or "", r.get("end_date") or "",
                r.get("working_days") or "",
                r.get("price_override") or r.get("helper_daily_rate_estimate") or "",
                est,
                act if act else "",
            ])
        writer.writerow(["", "", "", "", "", "", f"SUB-TOTAL {group_name}", "", "", group_est, group_act if group_act else ""])
        writer.writerow([])
        grand_est += group_est
        grand_act += group_act
    writer.writerow(["", "", "", "", "", "", "GRAND TOTAL", "", "", grand_est, grand_act if grand_act else ""])
    output.seek(0)
    prod_name = prod.get("name", "PRODUCTION")
    fname = _export_fname(prod_name, "GUARDS-BASECAMP", date_from, date_to, "csv")
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ─── Security Auto-fill from Locations ──────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/security-auto-fill", methods=["POST"])
def api_security_auto_fill(prod_id):
    """Auto-create security boat day overrides based on location schedule F days."""
    prod_or_404(prod_id)
    loc_schedules = get_location_schedules(prod_id)
    # Filter only F (filming) days for game locations
    filming_days = [ls for ls in loc_schedules
                    if ls['status'] == 'F' and ls['location_type'] == 'game']
    return jsonify({
        "message": f"Found {len(filming_days)} filming days at game locations",
        "filming_days": filming_days,
    })


# ─── FNB Budget Export ───────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/export/fnb-budget/csv")
def api_export_fnb_budget_csv(prod_id):
    """Export simplified FNB budget: totals by category only.
    Two columns: Up to Date (consumption) and Estimate (purchases).
    """
    from datetime import datetime as dt
    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    budget = get_fnb_budget_data(prod_id)

    out = io.StringIO()
    w = csv.writer(out)
    prod_name = prod.get("name", "PRODUCTION")
    w.writerow([f"{prod_name} - FNB BUDGET EXPORT"])
    gen_line = f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"
    if date_from and date_to:
        gen_line += f"  |  Period: {date_from} to {date_to}"
    w.writerow([gen_line])
    w.writerow([])
    w.writerow(["Category", "Up to Date ($)", "Estimate ($)", "Total ($)"])

    grand_utd = 0
    grand_est = 0
    for cat in budget.get('categories', []):
        utd = round(cat.get('consumption_total', 0), 2)
        est = round(cat.get('purchase_total', 0), 2)
        total = round(utd + est, 2)
        w.writerow([cat['name'], utd, est, total])
        grand_utd += utd
        grand_est += est

    w.writerow([])
    grand_total = round(grand_utd + grand_est, 2)
    w.writerow(["GRAND TOTAL", round(grand_utd, 2), round(grand_est, 2), grand_total])
    w.writerow([])
    balance = round(grand_est - grand_utd, 2)
    w.writerow([f"Balance (Estimate - Up to Date): ${balance}"])

    out.seek(0)
    fname = _export_fname(prod_name, "CATERING", date_from, date_to, "csv")
    return Response(out.read(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


# ─── FNB Tracking ───────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/fnb-tracking", methods=["GET"])
def api_get_fnb_tracking(prod_id):
    prod_or_404(prod_id)
    return jsonify(get_fnb_tracking(prod_id))


@app.route("/api/productions/<int:prod_id>/fnb-tracking", methods=["POST"])
def api_upsert_fnb_tracking(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data['production_id'] = prod_id
    if not data.get('date') or not data.get('category'):
        return jsonify({"error": "date and category required"}), 400
    result = upsert_fnb_tracking(data)
    return jsonify(result or {}), 200


@app.route("/api/fnb-tracking/<int:entry_id>", methods=["DELETE"])
def api_delete_fnb_tracking(entry_id):
    delete_fnb_tracking(entry_id)
    return jsonify({"deleted": entry_id})


@app.route("/api/productions/<int:prod_id>/fnb-summary", methods=["GET"])
def api_fnb_summary(prod_id):
    prod_or_404(prod_id)
    return jsonify(get_fnb_summary(prod_id))


# ─── FNB v2 (dynamic categories / items / entries) ──────────────────────────

@app.route("/api/productions/<int:prod_id>/fnb-categories", methods=["GET"])
def api_get_fnb_categories(prod_id):
    prod_or_404(prod_id)
    include_deleted = request.args.get('include_deleted', '').lower() == 'true'
    return jsonify(get_fnb_categories(prod_id, include_deleted=include_deleted))


@app.route("/api/productions/<int:prod_id>/fnb-categories", methods=["POST"])
def api_create_fnb_category(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data['production_id'] = prod_id
    if not data.get('name'):
        return jsonify({"error": "name required"}), 400
    try:
        result = create_fnb_category(data)
        return jsonify(result), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/fnb-categories/<int:cat_id>", methods=["PUT"])
def api_update_fnb_category(cat_id):
    data = request.json or {}
    validate_entity_name(data)
    result = update_fnb_category(cat_id, data)
    if not result:
        abort(404)
    return jsonify(result)


@app.route("/api/fnb-categories/<int:cat_id>", methods=["DELETE"])
def api_delete_fnb_category(cat_id):
    delete_fnb_category(cat_id)
    return jsonify({"deleted": cat_id})


@app.route("/api/fnb-categories/<int:cat_id>/duplicate", methods=["POST"])
def api_duplicate_fnb_category(cat_id):
    """Duplicate a FNB category with all its items."""
    result = duplicate_fnb_category(cat_id)
    if not result:
        return jsonify({"error": "Category not found"}), 404
    return jsonify(result), 201


@app.route("/api/productions/<int:prod_id>/fnb-items", methods=["GET"])
def api_get_fnb_items(prod_id):
    prod_or_404(prod_id)
    include_deleted = request.args.get('include_deleted', '').lower() == 'true'
    return jsonify(get_fnb_items(prod_id, include_deleted=include_deleted))


@app.route("/api/productions/<int:prod_id>/fnb-items", methods=["POST"])
def api_create_fnb_item(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data['production_id'] = prod_id
    if not data.get('name') or not data.get('category_id'):
        return jsonify({"error": "name and category_id required"}), 400
    try:
        result = create_fnb_item(data)
        return jsonify(result), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/fnb-items/<int:item_id>", methods=["PUT"])
def api_update_fnb_item(item_id):
    data = request.json or {}
    validate_entity_name(data)
    validate_numeric_fields(data, ['unit_price'])
    result = update_fnb_item(item_id, data)
    if not result:
        abort(404)
    return jsonify(result)


@app.route("/api/fnb-items/<int:item_id>", methods=["DELETE"])
def api_delete_fnb_item(item_id):
    delete_fnb_item(item_id)
    return jsonify({"deleted": item_id})


@app.route("/api/productions/<int:prod_id>/fnb-entries", methods=["GET"])
def api_get_fnb_entries(prod_id):
    prod_or_404(prod_id)
    entry_type = request.args.get('type')
    return jsonify(get_fnb_entries(prod_id, entry_type))


@app.route("/api/productions/<int:prod_id>/fnb-entries", methods=["POST"])
def api_upsert_fnb_entry(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    data['production_id'] = prod_id
    if not data.get('item_id') or not data.get('entry_type') or not data.get('date'):
        return jsonify({"error": "item_id, entry_type and date required"}), 400
    if data['entry_type'] not in ('purchase', 'consumption'):
        return jsonify({"error": "entry_type must be 'purchase' or 'consumption'"}), 400
    validate_numeric_fields(data, ['quantity', 'unit_price'])
    result = upsert_fnb_entry(data)
    return jsonify(result or {}), 200


@app.route("/api/fnb-entries/<int:entry_id>", methods=["DELETE"])
def api_delete_fnb_entry(entry_id):
    delete_fnb_entry(entry_id)
    return jsonify({"deleted": entry_id})


@app.route("/api/productions/<int:prod_id>/fnb-budget", methods=["GET"])
def api_fnb_budget(prod_id):
    prod_or_404(prod_id)
    return jsonify(get_fnb_budget_data(prod_id))


# ─── Global Budget Export (multi-sheet Excel) ────────────────────────────────

@app.route("/api/productions/<int:prod_id>/export/budget-global")
def api_export_budget_global(prod_id):
    """Export full budget as a multi-sheet Excel file (.xlsx).
    One sheet per category, plus a Summary sheet.
    Filename: KLAS7_BUDGET_YYMMDD.xlsx
    """
    from datetime import datetime as dt
    from collections import OrderedDict
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    prod_or_404(prod_id)

    wb = Workbook()

    # -- Style constants --
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    title_font = Font(bold=True, size=12)
    subtotal_font = Font(bold=True, size=10)
    money_fmt = '#,##0'
    money_fmt_dec = '#,##0.00'
    green_font = Font(bold=True, color="22C55E")
    thin_border = Border(
        bottom=Side(style='thin', color='CCCCCC')
    )

    def style_header_row(ws, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

    # Collect category totals for summary
    summary_data = OrderedDict()

    # ── Sheet 1: LOCATIONS ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Locations"
    loc_schedules = get_location_schedules(prod_id)
    loc_sites = get_location_sites(prod_id)
    site_pricing = {}
    for s in loc_sites:
        site_pricing[s['name']] = {
            'price_p': s.get('price_p') or 0,
            'price_f': s.get('price_f') or 0,
            'price_w': s.get('price_w') or 0,
            'global_deal': s.get('global_deal'),
        }

    loc_day_counts = {}
    for ls in loc_schedules:
        loc_name = ls['location_name']
        loc_day_counts.setdefault(loc_name, {'P': 0, 'F': 0, 'W': 0})
        if ls['status'] in ('P', 'F', 'W'):
            loc_day_counts[loc_name][ls['status']] += 1

    ws.append(["KLAS 7 - LOCATIONS BUDGET"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Location", "Type", "P Days", "F Days", "W Days", "Total Days",
               "$/P", "$/F", "$/W", "Global Deal", "Total ($)"])
    style_header_row(ws, 11)

    loc_grand = 0
    for loc_name, counts in sorted(loc_day_counts.items()):
        pricing = site_pricing.get(loc_name, {'price_p': 0, 'price_f': 0, 'price_w': 0, 'global_deal': None})
        site_info = next((s for s in loc_sites if s['name'] == loc_name), {})
        loc_type = site_info.get('location_type', '')
        if pricing['global_deal'] and pricing['global_deal'] > 0:
            total = pricing['global_deal']
        else:
            total = (counts['P'] * pricing['price_p'] +
                     counts['F'] * pricing['price_f'] +
                     counts['W'] * pricing['price_w'])
        loc_grand += total
        ws.append([loc_name, loc_type, counts['P'], counts['F'], counts['W'],
                   counts['P'] + counts['F'] + counts['W'],
                   pricing['price_p'], pricing['price_f'], pricing['price_w'],
                   pricing['global_deal'] or "", round(total, 2)])
    ws.append([])
    ws.append(["", "", "", "", "", "", "", "", "", "GRAND TOTAL", round(loc_grand, 2)])
    ws.cell(row=ws.max_row, column=11).font = green_font
    auto_width(ws)
    summary_data["LOCATIONS"] = round(loc_grand, 2)

    # ── Sheet 2: BOATS ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Boats")
    budget = get_budget(prod_id)
    boat_rows = [r for r in get_boat_assignments(prod_id, context='boats') if r.get("working_days")]
    ws.append(["KLAS 7 - BOATS BUDGET"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Department", "Function", "Boat", "Vendor", "Start", "End",
               "Working Days", "Rate/day", "Total Estimate", "Total Actual"])
    style_header_row(ws, 10)
    grand_est = 0
    grand_act = 0
    for r in boat_rows:
        est = r.get("amount_estimate") or 0
        act = r.get("amount_actual") or 0
        grand_est += est
        grand_act += act
        ws.append([
            r.get("department") or "BOATS",
            r.get("function_name") or r.get("name") or "",
            r.get("boat_name_override") or r.get("boat_name") or "",
            r.get("vendor") or "",
            r.get("start_date") or "", r.get("end_date") or "",
            r.get("working_days") or "",
            r.get("price_override") or r.get("boat_daily_rate_estimate") or "",
            est, act if act else "",
        ])
    ws.append([])
    ws.append(["", "", "", "", "", "", "GRAND TOTAL", "", round(grand_est, 2), round(grand_act, 2)])
    ws.cell(row=ws.max_row, column=9).font = green_font
    auto_width(ws)
    summary_data["BOATS"] = round(grand_est, 2)

    # ── Sheet 3: PICTURE BOATS ────────────────────────────────────────────────
    ws = wb.create_sheet("Picture Boats")
    pb_rows = [r for r in get_picture_boat_assignments(prod_id) if r.get("working_days")]
    ws.append(["KLAS 7 - PICTURE BOATS BUDGET"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Function", "Group", "Boat", "Captain", "Vendor",
               "Start", "End", "Working Days", "Rate/day (est.)",
               "Total Estimate", "Total Actual"])
    style_header_row(ws, 11)
    grand_est = 0
    grand_act = 0
    for r in pb_rows:
        est = r.get("amount_estimate") or 0
        act = r.get("amount_actual") or 0
        grand_est += est
        grand_act += act
        ws.append([
            r.get("function_name") or "",
            r.get("function_group") or "",
            r.get("boat_name_override") or r.get("boat_name") or "",
            r.get("captain") or "",
            r.get("vendor") or "",
            r.get("start_date") or "", r.get("end_date") or "",
            r.get("working_days") or "",
            r.get("price_override") or r.get("boat_daily_rate_estimate") or "",
            est, act if act else "",
        ])
    ws.append([])
    ws.append(["", "", "", "", "", "", "GRAND TOTAL", "", "", round(grand_est, 2), round(grand_act, 2)])
    ws.cell(row=ws.max_row, column=10).font = green_font
    auto_width(ws)
    summary_data["PICTURE BOATS"] = round(grand_est, 2)

    # ── Sheet 4: SECURITY BOATS ───────────────────────────────────────────────
    ws = wb.create_sheet("Security Boats")
    sb_rows = [r for r in get_security_boat_assignments(prod_id) if r.get("working_days")]
    ws.append(["KLAS 7 - SECURITY BOATS BUDGET"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Function", "Group", "Boat", "Captain", "Vendor",
               "Start", "End", "Working Days", "Rate/day (est.)",
               "Total Estimate", "Total Actual"])
    style_header_row(ws, 11)
    grand_est = 0
    grand_act = 0
    for r in sb_rows:
        est = r.get("amount_estimate") or 0
        act = r.get("amount_actual") or 0
        grand_est += est
        grand_act += act
        ws.append([
            r.get("function_name") or "",
            r.get("function_group") or "",
            r.get("boat_name_override") or r.get("boat_name") or "",
            r.get("captain") or "",
            r.get("vendor") or "",
            r.get("start_date") or "", r.get("end_date") or "",
            r.get("working_days") or "",
            r.get("price_override") or r.get("boat_daily_rate_estimate") or "",
            est, act if act else "",
        ])
    ws.append([])
    ws.append(["", "", "", "", "", "", "GRAND TOTAL", "", "", round(grand_est, 2), round(grand_act, 2)])
    ws.cell(row=ws.max_row, column=10).font = green_font
    auto_width(ws)
    summary_data["SECURITY BOATS"] = round(grand_est, 2)

    # ── Sheet 5: TRANSPORT ────────────────────────────────────────────────────
    ws = wb.create_sheet("Transport")
    tr_rows = [r for r in get_transport_assignments(prod_id) if r.get("working_days")]
    ws.append(["KLAS 7 - TRANSPORT BUDGET"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Function", "Group", "Vehicle", "Type", "Driver", "Vendor",
               "Start", "End", "Working Days", "Rate/day (est.)",
               "Total Estimate", "Total Actual"])
    style_header_row(ws, 12)
    grand_est = 0
    grand_act = 0
    for r in tr_rows:
        est = r.get("amount_estimate") or 0
        act = r.get("amount_actual") or 0
        grand_est += est
        grand_act += act
        ws.append([
            r.get("function_name") or "",
            r.get("function_group") or "",
            r.get("vehicle_name_override") or r.get("vehicle_name") or "",
            r.get("vehicle_type") or "",
            r.get("driver") or "",
            r.get("vendor") or "",
            r.get("start_date") or "", r.get("end_date") or "",
            r.get("working_days") or "",
            r.get("price_override") or r.get("vehicle_daily_rate_estimate") or "",
            est, act if act else "",
        ])
    ws.append([])
    ws.append(["", "", "", "", "", "", "GRAND TOTAL", "", "", "", round(grand_est, 2), round(grand_act, 2)])
    ws.cell(row=ws.max_row, column=11).font = green_font
    auto_width(ws)
    summary_data["TRANSPORT"] = round(grand_est, 2)

    # ── Sheet 6: FUEL ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Fuel")
    entries = get_fuel_entries(prod_id)
    machinery = get_fuel_machinery(prod_id)
    locked_prices = get_fuel_locked_prices()
    cur_diesel = float(get_setting("fuel_price_diesel", "0"))
    cur_petrol = float(get_setting("fuel_price_petrol", "0"))

    # Build consumer breakdown (same logic as fuel-budget CSV export)
    asgn_map = {}
    for ctx, fetcher in [
        ('boats', lambda: get_boat_assignments(prod_id, context='boats')),
        ('picture_boats', lambda: get_picture_boat_assignments(prod_id)),
        ('security_boats', lambda: get_security_boat_assignments(prod_id)),
        ('transport', lambda: get_transport_assignments(prod_id)),
    ]:
        for a in fetcher():
            asgn_map[(ctx, a['id'])] = (
                a.get('boat_name_override') or a.get('boat_name') or
                a.get('vehicle_name_override') or a.get('vehicle_name') or '?',
                a.get('function_name') or '?'
            )

    consumers = {}
    for e in entries:
        if e['source_type'] == 'machinery':
            continue  # handled separately below with proper name resolution
        key = (e['source_type'], e['assignment_id'])
        name_info = asgn_map.get(key, (f"#{e['assignment_id']}", '?'))
        consumer_key = f"{e['source_type'].upper()} | {name_info[0]} | {name_info[1]}"
        if consumer_key not in consumers:
            consumers[consumer_key] = {'diesel_l': 0, 'petrol_l': 0, 'cost_up_to_date': 0, 'cost_estimate': 0}
        ft = e.get('fuel_type', 'DIESEL')
        liters = e.get('liters', 0) or 0
        date = e.get('date', '')
        if date in locked_prices:
            price = locked_prices[date]['diesel_price'] if ft == 'DIESEL' else locked_prices[date]['petrol_price']
            consumers[consumer_key]['cost_up_to_date'] += liters * price
        else:
            price = cur_diesel if ft == 'DIESEL' else cur_petrol
            consumers[consumer_key]['cost_estimate'] += liters * price
        if ft == 'PETROL':
            consumers[consumer_key]['petrol_l'] += liters
        else:
            consumers[consumer_key]['diesel_l'] += liters

    # Machinery entries from fuel_entries with source_type='machinery'
    # Resolve machinery names for consumer keys
    machinery_names = {m['id']: m['name'] for m in machinery}
    for e in entries:
        if e['source_type'] != 'machinery':
            continue
        m_name = machinery_names.get(e['assignment_id'], f"Machine #{e['assignment_id']}")
        consumer_key = f"MACHINERY | {m_name}"
        if consumer_key not in consumers:
            consumers[consumer_key] = {'diesel_l': 0, 'petrol_l': 0, 'cost_up_to_date': 0, 'cost_estimate': 0}
        ft = e.get('fuel_type', 'DIESEL')
        liters = e.get('liters', 0) or 0
        date = e.get('date', '')
        if date in locked_prices:
            price = locked_prices[date]['diesel_price'] if ft == 'DIESEL' else locked_prices[date]['petrol_price']
            consumers[consumer_key]['cost_up_to_date'] += liters * price
        else:
            price = cur_diesel if ft == 'DIESEL' else cur_petrol
            consumers[consumer_key]['cost_estimate'] += liters * price
        if ft == 'PETROL':
            consumers[consumer_key]['petrol_l'] += liters
        else:
            consumers[consumer_key]['diesel_l'] += liters

    ws.append(["KLAS 7 - FUEL BUDGET"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Consumer", "Diesel (L)", "Petrol (L)", "Total (L)",
               "Cost Up to Date ($)", "Cost Estimate ($)", "Total Cost ($)"])
    style_header_row(ws, 7)
    fuel_grand_diesel = 0
    fuel_grand_petrol = 0
    fuel_grand_utd = 0
    fuel_grand_est = 0
    for name, data in sorted(consumers.items()):
        total_l = data['diesel_l'] + data['petrol_l']
        total_cost = data['cost_up_to_date'] + data['cost_estimate']
        ws.append([name, round(data['diesel_l'], 1), round(data['petrol_l'], 1),
                   round(total_l, 1), round(data['cost_up_to_date'], 2),
                   round(data['cost_estimate'], 2), round(total_cost, 2)])
        fuel_grand_diesel += data['diesel_l']
        fuel_grand_petrol += data['petrol_l']
        fuel_grand_utd += data['cost_up_to_date']
        fuel_grand_est += data['cost_estimate']
    ws.append([])
    fuel_total_l = fuel_grand_diesel + fuel_grand_petrol
    fuel_total_cost = fuel_grand_utd + fuel_grand_est
    avg_price = fuel_total_cost / fuel_total_l if fuel_total_l > 0 else 0
    ws.append(["GRAND TOTAL", round(fuel_grand_diesel, 1), round(fuel_grand_petrol, 1),
               round(fuel_total_l, 1), round(fuel_grand_utd, 2),
               round(fuel_grand_est, 2), round(fuel_total_cost, 2)])
    ws.cell(row=ws.max_row, column=7).font = green_font
    ws.append(["AVERAGE PRICE PER LITRE", "", "", "", "", "", round(avg_price, 4)])
    ws.append([])
    ws.append([f"Current Diesel price: ${cur_diesel}/L"])
    ws.append([f"Current Petrol price: ${cur_petrol}/L"])
    auto_width(ws)
    summary_data["FUEL"] = round(fuel_total_cost, 2)

    # ── Sheet 7: LABOUR ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Labor")
    lb_rows = [r for r in get_helper_assignments(prod_id) if r.get("working_days")]
    by_group = OrderedDict()
    for r in lb_rows:
        g = r.get("function_group") or r.get("helper_group") or "GENERAL"
        by_group.setdefault(g, []).append(r)

    ws.append(["KLAS 7 - LABOR BUDGET"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Group", "Function", "Worker", "Role", "Contact",
               "Start", "End", "Working Days", "Rate/day",
               "Total Estimate", "Total Actual"])
    style_header_row(ws, 11)
    grand_est = 0
    grand_act = 0
    for group_name, group_rows in by_group.items():
        group_est = 0
        group_act = 0
        for r in group_rows:
            est = r.get("amount_estimate") or 0
            act = r.get("amount_actual") or 0
            group_est += est
            group_act += act
            ws.append([
                group_name,
                r.get("function_name") or "",
                r.get("helper_name_override") or r.get("helper_name") or "",
                r.get("helper_role") or "",
                r.get("helper_contact") or "",
                r.get("start_date") or "", r.get("end_date") or "",
                r.get("working_days") or "",
                r.get("price_override") or r.get("helper_daily_rate_estimate") or "",
                est, act if act else "",
            ])
        ws.append(["", "", "", "", "", "", f"SUB-TOTAL {group_name}", "", "",
                   round(group_est, 2), round(group_act, 2) if group_act else ""])
        ws.cell(row=ws.max_row, column=10).font = subtotal_font
        ws.append([])
        grand_est += group_est
        grand_act += group_act
    ws.append(["", "", "", "", "", "", "GRAND TOTAL", "", "",
               round(grand_est, 2), round(grand_act, 2) if grand_act else ""])
    ws.cell(row=ws.max_row, column=10).font = green_font
    auto_width(ws)
    summary_data["LABOUR"] = round(grand_est, 2)

    # ── Sheet 8: GUARDS (merged Location + Base Camp) ───────────────────────
    ws = wb.create_sheet("Guards")
    ws.append(["KLAS 7 - GUARDS BUDGET"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])

    # Part A: Location Guards (from guard_location_schedules)
    guard_loc_data = get_guard_location_schedules(prod_id)
    type_by_name = {s['name']: s.get('location_type', 'game') for s in loc_sites}
    loc_guard_by_loc = {}
    for gls in guard_loc_data:
        loc_name = gls['location_name']
        nb = gls.get('nb_guards', 2)
        loc_type = type_by_name.get(loc_name, 'game')
        loc_guard_by_loc.setdefault(loc_name, {'days': 0, 'total_guard_days': 0, 'cost': 0, 'type': loc_type})
        loc_guard_by_loc[loc_name]['days'] += 1
        loc_guard_by_loc[loc_name]['total_guard_days'] += nb
        loc_guard_by_loc[loc_name]['cost'] += nb * 45

    ws.append(["LOCATION GUARDS"])
    ws.cell(row=ws.max_row, column=1).font = subtotal_font
    ws.append(["Location", "Type", "Active Days", "Guard-Days", "Rate/Guard/Day ($)", "Total ($)"])
    style_header_row(ws, 6)
    gl_grand = 0
    for loc_name, info in sorted(loc_guard_by_loc.items()):
        ws.append([loc_name, info['type'], info['days'], info['total_guard_days'], 45, round(info['cost'], 2)])
        gl_grand += info['cost']
    ws.append([])
    ws.append(["", "", "", "", "SUB-TOTAL LOCATION", round(gl_grand, 2)])
    ws.cell(row=ws.max_row, column=6).font = subtotal_font
    ws.append([])

    # Part B: Base Camp Guards
    gc_rows = [r for r in get_guard_camp_assignments(prod_id) if r.get("working_days")]
    by_group = OrderedDict()
    for r in gc_rows:
        g = r.get("function_group") or r.get("helper_group") or "GENERAL"
        by_group.setdefault(g, []).append(r)

    ws.append(["BASE CAMP GUARDS"])
    ws.cell(row=ws.max_row, column=1).font = subtotal_font
    ws.append(["Group", "Function", "Guard", "Role", "Contact",
               "Start", "End", "Working Days", "Rate/day",
               "Total Estimate", "Total Actual"])
    style_header_row(ws, 11)
    gc_grand_est = 0
    gc_grand_act = 0
    for group_name, group_rows in by_group.items():
        group_est = 0
        group_act = 0
        for r in group_rows:
            est = r.get("amount_estimate") or 0
            act = r.get("amount_actual") or 0
            group_est += est
            group_act += act
            ws.append([
                group_name,
                r.get("function_name") or "",
                r.get("helper_name_override") or r.get("helper_name") or "",
                r.get("helper_role") or "",
                r.get("helper_contact") or "",
                r.get("start_date") or "", r.get("end_date") or "",
                r.get("working_days") or "",
                r.get("price_override") or r.get("helper_daily_rate_estimate") or "",
                est, act if act else "",
            ])
        ws.append(["", "", "", "", "", "", f"SUB-TOTAL {group_name}", "", "",
                   round(group_est, 2), round(group_act, 2) if group_act else ""])
        ws.cell(row=ws.max_row, column=10).font = subtotal_font
        ws.append([])
        gc_grand_est += group_est
        gc_grand_act += group_act
    ws.append(["", "", "", "", "", "", "SUB-TOTAL BASE CAMP", "", "",
               round(gc_grand_est, 2), round(gc_grand_act, 2) if gc_grand_act else ""])
    ws.cell(row=ws.max_row, column=10).font = subtotal_font
    ws.append([])
    total_guards = gl_grand + gc_grand_est
    ws.append(["", "", "", "", "", "", "GRAND TOTAL GUARDS", "", "",
               round(total_guards, 2), ""])
    ws.cell(row=ws.max_row, column=10).font = green_font
    auto_width(ws)
    summary_data["GUARDS"] = round(total_guards, 2)

    # ── Sheet 10: FNB ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Catering")
    fnb_budget = get_fnb_budget_data(prod_id)
    ws.append(["KLAS 7 - CATERING BUDGET"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Category", "Up to Date ($)", "Estimate ($)", "Total ($)"])
    style_header_row(ws, 4)
    fnb_grand_utd = 0
    fnb_grand_est = 0
    for cat in fnb_budget.get('categories', []):
        utd = round(cat.get('consumption_total', 0), 2)
        est = round(cat.get('purchase_total', 0), 2)
        total = round(utd + est, 2)
        ws.append([cat['name'], utd, est, total])
        fnb_grand_utd += utd
        fnb_grand_est += est
    ws.append([])
    fnb_total = round(fnb_grand_utd + fnb_grand_est, 2)
    ws.append(["GRAND TOTAL", round(fnb_grand_utd, 2), round(fnb_grand_est, 2), fnb_total])
    ws.cell(row=ws.max_row, column=4).font = green_font
    ws.append([])
    balance = round(fnb_grand_est - fnb_grand_utd, 2)
    ws.append([f"Balance (Estimate - Up to Date): ${balance}"])
    auto_width(ws)
    summary_data["CATERING"] = round(fnb_grand_est, 2)  # Use purchase (estimate) as budget total

    # ── Insert Summary sheet at position 0 ────────────────────────────────────
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append(["KLAS 7 - GLOBAL BUDGET SUMMARY"])
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_summary.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws_summary.append([])
    ws_summary.append(["Category", "Total Estimate ($)"])
    style_header_row(ws_summary, 2)
    overall_total = 0
    for dept, total in summary_data.items():
        ws_summary.append([dept, total])
        ws_summary.cell(row=ws_summary.max_row, column=2).number_format = money_fmt_dec
        overall_total += total
    ws_summary.append([])
    ws_summary.append(["GRAND TOTAL", round(overall_total, 2)])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True, size=12)
    ws_summary.cell(row=ws_summary.max_row, column=2).font = Font(bold=True, size=12, color="22C55E")
    ws_summary.cell(row=ws_summary.max_row, column=2).number_format = money_fmt_dec
    auto_width(ws_summary)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"KLAS7_BUDGET_{dt.now().strftime('%y%m%d')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ─── ENRICHED BUDGET XLSX EXPORT (P5.8) ─────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/export/budget-xlsx")
def api_export_budget_xlsx(prod_id):
    """Export enriched multi-sheet budget XLSX with variance summary and history."""
    from datetime import datetime as dt
    from collections import OrderedDict
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    prod = prod_or_404(prod_id)
    prod_name = prod.get("name", "PROD") if isinstance(prod, dict) else "PROD"

    wb = Workbook()

    # -- Style constants --
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    title_font = Font(bold=True, size=12)
    subtotal_font = Font(bold=True, size=10)
    money_fmt = '#,##0.00'
    green_font = Font(bold=True, color="22C55E")
    red_font = Font(bold=True, color="EF4444")
    thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))

    def style_header_row(ws, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

    def add_total_row(ws, label_col, label, value_cols, values, font=green_font):
        row_data = [""] * (max(value_cols) + 1)
        row_data[label_col] = label
        for c, v in zip(value_cols, values):
            row_data[c] = v
        ws.append(row_data)
        for c in value_cols:
            ws.cell(row=ws.max_row, column=c + 1).font = font
            ws.cell(row=ws.max_row, column=c + 1).number_format = money_fmt

    # ── Collect budget data via get_budget ──
    budget = get_budget(prod_id)
    by_dept = budget.get("by_department", {})

    # Build summary data: dept -> {estimated, actual, variance, variance_pct, lines}
    summary_rows = OrderedDict()
    for dept_name, ddata in sorted(by_dept.items()):
        est = ddata.get("total_estimate", 0) or 0
        act = ddata.get("total_actual", 0) or 0
        var = act - est
        var_pct = round((var / est) * 100, 1) if est else 0
        lines = ddata.get("lines", [])
        summary_rows[dept_name] = {
            "estimated": round(est, 2),
            "actual": round(act, 2),
            "variance": round(var, 2),
            "variance_pct": var_pct,
            "lines": lines,
        }

    # ── Department sheets ──
    for dept_name, dept_info in summary_rows.items():
        ws = wb.create_sheet(dept_name[:31])  # Excel max 31 chars for sheet name
        ws.append([f"{prod_name} - {dept_name} BUDGET"])
        ws.cell(row=1, column=1).font = title_font
        ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
        ws.append([])
        ws.append(["Entity", "Detail", "Daily Rate", "Working Days",
                    "Start", "End", "Estimate ($)", "Actual ($)", "Variance ($)"])
        style_header_row(ws, 9)

        dept_est = 0
        dept_act = 0
        for row in dept_info["lines"]:
            r_est = row.get("amount_estimate", 0) or 0
            r_act = row.get("amount_actual", 0) or 0
            r_var = r_act - r_est
            entity = row.get("name") or row.get("function_name") or ""
            detail = row.get("boat") or row.get("detail") or row.get("vehicle") or ""
            rate = row.get("price_override") or row.get("daily_rate") or row.get("daily_rate_estimate") or ""
            days = row.get("working_days") or ""
            start = row.get("start_date") or ""
            end = row.get("end_date") or ""
            ws.append([entity, detail, rate, days, start, end,
                        round(r_est, 2), round(r_act, 2) if r_act else "", round(r_var, 2) if r_act else ""])
            # Number format
            for c in (7, 8, 9):
                ws.cell(row=ws.max_row, column=c).number_format = money_fmt
            dept_est += r_est
            dept_act += r_act

        ws.append([])
        ws.append(["", "", "", "", "", "TOTAL",
                    round(dept_est, 2), round(dept_act, 2), round(dept_act - dept_est, 2)])
        for c in (7, 8):
            ws.cell(row=ws.max_row, column=c).font = green_font
            ws.cell(row=ws.max_row, column=c).number_format = money_fmt
        variance_font = red_font if (dept_act - dept_est) > 0 else green_font
        ws.cell(row=ws.max_row, column=9).font = variance_font
        ws.cell(row=ws.max_row, column=9).number_format = money_fmt
        auto_width(ws)

    # ── Summary sheet (insert at position 0) ──
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append([f"{prod_name} - BUDGET VARIANCE SUMMARY"])
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_summary.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws_summary.append([])
    ws_summary.append(["Department", "Estimate ($)", "Actual ($)", "Variance ($)", "Variance (%)"])
    style_header_row(ws_summary, 5)

    grand_est = 0
    grand_act = 0
    for dept_name, info in summary_rows.items():
        ws_summary.append([dept_name, info["estimated"], info["actual"],
                            info["variance"],
                            f"{info['variance_pct']:+.1f}%" if info["actual"] else ""])
        for c in (2, 3, 4):
            ws_summary.cell(row=ws_summary.max_row, column=c).number_format = money_fmt
        if info["variance"] > 0:
            ws_summary.cell(row=ws_summary.max_row, column=4).font = red_font
        elif info["variance"] < 0:
            ws_summary.cell(row=ws_summary.max_row, column=4).font = green_font
        grand_est += info["estimated"]
        grand_act += info["actual"]

    ws_summary.append([])
    grand_var = grand_act - grand_est
    grand_pct = round((grand_var / grand_est) * 100, 1) if grand_est else 0
    ws_summary.append(["GRAND TOTAL", round(grand_est, 2), round(grand_act, 2),
                        round(grand_var, 2), f"{grand_pct:+.1f}%"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True, size=12)
    ws_summary.cell(row=ws_summary.max_row, column=2).font = Font(bold=True, size=12, color="22C55E")
    ws_summary.cell(row=ws_summary.max_row, column=2).number_format = money_fmt
    ws_summary.cell(row=ws_summary.max_row, column=3).font = Font(bold=True, size=12)
    ws_summary.cell(row=ws_summary.max_row, column=3).number_format = money_fmt
    var_font = Font(bold=True, size=12, color="EF4444") if grand_var > 0 else Font(bold=True, size=12, color="22C55E")
    ws_summary.cell(row=ws_summary.max_row, column=4).font = var_font
    ws_summary.cell(row=ws_summary.max_row, column=4).number_format = money_fmt
    auto_width(ws_summary)

    # ── History sheet ──
    ws_hist = wb.create_sheet("History")
    ws_hist.append([f"{prod_name} - RECENT MODIFICATIONS"])
    ws_hist.cell(row=1, column=1).font = title_font
    ws_hist.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws_hist.append([])
    ws_hist.append(["Date", "User", "Action", "Module", "Description"])
    style_header_row(ws_hist, 5)

    history = get_history(prod_id, limit=100)
    for entry in history:
        ws_hist.append([
            entry.get("created_at", ""),
            entry.get("user_nickname", "") or entry.get("username", ""),
            entry.get("action", ""),
            entry.get("table_name", ""),
            entry.get("human_description", "") or entry.get("description", ""),
        ])
    auto_width(ws_hist)

    # Remove default empty sheet if still present
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"{prod_name}_BUDGET_{dt.now().strftime('%y%m%d')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ─── LOGISTICS EXPORT ─────────────────────────────────────────────────────────

def _is_date_active(date_str, asgn):
    """Check if date_str is an active working day for an assignment dict."""
    import json as _json
    start = asgn.get("start_date", "")
    end = asgn.get("end_date", "")
    try:
        overrides = _json.loads(asgn.get("day_overrides") or "{}")
    except Exception:
        overrides = {}
    include_sun = bool(asgn.get("include_sunday", 1))

    if date_str in overrides:
        return bool(overrides[date_str]) and overrides[date_str] != 'empty'

    if start and end and start <= date_str <= end:
        if not include_sun:
            from datetime import datetime as _dt
            if _dt.strptime(date_str, "%Y-%m-%d").weekday() == 6:
                return False
        return True
    return False


@app.route("/api/productions/<int:prod_id>/export/logistics")
def api_export_logistics(prod_id):
    """Export full logistics/scheduling data as a multi-sheet Excel file (.xlsx).
    Filename: KLAS7_LOGISTICS_YYMMDD.xlsx
    """
    from datetime import datetime as dt
    from collections import OrderedDict
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    prod_or_404(prod_id)

    wb = Workbook()

    # -- Style constants (same as budget export) --
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    title_font = Font(bold=True, size=12)
    subtotal_font = Font(bold=True, size=10)
    section_font = Font(bold=True, size=11, color="3B82F6")
    money_fmt = '#,##0'
    money_fmt_dec = '#,##0.00'
    green_font = Font(bold=True, color="22C55E")
    thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))

    fill_p = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # green
    fill_f = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")  # yellow
    fill_w = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # blue

    def style_header_row(ws, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

    # -- Load all data upfront (read-only) --
    shooting_days = get_shooting_days(prod_id)
    loc_sites = get_location_sites(prod_id)
    loc_schedules = get_location_schedules(prod_id)
    boat_rows = get_boat_assignments(prod_id, context='boats')
    pb_rows = get_picture_boat_assignments(prod_id)
    sb_rows = get_security_boat_assignments(prod_id)
    transport_rows = get_transport_assignments(prod_id)
    fuel_entries = get_fuel_entries(prod_id)
    fuel_machinery = get_fuel_machinery(prod_id)
    locked_prices = get_fuel_locked_prices()
    cur_diesel = float(get_setting("fuel_price_diesel", "0"))
    cur_petrol = float(get_setting("fuel_price_petrol", "0"))
    helper_rows = get_helper_assignments(prod_id)
    guard_loc_data = get_guard_location_schedules(prod_id)
    gc_rows = get_guard_camp_assignments(prod_id)
    fnb_cats = get_fnb_categories(prod_id)
    fnb_items = get_fnb_items(prod_id)
    fnb_entries = get_fnb_entries(prod_id)

    # Shooting day lookup
    date_to_day = {d.get('date'): d.get('day_number', '') for d in shooting_days}

    def _date_range(min_date, max_date):
        """Generate all dates from min_date to max_date inclusive (YYYY-MM-DD strings)."""
        from datetime import timedelta
        if not min_date or not max_date:
            return []
        start = dt.strptime(min_date, "%Y-%m-%d").date()
        end = dt.strptime(max_date, "%Y-%m-%d").date()
        return [(start + timedelta(days=i)).strftime("%Y-%m-%d") for i in range((end - start).days + 1)]

    def _fmt_date_header(d):
        """Format date header: 'D{num} MM-DD' for shooting days, just 'MM-DD' otherwise."""
        day_num = date_to_day.get(d)
        if day_num:
            return f"D{day_num} {d[5:]}"
        return d[5:]

    def _assignments_date_range(assignments):
        """Compute min/max date range from assignment start/end dates."""
        starts = [r['start_date'] for r in assignments if r.get('start_date') and r.get('working_days')]
        ends = [r['end_date'] for r in assignments if r.get('end_date') and r.get('working_days')]
        if not starts or not ends:
            return []
        return _date_range(min(starts), max(ends))

    # ── Sheet 1: PDT ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "PDT"
    ws.append(["KLAS 7 - SHOOTING SCHEDULE (PDT)"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Day #", "Date", "Status", "Location", "Game",
               "Rehearsal", "Animateur", "Game Time",
               "Candidats Depart", "Tide Height", "Tide Status",
               "Nb Candidats", "Reward", "Conseil", "Events", "Notes"])
    style_header_row(ws, 16)
    for d in shooting_days:
        evts = d.get('events', [])
        evt_str = ", ".join(f"{e.get('event_type','')}: {e.get('event_name','')}" for e in evts) if evts else ""
        ws.append([
            d.get('day_number', ''),
            d.get('date', ''),
            d.get('status', ''),
            d.get('location', ''),
            d.get('game_name', ''),
            d.get('heure_rehearsal', ''),
            d.get('heure_animateur', ''),
            d.get('heure_game', ''),
            d.get('heure_depart_candidats', ''),
            d.get('maree_hauteur', ''),
            d.get('maree_statut', ''),
            d.get('nb_candidats', ''),
            d.get('recompense', ''),
            d.get('conseil_soir', ''),
            evt_str,
            d.get('notes', ''),
        ])
    auto_width(ws)

    # ── Sheet 2: LOCATIONS (matrix) ──────────────────────────────────────────
    ws = wb.create_sheet("Locations")
    ws.append(["KLAS 7 - LOCATIONS SCHEDULE"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])

    # Build schedule matrix: {loc_name: {date: status}}
    loc_matrix = {}
    for ls in loc_schedules:
        loc_matrix.setdefault(ls['location_name'], {})[ls['date']] = ls.get('status', '')

    # Date range: first to last location schedule date
    loc_all_dates_set = sorted(set(ls['date'] for ls in loc_schedules if ls.get('date')))
    loc_dates = _date_range(loc_all_dates_set[0], loc_all_dates_set[-1]) if loc_all_dates_set else []

    # Header row: Location, Type, then one col per date
    header = ["Location", "Type"] + [_fmt_date_header(d) for d in loc_dates] + ["P", "F", "W", "Total"]
    ws.append(header)
    style_header_row(ws, len(header))
    header_row_num = ws.max_row

    loc_names = sorted(loc_matrix.keys())
    for loc_name in loc_names:
        site_info = next((s for s in loc_sites if s['name'] == loc_name), {})
        row_data = [loc_name, site_info.get('location_type', '')]
        p_count = f_count = w_count = 0
        for dt_str in loc_dates:
            status = loc_matrix.get(loc_name, {}).get(dt_str, '')
            row_data.append(status)
            if status == 'P': p_count += 1
            elif status == 'F': f_count += 1
            elif status == 'W': w_count += 1
        row_data += [p_count, f_count, w_count, p_count + f_count + w_count]
        ws.append(row_data)
        # Color cells
        row_num = ws.max_row
        for col_idx, dt_str in enumerate(loc_dates, start=3):
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.value == 'P': cell.fill = fill_p
            elif cell.value == 'F': cell.fill = fill_f
            elif cell.value == 'W': cell.fill = fill_w
            cell.alignment = Alignment(horizontal='center')

    # Totals row per date
    ws.append([])
    totals_row = ["TOTAL", ""]
    for dt_str in loc_dates:
        count = sum(1 for ln in loc_names if loc_matrix.get(ln, {}).get(dt_str, '') in ('P', 'F', 'W'))
        totals_row.append(count)
    totals_row += ["", "", "", ""]
    ws.append(totals_row)
    ws.cell(row=ws.max_row, column=1).font = subtotal_font

    if loc_dates:
        ws.freeze_panes = "C5"
    auto_width(ws)

    # ── Helper: build assignment matrix ────────────────────────────────────────
    fill_active = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    def _write_assignment_matrix(ws, title, assignments, label_fn):
        """Write a matrix sheet: rows=assignments, cols=dates, cells=1 if active.
        Date range = first start_date to last end_date of active assignments."""
        ws.append([f"KLAS 7 - {title}"])
        ws.cell(row=1, column=1).font = title_font
        ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
        ws.append([])

        # Compute per-module date range
        sheet_dates = _assignments_date_range(assignments)
        sheet_headers = [_fmt_date_header(d) for d in sheet_dates]

        header = ["Assignment"] + sheet_headers + ["Total"]
        ws.append(header)
        style_header_row(ws, len(header))

        matrix_rows = []
        for r in assignments:
            if not r.get("working_days"):
                continue
            label = label_fn(r)
            day_map = {}
            for d in sheet_dates:
                if _is_date_active(d, r):
                    day_map[d] = 1
            matrix_rows.append((label, day_map, r.get("working_days", 0)))

        for label, day_map, wd in matrix_rows:
            row = [label] + [day_map.get(d, "") for d in sheet_dates] + [wd]
            ws.append(row)
            row_num = ws.max_row
            for col_idx, d in enumerate(sheet_dates, start=2):
                cell = ws.cell(row=row_num, column=col_idx)
                if cell.value == 1:
                    cell.fill = fill_active
                    cell.alignment = Alignment(horizontal='center')

        # Totals row
        ws.append([])
        totals = ["TOTAL / DAY"]
        for d in sheet_dates:
            totals.append(sum(1 for _, dm, _ in matrix_rows if dm.get(d)))
        totals.append(sum(wd for _, _, wd in matrix_rows))
        ws.append(totals)
        ws.cell(row=ws.max_row, column=1).font = subtotal_font
        ws.cell(row=ws.max_row, column=len(sheet_dates) + 2).font = green_font

        if sheet_dates:
            ws.freeze_panes = "B5"
        auto_width(ws)
        return matrix_rows

    # ── Sheet 3: BOATS (matrix) ─────────────────────────────────────────────
    ws = wb.create_sheet("Boats")
    boat_matrix = _write_assignment_matrix(ws, "BOATS SCHEDULE", boat_rows,
        lambda r: f"{r.get('function_name','') or ''} — {r.get('boat_name_override') or r.get('boat_name') or ''}")

    # ── Sheet 4: PICTURE BOATS (matrix) ─────────────────────────────────────
    ws = wb.create_sheet("Picture Boats")
    pb_matrix = _write_assignment_matrix(ws, "PICTURE BOATS SCHEDULE", pb_rows,
        lambda r: f"{r.get('function_name','') or ''} — {r.get('boat_name_override') or r.get('boat_name') or ''}")

    # ── Sheet 5: SECURITY BOATS (matrix) ────────────────────────────────────
    ws = wb.create_sheet("Security Boats")
    sb_matrix = _write_assignment_matrix(ws, "SECURITY BOATS SCHEDULE", sb_rows,
        lambda r: f"{r.get('function_name','') or ''} — {r.get('boat_name_override') or r.get('boat_name') or ''}")

    # ── Sheet 6: TRANSPORT (matrix) ─────────────────────────────────────────
    ws = wb.create_sheet("Transport")
    tr_matrix = _write_assignment_matrix(ws, "TRANSPORT SCHEDULE", transport_rows,
        lambda r: f"{r.get('function_name','') or ''} — {r.get('vehicle_name_override') or r.get('vehicle_name') or ''}")

    # ── Sheet 7: FUEL (matrix: consumer × date → liters) ───────────────────
    ws = wb.create_sheet("Fuel")
    ws.append(["KLAS 7 - FUEL CONSUMPTION"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])

    # Build assignment name map for fuel
    asgn_map = {}
    for ctx, fetcher in [
        ('boats', lambda: boat_rows),
        ('picture_boats', lambda: pb_rows),
        ('security_boats', lambda: sb_rows),
        ('transport', lambda: transport_rows),
    ]:
        for a in fetcher():
            asgn_map[(ctx, a['id'])] = (
                a.get('boat_name_override') or a.get('boat_name') or
                a.get('vehicle_name_override') or a.get('vehicle_name') or '?',
                a.get('function_name') or '?'
            )
    machinery_names = {m['id']: m['name'] for m in fuel_machinery}

    # Build fuel matrix: {consumer_label: {date: liters}}
    fuel_matrix = {}
    for e in fuel_entries:
        if e['source_type'] == 'machinery':
            mid = e['assignment_id']
            label = f"MACHINERY — {machinery_names.get(mid, f'#{mid}')}"
        else:
            info = asgn_map.get((e['source_type'], e['assignment_id']), (f"#{e['assignment_id']}", '?'))
            label = f"{e['source_type'].upper()} — {info[0]}"
        liters = e.get('liters', 0) or 0
        date = e.get('date', '')
        fuel_matrix.setdefault(label, {})[date] = fuel_matrix.get(label, {}).get(date, 0) + liters

    # Fuel date range: first to last fuel entry date
    fuel_entry_dates = sorted(set(e['date'] for e in fuel_entries if e.get('date')))
    fuel_dates = _date_range(fuel_entry_dates[0], fuel_entry_dates[-1]) if fuel_entry_dates else []

    header = ["Consumer"] + [_fmt_date_header(d) for d in fuel_dates] + ["Total (L)"]
    ws.append(header)
    style_header_row(ws, len(header))

    fuel_grand_liters = 0
    fuel_consumers = sorted(fuel_matrix.keys())
    for label in fuel_consumers:
        day_data = fuel_matrix[label]
        total_l = sum(day_data.values())
        fuel_grand_liters += total_l
        row = [label]
        for d in fuel_dates:
            val = day_data.get(d, '')
            row.append(round(val, 1) if val else '')
        row.append(round(total_l, 1))
        ws.append(row)

    # Totals row
    ws.append([])
    totals = ["TOTAL / DAY"]
    for d in fuel_dates:
        day_total = sum(fuel_matrix.get(c, {}).get(d, 0) for c in fuel_consumers)
        totals.append(round(day_total, 1) if day_total else '')
    totals.append(round(fuel_grand_liters, 1))
    ws.append(totals)
    ws.cell(row=ws.max_row, column=1).font = subtotal_font
    ws.cell(row=ws.max_row, column=len(fuel_dates) + 2).font = green_font

    # Machinery reference
    if fuel_machinery:
        ws.append([])
        ws.append([])
        ws.append(["MACHINERY REFERENCE"])
        ws.cell(row=ws.max_row, column=1).font = section_font
        ws.append(["Name", "Fuel Type", "Start", "End", "Liters/Day"])
        style_header_row(ws, 5)
        for m in fuel_machinery:
            ws.append([m.get('name', ''), m.get('fuel_type', ''),
                       m.get('start_date', ''), m.get('end_date', ''),
                       m.get('liters_per_day', '')])

    if fuel_dates:
        ws.freeze_panes = "B5"
    auto_width(ws)

    # ── Sheet 8: LABOUR (matrix) ────────────────────────────────────────────
    ws = wb.create_sheet("Labor")
    lb_matrix = _write_assignment_matrix(ws, "LABOR SCHEDULE", helper_rows,
        lambda r: f"{r.get('function_name','') or ''} — {r.get('helper_name_override') or r.get('helper_name') or ''}")

    # ── Sheet 9: GUARDS (matrix + base camp) ─────────────────────────────────
    ws = wb.create_sheet("Guards")
    ws.append(["KLAS 7 - GUARDS SCHEDULE"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])

    # Section A: Location guards matrix
    ws.append(["LOCATION GUARDS"])
    ws.cell(row=ws.max_row, column=1).font = section_font

    guard_matrix = {}
    for gls in guard_loc_data:
        loc_name = gls['location_name']
        guard_matrix.setdefault(loc_name, {})[gls['date']] = gls.get('nb_guards', 1)

    # Guard location date range
    gl_date_set = sorted(set(gls['date'] for gls in guard_loc_data if gls.get('date')))
    gl_dates = _date_range(gl_date_set[0], gl_date_set[-1]) if gl_date_set else []

    g_header = ["Location"] + [_fmt_date_header(d) for d in gl_dates] + ["Total Guard-Days"]
    ws.append(g_header)
    style_header_row(ws, len(g_header))

    gl_locations = sorted(guard_matrix.keys())
    for loc_name in gl_locations:
        row_data = [loc_name]
        total_gd = 0
        for dt_str in gl_dates:
            nb = guard_matrix.get(loc_name, {}).get(dt_str, '')
            row_data.append(nb)
            if isinstance(nb, (int, float)):
                total_gd += nb
        row_data.append(total_gd)
        ws.append(row_data)

    # Totals per date
    ws.append([])
    totals_row = ["TOTAL / DAY"]
    for dt_str in gl_dates:
        count = sum(guard_matrix.get(ln, {}).get(dt_str, 0) for ln in gl_locations)
        totals_row.append(count)
    totals_row.append(sum(
        sum(v for v in guard_matrix.get(ln, {}).values() if isinstance(v, (int, float)))
        for ln in gl_locations
    ))
    ws.append(totals_row)
    ws.cell(row=ws.max_row, column=1).font = subtotal_font

    if gl_dates:
        ws.freeze_panes = "B6"

    # Section B: Base Camp Guards (matrix)
    ws.append([])
    ws.append([])
    ws.append(["BASE CAMP GUARDS"])
    ws.cell(row=ws.max_row, column=1).font = section_font

    # Base camp date range from assignments
    gc_dates = _assignments_date_range(gc_rows)
    gc_header = ["Guard"] + [_fmt_date_header(d) for d in gc_dates] + ["Total"]
    ws.append(gc_header)
    style_header_row(ws, len(gc_header))

    gc_matrix_rows = []
    for r in gc_rows:
        if not r.get("working_days"):
            continue
        label = f"{r.get('function_name','') or ''} — {r.get('helper_name_override') or r.get('helper_name') or ''}"
        day_map = {}
        for d in gc_dates:
            if _is_date_active(d, r):
                day_map[d] = 1
        gc_matrix_rows.append((label, day_map, r.get("working_days", 0)))

    for label, day_map, wd in gc_matrix_rows:
        row = [label] + [day_map.get(d, "") for d in gc_dates] + [wd]
        ws.append(row)
        row_num = ws.max_row
        for col_idx, d in enumerate(gc_dates, start=2):
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.value == 1:
                cell.fill = fill_active
                cell.alignment = Alignment(horizontal='center')

    # Base camp totals per date
    ws.append([])
    gc_totals = ["TOTAL / DAY"]
    for d in gc_dates:
        gc_totals.append(sum(1 for _, dm, _ in gc_matrix_rows if dm.get(d)))
    gc_totals.append(sum(wd for _, _, wd in gc_matrix_rows))
    ws.append(gc_totals)
    ws.cell(row=ws.max_row, column=1).font = subtotal_font

    auto_width(ws)

    # ── Sheet 10: FNB ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Catering")
    ws.append(["KLAS 7 - CATERING"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Category", "Item", "Unit", "Unit Price",
               "Purchased Qty", "Consumed Qty", "Purchase Cost", "Consumption Cost"])
    style_header_row(ws, 8)

    # Build item-level aggregation
    item_totals = {}
    for e in fnb_entries:
        iid = e.get('item_id')
        if iid is None:
            continue
        item_totals.setdefault(iid, {'purchase': 0, 'consumption': 0})
        etype = e.get('entry_type', '')
        qty = e.get('quantity', 0) or 0
        if etype in item_totals[iid]:
            item_totals[iid][etype] += qty

    # Group items by category
    cat_map = {c['id']: c['name'] for c in fnb_cats}
    grand_purchase_cost = 0
    grand_consumption_cost = 0
    for cat in fnb_cats:
        cat_items = [it for it in fnb_items if it.get('category_id') == cat['id']]
        for it in cat_items:
            iid = it['id']
            totals = item_totals.get(iid, {'purchase': 0, 'consumption': 0})
            unit_price = it.get('unit_price', 0) or 0
            p_cost = round(totals['purchase'] * unit_price, 2)
            c_cost = round(totals['consumption'] * unit_price, 2)
            grand_purchase_cost += p_cost
            grand_consumption_cost += c_cost
            ws.append([
                cat['name'],
                it.get('name', ''),
                it.get('unit', ''),
                unit_price,
                round(totals['purchase'], 2) if totals['purchase'] else "",
                round(totals['consumption'], 2) if totals['consumption'] else "",
                p_cost if p_cost else "",
                c_cost if c_cost else "",
            ])
    ws.append([])
    ws.append(["", "", "", "TOTAL", "", "", round(grand_purchase_cost, 2), round(grand_consumption_cost, 2)])
    ws.cell(row=ws.max_row, column=7).font = green_font
    auto_width(ws)

    # ── Insert Summary sheet at position 0 ───────────────────────────────────
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append(["KLAS 7 - LOGISTICS OVERVIEW"])
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_summary.append([f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}"])
    ws_summary.append([])

    # Date range from shooting days
    pdt_dates = sorted(set(d.get('date', '') for d in shooting_days if d.get('date')))
    if pdt_dates:
        ws_summary.append([f"Production dates: {pdt_dates[0]} to {pdt_dates[-1]}"])
    ws_summary.append([])

    ws_summary.append(["Module", "Items / Assignments", "Details"])
    style_header_row(ws_summary, 3)

    active_boats = len(boat_matrix)
    active_pb = len(pb_matrix)
    active_sb = len(sb_matrix)
    active_transport = len(tr_matrix)
    active_helpers = len(lb_matrix)
    active_gc = len([r for r in gc_rows if r.get("working_days")])

    ws_summary.append(["PDT", len(shooting_days), f"{len(shooting_days)} shooting days"])
    ws_summary.append(["LOCATIONS", len(loc_sites), f"{len(loc_names)} locations with schedules"])
    ws_summary.append(["BOATS", active_boats, f"{active_boats} active assignments"])
    ws_summary.append(["PICTURE BOATS", active_pb, f"{active_pb} active assignments"])
    ws_summary.append(["SECURITY BOATS", active_sb, f"{active_sb} active assignments"])
    ws_summary.append(["TRANSPORT", active_transport, f"{active_transport} active assignments"])
    ws_summary.append(["FUEL", len(fuel_entries), f"{round(fuel_grand_liters, 0)} total liters logged"])
    ws_summary.append(["LABOR", active_helpers, f"{active_helpers} active assignments"])
    ws_summary.append(["GUARDS", len(gl_locations), f"{len(gl_locations)} guard locations + {active_gc} base camp"])
    ws_summary.append(["CATERING", len(fnb_items), f"{len(fnb_cats)} categories, {len(fnb_items)} items"])

    auto_width(ws_summary)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"KLAS7_LOGISTICS_{dt.now().strftime('%y%m%d')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ─── Async Exports ───────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/export/budget-global/async", methods=["POST"])
def api_export_budget_global_async(prod_id):
    """Start budget XLSX export in background. Returns job_id for polling."""
    prod_or_404(prod_id)
    _cleanup_old_exports()

    # Capture date params from request to pass to background thread
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    qs = f"?from={date_from}&to={date_to}" if (date_from or date_to) else ""

    job_id = str(uuid.uuid4())[:8]
    _export_jobs[job_id] = {"status": "processing", "created_at": _time.time(), "path": None, "filename": None}

    def _do_export():
        try:
            # Call the sync export function within app context, passing date params
            with app.test_request_context(f"/api/productions/{prod_id}/export/budget-global{qs}"):
                resp = api_export_budget_global(prod_id)
                data = resp.get_data()
                fname = resp.headers.get("Content-Disposition", "").split("filename=")[-1] or "export.xlsx"
                fpath = os.path.join(_EXPORT_DIR, f"{job_id}.xlsx")
                with open(fpath, "wb") as f:
                    f.write(data)
                _export_jobs[job_id]["status"] = "done"
                _export_jobs[job_id]["path"] = fpath
                _export_jobs[job_id]["filename"] = fname
        except Exception as e:
            _export_jobs[job_id]["status"] = "error"
            _export_jobs[job_id]["error"] = str(e)

    t = threading.Thread(target=_do_export, daemon=True)
    t.start()
    return jsonify({"job_id": job_id, "status": "processing"}), 202


@app.route("/api/productions/<int:prod_id>/export/logistics/async", methods=["POST"])
def api_export_logistics_async(prod_id):
    """Start logistics XLSX export in background. Returns job_id for polling."""
    prod_or_404(prod_id)
    _cleanup_old_exports()

    # Capture date params from request to pass to background thread
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    qs = f"?from={date_from}&to={date_to}" if (date_from or date_to) else ""

    job_id = str(uuid.uuid4())[:8]
    _export_jobs[job_id] = {"status": "processing", "created_at": _time.time(), "path": None, "filename": None}

    def _do_export():
        try:
            with app.test_request_context(f"/api/productions/{prod_id}/export/logistics{qs}"):
                resp = api_export_logistics(prod_id)
                data = resp.get_data()
                fname = resp.headers.get("Content-Disposition", "").split("filename=")[-1] or "export.xlsx"
                fpath = os.path.join(_EXPORT_DIR, f"{job_id}.xlsx")
                with open(fpath, "wb") as f:
                    f.write(data)
                _export_jobs[job_id]["status"] = "done"
                _export_jobs[job_id]["path"] = fpath
                _export_jobs[job_id]["filename"] = fname
        except Exception as e:
            _export_jobs[job_id]["status"] = "error"
            _export_jobs[job_id]["error"] = str(e)

    t = threading.Thread(target=_do_export, daemon=True)
    t.start()
    return jsonify({"job_id": job_id, "status": "processing"}), 202


# ─── Export Preferences API (AXE 2.2) ─────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/export-defaults/<module>", methods=["GET"])
def api_get_export_defaults(prod_id, module):
    """Get smart export date defaults for a module.
    Priority: 1) last user export, 2) module data range, 3) production dates."""
    prod = prod_or_404(prod_id)
    user_id = getattr(g, 'user_id', None)

    # 1. Check user's last export preference
    if user_id:
        pref = get_export_preference(user_id, prod_id, module)
        if pref and pref.get('last_export_from'):
            return jsonify({
                "from": pref['last_export_from'],
                "to": pref['last_export_to'],
                "source": "last_export",
            })

    # 2. Get date range from module data
    module_range = get_module_date_range(prod_id, module)
    if module_range and module_range.get('first_date'):
        return jsonify({
            "from": module_range['first_date'],
            "to": module_range['last_date'],
            "source": "module_data",
        })

    # 3. Fallback to production dates
    return jsonify({
        "from": prod.get('start_date', ''),
        "to": prod.get('end_date', ''),
        "source": "production",
    })


@app.route("/api/productions/<int:prod_id>/export-defaults/<module>", methods=["POST"])
def api_save_export_defaults(prod_id, module):
    """Save the user's last export date range for a module."""
    prod_or_404(prod_id)
    user_id = getattr(g, 'user_id', None)
    if not user_id:
        return jsonify({"error": "Authentication required"}), 401
    data = request.json or {}
    save_export_preference(user_id, prod_id, module,
                           data.get('from', ''), data.get('to', ''))
    return jsonify({"ok": True})


# ─── PDF & Advanced Exports (AXE 2.3) ────────────────────────────────────────

def _pdf_header(page, prod_name, title, date_range=None):
    """Draw a standard PDF header with production name, title, and date range."""
    import fitz
    w = page.rect.width
    # Production name
    page.insert_text(fitz.Point(40, 40), prod_name.upper(),
                     fontsize=10, fontname="helv", color=(0.4, 0.4, 0.4))
    # Title
    page.insert_text(fitz.Point(40, 62), title,
                     fontsize=18, fontname="hebo", color=(0.1, 0.1, 0.1))
    # Generation date + date range
    from datetime import datetime as _dt
    gen_text = f"Generated: {_dt.now().strftime('%Y-%m-%d %H:%M')}"
    if date_range:
        gen_text += f"  |  Period: {date_range}"
    page.insert_text(fitz.Point(40, 82), gen_text,
                     fontsize=8, fontname="helv", color=(0.5, 0.5, 0.5))
    # Separator line
    page.draw_line(fitz.Point(40, 90), fitz.Point(w - 40, 90),
                   color=(0.8, 0.8, 0.8), width=0.5)
    return 100  # y position after header


def _pdf_table(page, y, headers, rows, col_widths, w_total,
               header_bg=(0.17, 0.17, 0.17), header_fg=(1, 1, 1),
               row_height=18, font_size=8):
    """Draw a table on a PDF page. Returns y position after the table."""
    import fitz
    x_start = 40
    # Header row
    x = x_start
    for i, h in enumerate(headers):
        cw = col_widths[i]
        rect = fitz.Rect(x, y, x + cw, y + row_height + 2)
        page.draw_rect(rect, color=None, fill=header_bg)
        page.insert_text(fitz.Point(x + 4, y + row_height - 4),
                         h, fontsize=font_size, fontname="hebo", color=header_fg)
        x += cw
    y += row_height + 2

    # Data rows
    for ri, row in enumerate(rows):
        bg = (0.96, 0.96, 0.96) if ri % 2 else (1, 1, 1)
        x = x_start
        for i, val in enumerate(row):
            cw = col_widths[i]
            rect = fitz.Rect(x, y, x + cw, y + row_height)
            page.draw_rect(rect, color=None, fill=bg)
            text = str(val) if val is not None else ""
            # Right-align numbers
            if isinstance(val, (int, float)):
                text = f"{val:,.0f}" if isinstance(val, (int, float)) and val == int(val) else f"{val:,.2f}"
                tw = fitz.get_text_length(text, fontname="helv", fontsize=font_size)
                page.insert_text(fitz.Point(x + cw - tw - 4, y + row_height - 4),
                                 text, fontsize=font_size, fontname="helv", color=(0.15, 0.15, 0.15))
            else:
                # Truncate if too long
                max_chars = int(cw / (font_size * 0.45))
                if len(text) > max_chars:
                    text = text[:max_chars - 1] + "…"
                page.insert_text(fitz.Point(x + 4, y + row_height - 4),
                                 text, fontsize=font_size, fontname="helv", color=(0.15, 0.15, 0.15))
            x += cw
        y += row_height
    return y


@app.route("/api/productions/<int:prod_id>/export/budget-pdf")
def api_export_budget_pdf(prod_id):
    """Export consolidated budget as a print-friendly PDF with logo, date, department breakdown."""
    import fitz
    from datetime import datetime as dt

    prod = prod_or_404(prod_id)
    date_from, date_to = _export_date_params()
    budget = get_budget(prod_id)
    prod_name = prod.get("name", "PRODUCTION")
    date_str = dt.now().strftime("%y%m%d")

    # Filter budget rows by date range if provided
    if date_from or date_to:
        budget["rows"] = _filter_assignments_by_date(budget["rows"], date_from, date_to)
        # Rebuild by_department from filtered rows
        by_dept = {}
        for r in budget["rows"]:
            dept = r.get("department") or r.get("dept_name") or "OTHER"
            if dept not in by_dept:
                by_dept[dept] = {"lines": [], "total_estimate": 0, "total_actual": 0}
            by_dept[dept]["lines"].append(r)
            by_dept[dept]["total_estimate"] += r.get("amount_estimate") or 0
            by_dept[dept]["total_actual"] += (r.get("amount_actual") or 0)
        budget["by_department"] = by_dept
        budget["grand_total_estimate"] = sum(d["total_estimate"] for d in by_dept.values())
        budget["grand_total_actual"] = sum(d["total_actual"] for d in by_dept.values())

    doc = fitz.open()
    page = doc.new_page(width=595, height=842)  # A4
    w = page.rect.width

    # Header
    range_str = f"{date_from} to {date_to}" if date_from and date_to else f"{prod.get('start_date', 'N/A')} to {prod.get('end_date', 'N/A')}"
    y = _pdf_header(page, prod_name, "CONSOLIDATED BUDGET", range_str)

    # Grand totals section
    grand_est = budget["grand_total_estimate"]
    grand_act = budget["grand_total_actual"]
    page.insert_text(fitz.Point(40, y + 5), "GRAND TOTAL ESTIMATE",
                     fontsize=9, fontname="hebo", color=(0.3, 0.3, 0.3))
    page.insert_text(fitz.Point(220, y + 5), f"${grand_est:,.0f}",
                     fontsize=14, fontname="hebo", color=(0.13, 0.55, 0.13))
    if grand_act > 0:
        page.insert_text(fitz.Point(350, y + 5), f"ACTUAL: ${grand_act:,.0f}",
                         fontsize=9, fontname="hebo", color=(0.2, 0.5, 0.8))
    y += 30

    # Summary table by department
    headers = ["Department", "Lines", "Estimate ($)", "Actual ($)", "Variance"]
    col_widths = [140, 50, 120, 120, 85]
    dept_rows = []
    for dept_name, dept_data in budget["by_department"].items():
        est = dept_data["total_estimate"]
        act = dept_data.get("total_actual", 0) or 0
        variance = f"{((act - est) / est * 100):+.1f}%" if est > 0 and act > 0 else "—"
        dept_rows.append([dept_name, len(dept_data["lines"]), est, act or "—", variance])

    y = _pdf_table(page, y, headers, dept_rows, col_widths, w)
    y += 15

    # Detailed breakdown per department
    detail_headers = ["Item", "Detail", "Days", "$/Day", "Total ($)"]
    detail_widths = [140, 140, 45, 80, 110]

    for dept_name, dept_data in budget["by_department"].items():
        # Check if we need a new page
        needed = 30 + len(dept_data["lines"]) * 18
        if y + needed > 790:
            page = doc.new_page(width=595, height=842)
            y = _pdf_header(page, prod_name, "CONSOLIDATED BUDGET (cont.)")

        # Department header
        page.insert_text(fitz.Point(40, y + 3), dept_name,
                         fontsize=10, fontname="hebo", color=(0.1, 0.1, 0.1))
        dept_total = dept_data["total_estimate"]
        total_text = f"${dept_total:,.0f}"
        tw = fitz.get_text_length(total_text, fontname="hebo", fontsize=10)
        page.insert_text(fitz.Point(w - 40 - tw, y + 3), total_text,
                         fontsize=10, fontname="hebo", color=(0.13, 0.55, 0.13))
        y += 18

        detail_rows = []
        for line in dept_data["lines"]:
            detail_rows.append([
                line.get("name", ""),
                line.get("boat", "") or line.get("detail", ""),
                line.get("working_days", ""),
                line.get("unit_price_estimate", 0),
                line.get("amount_estimate", 0),
            ])

        # Paginate rows if needed
        while detail_rows:
            space = int((790 - y) / 18) - 1  # rows that fit
            if space < 3:
                page = doc.new_page(width=595, height=842)
                y = _pdf_header(page, prod_name, "CONSOLIDATED BUDGET (cont.)")
                space = int((790 - y) / 18) - 1
            batch = detail_rows[:space]
            detail_rows = detail_rows[space:]
            y = _pdf_table(page, y, detail_headers, batch, detail_widths, w)

        y += 10

    # Footer on last page
    page.insert_text(fitz.Point(40, 820), f"ShootLogix — {prod_name} — Budget Report",
                     fontsize=7, fontname="helv", color=(0.6, 0.6, 0.6))

    buf = io.BytesIO()
    doc.save(buf)
    doc.close()
    buf.seek(0)

    fname = _export_fname(prod_name, "BUDGET", date_from, date_to, "pdf")
    return Response(
        buf.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@app.route("/api/productions/<int:prod_id>/export/daily-report-pdf")
def api_export_daily_report_pdf(prod_id):
    """Export Daily Report: one page per shooting day with all resources mobilized + cost total."""
    import fitz
    from datetime import datetime as dt

    prod = prod_or_404(prod_id)
    daily = get_daily_budget(prod_id)
    prod_name = prod.get("name", "PRODUCTION")
    date_str = dt.now().strftime("%y%m%d")

    # Optional date range filter from query params
    date_from = request.args.get("from")
    date_to = request.args.get("to")

    days = daily.get("days", [])
    if date_from:
        days = [d for d in days if d["date"] >= date_from]
    if date_to:
        days = [d for d in days if d["date"] <= date_to]

    if not days:
        return jsonify({"error": "No shooting days in selected range"}), 404

    # Pre-load all assignments for resource details
    boat_asgns = get_boat_assignments(prod_id, context='boats')
    pb_asgns = get_picture_boat_assignments(prod_id)
    sb_asgns = get_security_boat_assignments(prod_id)
    tr_asgns = get_transport_assignments(prod_id)
    lb_asgns = get_helper_assignments(prod_id)
    gc_asgns = get_guard_camp_assignments(prod_id)

    from database import _is_assignment_active_on

    doc = fitz.open()

    for day_info in days:
        page = doc.new_page(width=595, height=842)  # A4
        date = day_info["date"]
        day_type = day_info.get("day_type", "standard").upper()
        location = day_info.get("location", "")
        day_num = day_info.get("day_number", "")

        range_str = f"{date_from} to {date_to}" if date_from and date_to else None
        y = _pdf_header(page, prod_name, f"DAILY REPORT — Day {day_num}", range_str)

        # Day info bar
        page.insert_text(fitz.Point(40, y + 3), f"Date: {date}",
                         fontsize=10, fontname="hebo", color=(0.1, 0.1, 0.1))
        page.insert_text(fitz.Point(200, y + 3), f"Type: {day_type}",
                         fontsize=10, fontname="hebo",
                         color=(0.8, 0.2, 0.2) if day_type in ("GAME", "ARENA") else (0.3, 0.3, 0.3))
        if location:
            page.insert_text(fitz.Point(320, y + 3), f"Location: {location}",
                             fontsize=10, fontname="helv", color=(0.3, 0.3, 0.3))
        y += 22

        # Cost summary bar
        total = day_info.get("total", 0)
        page.insert_text(fitz.Point(40, y), f"DAILY TOTAL: ${total:,.0f}",
                         fontsize=12, fontname="hebo", color=(0.13, 0.55, 0.13))
        y += 20

        # Department cost breakdown
        dept_keys = [("boats", "Boats"), ("picture_boats", "Picture Boats"),
                     ("security_boats", "Security Boats"), ("transport", "Transport"),
                     ("labour", "Labor"), ("guards", "Guards"),
                     ("locations", "Locations"), ("fnb", "Catering"), ("fuel", "Fuel")]

        cost_headers = ["Department", "Cost ($)"]
        cost_widths = [250, 120]
        cost_rows = []
        for key, label in dept_keys:
            val = day_info.get(key, 0)
            if val > 0:
                cost_rows.append([label, val])

        if cost_rows:
            y = _pdf_table(page, y, cost_headers, cost_rows, cost_widths, 595)
            y += 15

        # Active resources detail (compact: name list instead of full table for performance)
        def _list_active_compact(title, assignments, name_key, rate_key):
            nonlocal y, page
            active = [a for a in assignments if _is_assignment_active_on(date, a)]
            if not active:
                return
            if y + 30 > 790:
                page = doc.new_page(width=595, height=842)
                y = _pdf_header(page, prod_name, f"DAILY REPORT — Day {day_num} (cont.)")

            total_rate = sum(a.get("price_override") or a.get(rate_key) or 0 for a in active)
            page.insert_text(fitz.Point(40, y + 3),
                             f"{title} ({len(active)}) — ${total_rate:,.0f}",
                             fontsize=9, fontname="hebo", color=(0.2, 0.2, 0.2))
            y += 14

            # Compact list: names on a single line, wrapped if needed
            names = []
            for a in active:
                n = a.get("boat_name_override") or a.get(name_key) or a.get("helper_name_override") or a.get("vehicle_name_override") or ""
                rate = a.get("price_override") or a.get(rate_key) or 0
                names.append(f"{n} (${rate:,.0f})")
            text = ", ".join(names)
            # Word-wrap at ~90 chars per line
            lines = []
            while text:
                if len(text) <= 90:
                    lines.append(text)
                    break
                idx = text.rfind(", ", 0, 90)
                if idx == -1:
                    idx = 90
                else:
                    idx += 2
                lines.append(text[:idx])
                text = text[idx:]
            for line in lines:
                if y + 12 > 790:
                    page = doc.new_page(width=595, height=842)
                    y = _pdf_header(page, prod_name, f"DAILY REPORT — Day {day_num} (cont.)")
                page.insert_text(fitz.Point(55, y + 3), line,
                                 fontsize=7, fontname="helv", color=(0.35, 0.35, 0.35))
                y += 11
            y += 4

        _list_active_compact("BOATS", boat_asgns, "boat_name", "boat_daily_rate_estimate")
        _list_active_compact("PICTURE BOATS", pb_asgns, "boat_name", "boat_daily_rate_estimate")
        _list_active_compact("SECURITY BOATS", sb_asgns, "boat_name", "boat_daily_rate_estimate")
        _list_active_compact("TRANSPORT", tr_asgns, "vehicle_name", "vehicle_daily_rate_estimate")
        _list_active_compact("LABOUR", lb_asgns, "helper_name", "helper_daily_rate_estimate")
        _list_active_compact("GUARDS (BASE CAMP)", gc_asgns, "helper_name", "helper_daily_rate_estimate")

        # Footer
        page.insert_text(fitz.Point(40, 820),
                         f"ShootLogix — {prod_name} — Daily Report {date}",
                         fontsize=7, fontname="helv", color=(0.6, 0.6, 0.6))

    buf = io.BytesIO()
    doc.save(buf)
    doc.close()
    buf.seek(0)

    range_suffix = f"_{date_from}_{date_to}".replace("-", "") if date_from and date_to else ""
    fname = f"{prod_name}_DAILY_REPORT_{date_str}{range_suffix}.pdf"
    return Response(
        buf.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@app.route("/api/productions/<int:prod_id>/export/vendor-summary")
def api_export_vendor_summary(prod_id):
    """Export Vendor Summary: aggregation of costs by vendor over selected date range.
    Returns CSV with vendor, department, lines count, total estimate, total actual.
    """
    from datetime import datetime as dt

    prod = prod_or_404(prod_id)
    budget = get_budget(prod_id)
    prod_name = prod.get("name", "PRODUCTION")
    date_str = dt.now().strftime("%y%m%d")

    # Optional date range filter
    date_from = request.args.get("from")
    date_to = request.args.get("to")

    rows = budget["rows"]

    # Filter by date range if provided
    if date_from or date_to:
        filtered = []
        for r in rows:
            start = (r.get("start_date") or "")[:10]
            end = (r.get("end_date") or "")[:10]
            if not start and not end:
                # FNB/FUEL with no dates — always include
                filtered.append(r)
                continue
            if date_from and end and end < date_from:
                continue
            if date_to and start and start > date_to:
                continue
            filtered.append(r)
        rows = filtered

    # Aggregate by vendor
    vendor_data = {}
    for r in rows:
        vendor = r.get("vendor") or "—"
        dept = r.get("department") or r.get("dept_name") or "OTHER"
        key = (vendor, dept)
        if key not in vendor_data:
            vendor_data[key] = {"vendor": vendor, "department": dept, "lines": 0,
                                "total_estimate": 0, "total_actual": 0}
        vendor_data[key]["lines"] += 1
        vendor_data[key]["total_estimate"] += r.get("amount_estimate") or 0
        vendor_data[key]["total_actual"] += r.get("amount_actual") or 0

    # Sort by vendor then department
    sorted_data = sorted(vendor_data.values(), key=lambda x: (x["vendor"], x["department"]))

    # Build CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Vendor", "Department", "Lines", "Total Estimate ($)", "Total Actual ($)"])
    for d in sorted_data:
        writer.writerow([d["vendor"], d["department"], d["lines"],
                         round(d["total_estimate"], 2), round(d["total_actual"], 2)])

    # Grand totals row
    writer.writerow([])
    writer.writerow(["GRAND TOTAL", "", sum(d["lines"] for d in sorted_data),
                     round(sum(d["total_estimate"] for d in sorted_data), 2),
                     round(sum(d["total_actual"] for d in sorted_data), 2)])

    # Vendor totals
    writer.writerow([])
    writer.writerow(["--- VENDOR TOTALS ---"])
    vendor_totals = {}
    for d in sorted_data:
        v = d["vendor"]
        vendor_totals.setdefault(v, {"estimate": 0, "actual": 0})
        vendor_totals[v]["estimate"] += d["total_estimate"]
        vendor_totals[v]["actual"] += d["total_actual"]
    for v in sorted(vendor_totals.keys()):
        writer.writerow([v, "ALL", "",
                         round(vendor_totals[v]["estimate"], 2),
                         round(vendor_totals[v]["actual"], 2)])

    range_suffix = ""
    if date_from and date_to:
        range_suffix = f"_{date_from}_{date_to}".replace("-", "")
    fname = f"{prod_name}_VENDOR_SUMMARY_{date_str}{range_suffix}.csv"

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@app.route("/api/productions/<int:prod_id>/export/vendor-summary-pdf")
def api_export_vendor_summary_pdf(prod_id):
    """Export Vendor Summary as print-friendly PDF."""
    import fitz
    from datetime import datetime as dt

    prod = prod_or_404(prod_id)
    budget = get_budget(prod_id)
    prod_name = prod.get("name", "PRODUCTION")
    date_str = dt.now().strftime("%y%m%d")

    date_from = request.args.get("from")
    date_to = request.args.get("to")

    rows = budget["rows"]
    if date_from or date_to:
        filtered = []
        for r in rows:
            start = (r.get("start_date") or "")[:10]
            end = (r.get("end_date") or "")[:10]
            if not start and not end:
                filtered.append(r)
                continue
            if date_from and end and end < date_from:
                continue
            if date_to and start and start > date_to:
                continue
            filtered.append(r)
        rows = filtered

    # Aggregate by vendor
    vendor_data = {}
    for r in rows:
        vendor = r.get("vendor") or "—"
        vendor_data.setdefault(vendor, {"lines": [], "total_estimate": 0, "total_actual": 0})
        vendor_data[vendor]["lines"].append(r)
        vendor_data[vendor]["total_estimate"] += r.get("amount_estimate") or 0
        vendor_data[vendor]["total_actual"] += r.get("amount_actual") or 0

    # Sort vendors by total descending
    sorted_vendors = sorted(vendor_data.items(), key=lambda x: -x[1]["total_estimate"])

    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    range_str = f"{date_from} to {date_to}" if date_from and date_to else None
    y = _pdf_header(page, prod_name, "VENDOR SUMMARY", range_str)

    # Grand total
    grand = sum(v["total_estimate"] for _, v in sorted_vendors)
    page.insert_text(fitz.Point(40, y + 3), f"TOTAL ALL VENDORS: ${grand:,.0f}",
                     fontsize=12, fontname="hebo", color=(0.13, 0.55, 0.13))
    y += 25

    # Summary table
    sum_headers = ["Vendor", "Lines", "Estimate ($)", "Actual ($)", "Share (%)"]
    sum_widths = [180, 45, 110, 110, 70]
    sum_rows = []
    for vendor, vdata in sorted_vendors:
        share = f"{vdata['total_estimate'] / grand * 100:.1f}%" if grand > 0 else "—"
        sum_rows.append([vendor, len(vdata["lines"]), vdata["total_estimate"],
                         vdata["total_actual"] or "—", share])
    y = _pdf_table(page, y, sum_headers, sum_rows, sum_widths, 595)
    y += 20

    # Detail per vendor
    detail_headers = ["Dept", "Item", "Detail", "Days", "Total ($)"]
    detail_widths = [90, 130, 120, 40, 100]

    for vendor, vdata in sorted_vendors:
        needed = 30 + len(vdata["lines"]) * 18
        if y + min(needed, 100) > 790:
            page = doc.new_page(width=595, height=842)
            y = _pdf_header(page, prod_name, "VENDOR SUMMARY (cont.)")

        page.insert_text(fitz.Point(40, y + 3), vendor,
                         fontsize=10, fontname="hebo", color=(0.1, 0.1, 0.1))
        vt = f"${vdata['total_estimate']:,.0f}"
        tw = fitz.get_text_length(vt, fontname="hebo", fontsize=10)
        page.insert_text(fitz.Point(555 - tw, y + 3), vt,
                         fontsize=10, fontname="hebo", color=(0.13, 0.55, 0.13))
        y += 18

        d_rows = []
        for line in vdata["lines"]:
            d_rows.append([
                line.get("department") or line.get("dept_name", ""),
                line.get("name", ""),
                line.get("boat", "") or line.get("detail", ""),
                line.get("working_days", ""),
                line.get("amount_estimate", 0),
            ])

        while d_rows:
            space = int((790 - y) / 18) - 1
            if space < 3:
                page = doc.new_page(width=595, height=842)
                y = _pdf_header(page, prod_name, "VENDOR SUMMARY (cont.)")
                space = int((790 - y) / 18) - 1
            batch = d_rows[:space]
            d_rows = d_rows[space:]
            y = _pdf_table(page, y, detail_headers, batch, detail_widths, 595, font_size=7)
        y += 10

    # Footer
    page.insert_text(fitz.Point(40, 820),
                     f"ShootLogix — {prod_name} — Vendor Summary",
                     fontsize=7, fontname="helv", color=(0.6, 0.6, 0.6))

    buf = io.BytesIO()
    doc.save(buf)
    doc.close()
    buf.seek(0)

    range_suffix = f"_{date_from}_{date_to}".replace("-", "") if date_from and date_to else ""
    fname = f"{prod_name}_VENDOR_SUMMARY_{date_str}{range_suffix}.pdf"
    return Response(
        buf.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ─── Today View (P3.2) ────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/today", methods=["GET"])
def api_today(prod_id):
    """Return all active resources for a given date (defaults to today)."""
    prod_or_404(prod_id)
    from datetime import datetime as dt

    target_date = request.args.get("date") or dt.now().strftime("%Y-%m-%d")

    # ── 1. PDT / Schedule info for the day ──
    shooting_days = get_shooting_days(prod_id)
    day_info = None
    for d in shooting_days:
        if d.get("date") == target_date:
            day_info = d
            break

    schedule = None
    if day_info:
        schedule = {
            "day_number": day_info.get("day_number"),
            "date": day_info.get("date"),
            "location": day_info.get("location"),
            "game_name": day_info.get("game_name"),
            "status": day_info.get("status"),
            "heure_rehearsal": day_info.get("heure_rehearsal"),
            "heure_animateur": day_info.get("heure_animateur"),
            "heure_game": day_info.get("heure_game"),
            "heure_depart_candidats": day_info.get("heure_depart_candidats"),
            "maree_hauteur": day_info.get("maree_hauteur"),
            "maree_statut": day_info.get("maree_statut"),
            "nb_candidats": day_info.get("nb_candidats"),
            "conseil_soir": day_info.get("conseil_soir"),
            "notes": day_info.get("notes"),
            "events": day_info.get("events", []),
        }

    # Helper: check if an assignment covers a date (respecting day_overrides)
    def _is_active_on(asgn, date_str):
        start = asgn.get("start_date", "")
        end = asgn.get("end_date", "")
        if not start or not end:
            return False
        overrides = {}
        raw = asgn.get("day_overrides")
        if raw:
            try:
                overrides = json.loads(raw) if isinstance(raw, str) else raw
            except Exception:
                pass
        # If explicitly overridden on this date
        if date_str in overrides:
            return overrides[date_str] not in ("empty", "", None)
        # Otherwise check date range
        return start <= date_str <= end

    # ── 2. Boats (main boats) ──
    boat_assignments = get_boat_assignments(prod_id, context='boats')
    boats_today = []
    for a in boat_assignments:
        if _is_active_on(a, target_date):
            boats_today.append({
                "id": a.get("id"),
                "boat_name": a.get("boat_name") or a.get("boat_name_override", ""),
                "function_name": a.get("function_name"),
                "function_group": a.get("function_group"),
                "captain": a.get("captain"),
                "capacity": a.get("boat_capacity"),
                "status": a.get("assignment_status"),
                "wave_rating": a.get("wave_rating"),
                "image_path": a.get("image_path"),
                "color": a.get("color"),
                "notes": a.get("notes"),
            })

    # ── 3. Picture Boats ──
    pb_assignments = get_picture_boat_assignments(prod_id)
    picture_boats_today = []
    for a in pb_assignments:
        if _is_active_on(a, target_date):
            picture_boats_today.append({
                "id": a.get("id"),
                "boat_name": a.get("boat_name") or a.get("boat_name_override", ""),
                "function_name": a.get("function_name"),
                "function_group": a.get("function_group"),
                "status": a.get("assignment_status"),
                "image_path": a.get("image_path"),
                "color": a.get("color"),
                "notes": a.get("notes"),
            })

    # ── 4. Security Boats ──
    sb_assignments = get_security_boat_assignments(prod_id)
    security_boats_today = []
    for a in sb_assignments:
        if _is_active_on(a, target_date):
            security_boats_today.append({
                "id": a.get("id"),
                "boat_name": a.get("boat_name") or a.get("boat_name_override", ""),
                "function_name": a.get("function_name"),
                "function_group": a.get("function_group"),
                "status": a.get("assignment_status"),
                "image_path": a.get("image_path"),
                "color": a.get("color"),
                "notes": a.get("notes"),
            })

    # ── 5. Transport ──
    tr_assignments = get_transport_assignments(prod_id)
    transport_today = []
    for a in tr_assignments:
        if _is_active_on(a, target_date):
            transport_today.append({
                "id": a.get("id"),
                "vehicle_name": a.get("vehicle_name") or a.get("vehicle_name_override", ""),
                "vehicle_type": a.get("vehicle_type"),
                "function_name": a.get("function_name"),
                "function_group": a.get("function_group"),
                "status": a.get("assignment_status"),
                "color": a.get("color"),
                "notes": a.get("notes"),
            })

    # ── 6. Labour (helpers) ──
    lb_assignments = get_helper_assignments(prod_id)
    labour_today = []
    for a in lb_assignments:
        if _is_active_on(a, target_date):
            labour_today.append({
                "id": a.get("id"),
                "helper_name": a.get("helper_name") or a.get("helper_name_override", ""),
                "role": a.get("helper_role"),
                "function_name": a.get("function_name"),
                "function_group": a.get("function_group"),
                "status": a.get("assignment_status"),
                "color": a.get("color"),
                "notes": a.get("notes"),
            })

    # ── 7. Guards (camp) ──
    gc_assignments = get_guard_camp_assignments(prod_id)
    guards_today = []
    for a in gc_assignments:
        if _is_active_on(a, target_date):
            guards_today.append({
                "id": a.get("id"),
                "helper_name": a.get("helper_name") or a.get("helper_name_override", ""),
                "role": a.get("helper_role"),
                "function_name": a.get("function_name"),
                "function_group": a.get("function_group"),
                "status": a.get("assignment_status"),
                "color": a.get("color"),
                "notes": a.get("notes"),
            })

    # ── 8. Fuel for the day ──
    fuel_entries = get_fuel_entries(prod_id)
    fuel_today = [e for e in fuel_entries if e.get("date") == target_date]
    cur_diesel = float(get_setting("fuel_price_diesel", "0"))
    cur_petrol = float(get_setting("fuel_price_petrol", "0"))
    locked_prices = get_fuel_locked_prices()
    fuel_liters = sum(e.get("liters", 0) or 0 for e in fuel_today)
    fuel_cost = 0
    for e in fuel_today:
        liters = e.get("liters", 0) or 0
        ft = e.get("fuel_type", "DIESEL")
        if target_date in locked_prices:
            price = locked_prices[target_date]["diesel_price"] if ft == "DIESEL" else locked_prices[target_date]["petrol_price"]
        else:
            price = cur_diesel if ft == "DIESEL" else cur_petrol
        fuel_cost += liters * price

    fuel_summary = {
        "entries": len(fuel_today),
        "total_liters": round(fuel_liters, 1),
        "total_cost": round(fuel_cost, 2),
        "diesel_price": cur_diesel,
        "petrol_price": cur_petrol,
    }

    # ── 9. Location schedules for the day ──
    loc_schedules = get_location_schedules(prod_id)
    locations_today = [
        {"location_name": ls["location_name"], "status": ls.get("status"), "notes": ls.get("notes")}
        for ls in loc_schedules if ls.get("date") == target_date
    ]

    return jsonify({
        "date": target_date,
        "schedule": schedule,
        "boats": boats_today,
        "picture_boats": picture_boats_today,
        "security_boats": security_boats_today,
        "transport": transport_today,
        "labour": labour_today,
        "guards": guards_today,
        "fuel": fuel_summary,
        "locations": locations_today,
        "counts": {
            "boats": len(boats_today),
            "picture_boats": len(picture_boats_today),
            "security_boats": len(security_boats_today),
            "transport": len(transport_today),
            "labour": len(labour_today),
            "guards": len(guards_today),
            "fleet_total": len(boats_today) + len(picture_boats_today) + len(security_boats_today),
            "crew_total": len(labour_today) + len(guards_today),
        },
    })


# ─── Dashboard ───────────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/dashboard", methods=["GET"])
def api_dashboard(prod_id):
    """Return budget summary, KPIs, and alerts for the dashboard."""
    prod_or_404(prod_id)
    from datetime import datetime as dt

    shooting_days = get_shooting_days(prod_id)
    budget = get_budget(prod_id)

    # Gather all department totals
    departments = {}

    # Boats
    boat_rows = [r for r in get_boat_assignments(prod_id, context='boats') if r.get("working_days")]
    departments["boats"] = {
        "estimate": sum(r.get("amount_estimate") or 0 for r in boat_rows),
        "actual": sum(r.get("amount_actual") or 0 for r in boat_rows),
        "count": len(boat_rows),
    }

    # Picture boats
    pb_rows = [r for r in get_picture_boat_assignments(prod_id) if r.get("working_days")]
    departments["picture_boats"] = {
        "estimate": sum(r.get("amount_estimate") or 0 for r in pb_rows),
        "actual": sum(r.get("amount_actual") or 0 for r in pb_rows),
        "count": len(pb_rows),
    }

    # Security boats
    sb_rows = [r for r in get_security_boat_assignments(prod_id) if r.get("working_days")]
    departments["security_boats"] = {
        "estimate": sum(r.get("amount_estimate") or 0 for r in sb_rows),
        "actual": sum(r.get("amount_actual") or 0 for r in sb_rows),
        "count": len(sb_rows),
    }

    # Transport
    tr_rows = [r for r in get_transport_assignments(prod_id) if r.get("working_days")]
    departments["transport"] = {
        "estimate": sum(r.get("amount_estimate") or 0 for r in tr_rows),
        "actual": sum(r.get("amount_actual") or 0 for r in tr_rows),
        "count": len(tr_rows),
    }

    # Labour
    lb_rows = [r for r in get_helper_assignments(prod_id) if r.get("working_days")]
    departments["labour"] = {
        "estimate": sum(r.get("amount_estimate") or 0 for r in lb_rows),
        "actual": sum(r.get("amount_actual") or 0 for r in lb_rows),
        "count": len(lb_rows),
    }

    # Guards (base camp)
    gc_rows = get_guard_camp_assignments(prod_id)
    gc_active = [r for r in gc_rows if r.get("working_days")]
    departments["guards"] = {
        "estimate": sum(r.get("amount_estimate") or 0 for r in gc_active),
        "actual": sum(r.get("amount_actual") or 0 for r in gc_active),
        "count": len(gc_active),
    }

    # Fuel
    fuel_entries = get_fuel_entries(prod_id)
    cur_diesel = float(get_setting("fuel_price_diesel", "0"))
    cur_petrol = float(get_setting("fuel_price_petrol", "0"))
    locked_prices = get_fuel_locked_prices()
    fuel_total = 0
    fuel_liters = 0
    for e in fuel_entries:
        liters = e.get("liters", 0) or 0
        fuel_liters += liters
        ft = e.get("fuel_type", "DIESEL")
        date = e.get("date", "")
        if date in locked_prices:
            price = locked_prices[date]["diesel_price"] if ft == "DIESEL" else locked_prices[date]["petrol_price"]
        else:
            price = cur_diesel if ft == "DIESEL" else cur_petrol
        fuel_total += liters * price
    departments["fuel"] = {
        "estimate": fuel_total,
        "actual": fuel_total,
        "count": len(fuel_entries),
        "liters": round(fuel_liters, 0),
    }

    # Locations
    loc_schedules = get_location_schedules(prod_id)
    loc_sites = get_location_sites(prod_id)
    site_pricing = {}
    for s in loc_sites:
        site_pricing[s["name"]] = {
            "price_p": s.get("price_p") or 0,
            "price_f": s.get("price_f") or 0,
            "price_w": s.get("price_w") or 0,
            "global_deal": s.get("global_deal"),
        }
    loc_day_counts = {}
    for ls in loc_schedules:
        loc_name = ls["location_name"]
        loc_day_counts.setdefault(loc_name, {"P": 0, "F": 0, "W": 0})
        if ls["status"] in ("P", "F", "W"):
            loc_day_counts[loc_name][ls["status"]] += 1
    loc_total = 0
    for loc_name, counts in loc_day_counts.items():
        pricing = site_pricing.get(loc_name, {"price_p": 0, "price_f": 0, "price_w": 0, "global_deal": None})
        if pricing["global_deal"] and pricing["global_deal"] > 0:
            loc_total += pricing["global_deal"]
        else:
            loc_total += (counts["P"] * pricing["price_p"] +
                         counts["F"] * pricing["price_f"] +
                         counts["W"] * pricing["price_w"])
    departments["locations"] = {
        "estimate": loc_total,
        "actual": loc_total,
        "count": len(loc_sites),
    }

    # FNB
    fnb_budget = get_fnb_budget_data(prod_id)
    fnb_total = 0
    for cat in fnb_budget.get("categories", []):
        fnb_total += (cat.get("consumption_total", 0) or 0) + (cat.get("purchase_total", 0) or 0)
    departments["fnb"] = {
        "estimate": fnb_total,
        "actual": fnb_total,
        "count": len(fnb_budget.get("categories", [])),
    }

    # Compute grand totals
    total_estimate = sum(d["estimate"] for d in departments.values())
    total_actual = sum(d["actual"] for d in departments.values())

    # Compute KPIs
    pdt_dates = sorted(set(d.get("date", "") for d in shooting_days if d.get("date")))
    days_elapsed = 0
    days_remaining = 0
    today = dt.now().strftime("%Y-%m-%d")
    if pdt_dates:
        days_elapsed = sum(1 for d in pdt_dates if d <= today)
        days_remaining = sum(1 for d in pdt_dates if d > today)

    burn_rate = total_actual / max(days_elapsed, 1)

    # Variance % per department
    for dept_name, dept_data in departments.items():
        est = dept_data["estimate"]
        act = dept_data["actual"]
        if est > 0:
            dept_data["variance_pct"] = round((act - est) / est * 100, 1)
            dept_data["usage_pct"] = round(act / est * 100, 1)
        else:
            dept_data["variance_pct"] = 0.0
            dept_data["usage_pct"] = 0.0

    # Alerts (75% = caution, 90% = warning, 100%+ = over_budget)
    alerts = []
    for dept_name, dept_data in departments.items():
        est = dept_data["estimate"]
        act = dept_data["actual"]
        if est > 0 and act > 0:
            pct = round(act / est * 100)
            if pct >= 100:
                alerts.append({
                    "type": "over_budget",
                    "dept": dept_name,
                    "pct": pct,
                    "msg": f"{dept_name.replace('_', ' ').title()} is at {pct}% of estimate"
                })
            elif pct >= 90:
                alerts.append({
                    "type": "warning",
                    "dept": dept_name,
                    "pct": pct,
                    "msg": f"{dept_name.replace('_', ' ').title()} approaching budget ({pct}%)"
                })
            elif pct >= 75:
                alerts.append({
                    "type": "caution",
                    "dept": dept_name,
                    "pct": pct,
                    "msg": f"{dept_name.replace('_', ' ').title()} at {pct}% of budget"
                })

    # Cumulative burn rate data per shooting day (for burn rate chart)
    burn_data = []
    if pdt_dates and total_actual > 0:
        elapsed_dates = [d for d in pdt_dates if d <= today]
        if elapsed_dates:
            daily_rate = total_actual / len(elapsed_dates)
            cumulative = 0
            for i, d in enumerate(pdt_dates):
                if d <= today:
                    cumulative += daily_rate
                    burn_data.append({"date": d, "cumulative": round(cumulative, 2), "is_actual": True})
                else:
                    # Linear projection from current burn rate
                    cumulative += daily_rate
                    burn_data.append({"date": d, "cumulative": round(cumulative, 2), "is_actual": False})

    # Next arena day
    next_arena = None
    for day in shooting_days:
        if day.get("date", "") > today:
            # Check if it has arena events
            events = get_events_for_day(day["id"])
            for ev in events:
                if ev.get("event_type") == "arena":
                    next_arena = day["date"]
                    break
            if next_arena:
                break

    return jsonify_cached({
        "departments": departments,
        "total_estimate": round(total_estimate, 2),
        "total_actual": round(total_actual, 2),
        "kpis": {
            "shooting_days_total": len(shooting_days),
            "days_elapsed": days_elapsed,
            "days_remaining": days_remaining,
            "burn_rate_per_day": round(burn_rate, 2),
            "projected_total": round(burn_rate * (days_elapsed + days_remaining), 2),
            "next_arena": next_arena,
            "fuel_liters": round(fuel_liters, 0),
        },
        "alerts": alerts,
        "burn_data": burn_data,
    })


# ─── Executive Dashboard KPIs (P5.1) ──────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/dashboard/kpis", methods=["GET"])
def api_dashboard_kpis(prod_id):
    """Return executive KPIs: fleet coverage, crew coverage, unconfirmed assignments, breakdowns."""
    prod_or_404(prod_id)
    from datetime import datetime as dt, timedelta

    from database import _is_assignment_active_on

    today = dt.now().strftime("%Y-%m-%d")
    horizon = [(dt.now() + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(3)]

    # All fleet assignments (boats + picture boats + security boats + transport)
    boat_assigns = get_boat_assignments(prod_id, context='boats')
    pb_assigns = get_picture_boat_assignments(prod_id)
    sb_assigns = get_security_boat_assignments(prod_id)
    tr_assigns = get_transport_assignments(prod_id)
    fleet_all = boat_assigns + pb_assigns + sb_assigns + tr_assigns

    # Fleet coverage: % of functions with a vessel/vehicle assigned for next 3 days
    functions_needing = 0
    functions_covered = 0
    for a in fleet_all:
        for d in horizon:
            if _is_assignment_active_on(d, a):
                functions_needing += 1
                has_asset = (a.get("boat_id") or a.get("picture_boat_id") or
                             a.get("security_boat_id") or a.get("vehicle_id"))
                if has_asset and a.get("assignment_status") != "breakdown":
                    functions_covered += 1
                break
    fleet_coverage = round(functions_covered / max(functions_needing, 1) * 100, 1)

    # Crew coverage: % of guard + labour posts filled
    helper_assigns = get_helper_assignments(prod_id)
    guard_assigns = get_guard_camp_assignments(prod_id)
    crew_all = helper_assigns + guard_assigns

    crew_needing = 0
    crew_covered = 0
    for a in crew_all:
        for d in horizon:
            if _is_assignment_active_on(d, a):
                crew_needing += 1
                has_person = a.get("helper_id") or a.get("helper_name_override")
                if has_person and a.get("assignment_status") not in ("estimate", "off"):
                    crew_covered += 1
                break
    crew_coverage = round(crew_covered / max(crew_needing, 1) * 100, 1)

    # Unconfirmed at J-2 (assignments still 'estimate' within 2 days)
    j2 = [(dt.now() + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(3)]
    all_assigns = fleet_all + crew_all
    unconfirmed = 0
    for a in all_assigns:
        if a.get("assignment_status") == "estimate":
            for d in j2:
                if _is_assignment_active_on(d, a):
                    unconfirmed += 1
                    break

    # Breakdowns: fleet items with status 'breakdown'
    breakdowns = sum(1 for a in fleet_all if a.get("assignment_status") == "breakdown")

    return jsonify({
        "fleet_coverage": fleet_coverage,
        "crew_coverage": crew_coverage,
        "unconfirmed_assignments": unconfirmed,
        "breakdowns": breakdowns,
    })


@app.route("/api/productions/<int:prod_id>/dashboard/alerts", methods=["GET"])
def api_dashboard_alerts(prod_id):
    """Return executive alerts: unconfirmed at J-2, breakdowns, budget overruns >10%."""
    prod_or_404(prod_id)
    from datetime import datetime as dt, timedelta

    from database import _is_assignment_active_on

    today = dt.now().strftime("%Y-%m-%d")
    j2 = [(dt.now() + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(3)]
    alerts = []

    # All assignments
    boat_assigns = get_boat_assignments(prod_id, context='boats')
    pb_assigns = get_picture_boat_assignments(prod_id)
    sb_assigns = get_security_boat_assignments(prod_id)
    tr_assigns = get_transport_assignments(prod_id)
    helper_assigns = get_helper_assignments(prod_id)
    guard_assigns = get_guard_camp_assignments(prod_id)
    fleet_all = boat_assigns + pb_assigns + sb_assigns + tr_assigns
    all_assigns = fleet_all + helper_assigns + guard_assigns

    # 1. Unconfirmed assignments at J-2
    for a in all_assigns:
        if a.get("assignment_status") == "estimate":
            for d in j2:
                if _is_assignment_active_on(d, a):
                    name = (a.get("function_name") or a.get("boat_name_override")
                            or a.get("helper_name_override") or f"#{a.get('id','?')}")
                    alerts.append({
                        "type": "unconfirmed",
                        "severity": "warning",
                        "msg": f"Unconfirmed assignment: {name} (estimate, active within 2 days)",
                        "entity_id": a.get("id"),
                    })
                    break

    # 2. Breakdowns
    for a in fleet_all:
        if a.get("assignment_status") == "breakdown":
            name = (a.get("boat_name") or a.get("name") or a.get("vehicle_name")
                    or a.get("boat_name_override") or f"#{a.get('id','?')}")
            alerts.append({
                "type": "breakdown",
                "severity": "danger",
                "msg": f"Breakdown: {name}",
                "entity_id": a.get("id"),
            })

    # 3. Budget overruns >10% per department
    dept_data = {}
    def _add_dept(key, rows):
        est = sum(r.get("amount_estimate") or 0 for r in rows if r.get("working_days"))
        act = sum(r.get("amount_actual") or 0 for r in rows if r.get("working_days"))
        dept_data[key] = {"estimate": est, "actual": act}

    _add_dept("boats", boat_assigns)
    _add_dept("picture_boats", pb_assigns)
    _add_dept("security_boats", sb_assigns)
    _add_dept("transport", tr_assigns)
    _add_dept("labour", helper_assigns)
    _add_dept("guards", guard_assigns)

    dept_labels = {
        "boats": "Boats", "picture_boats": "Picture Boats",
        "security_boats": "Security Boats", "transport": "Transport",
        "labour": "Labour", "guards": "Guards",
    }
    for key, d in dept_data.items():
        est = d["estimate"]
        act = d["actual"]
        if est > 0:
            overrun_pct = round((act - est) / est * 100, 1)
            if overrun_pct > 10:
                alerts.append({
                    "type": "budget_overrun",
                    "severity": "danger" if overrun_pct > 25 else "warning",
                    "msg": f"{dept_labels.get(key, key)}: +{overrun_pct}% over budget",
                    "dept": key,
                    "overrun_pct": overrun_pct,
                })
                _notify_budget_overrun(prod_id, dept_labels.get(key, key), overrun_pct)

    return jsonify({"alerts": alerts, "count": len(alerts)})


@app.route("/api/productions/<int:prod_id>/dashboard/burnrate", methods=["GET"])
def api_dashboard_burnrate(prod_id):
    """Return burn rate data: cumulative spend per day, linear projection, % consumed."""
    prod_or_404(prod_id)
    from datetime import datetime as dt

    shooting_days = get_shooting_days(prod_id)
    today = dt.now().strftime("%Y-%m-%d")

    # Get daily budget breakdown
    daily = get_daily_budget(prod_id)
    days_list = daily.get("days", [])

    # Cumulative spend per day
    cumulative = 0
    burn_data = []
    for d in days_list:
        total_day = d.get("total", 0) or 0
        if d["date"] <= today:
            cumulative += total_day
            burn_data.append({
                "date": d["date"],
                "day_number": d.get("day_number"),
                "daily": round(total_day, 2),
                "cumulative": round(cumulative, 2),
                "is_actual": True,
            })
        else:
            burn_data.append({
                "date": d["date"],
                "day_number": d.get("day_number"),
                "daily": round(total_day, 2),
                "cumulative": None,
                "is_actual": False,
            })

    # Compute totals
    total_spent = cumulative
    days_elapsed = sum(1 for d in days_list if d["date"] <= today)
    total_days = len(days_list)
    days_remaining = total_days - days_elapsed

    # Linear projection
    daily_rate = total_spent / max(days_elapsed, 1)
    projected_cumulative = cumulative
    for d in burn_data:
        if not d["is_actual"]:
            projected_cumulative += daily_rate
            d["projected_cumulative"] = round(projected_cumulative, 2)

    # Total budget estimate
    boat_assigns = get_boat_assignments(prod_id, context='boats')
    pb_assigns = get_picture_boat_assignments(prod_id)
    sb_assigns = get_security_boat_assignments(prod_id)
    tr_assigns = get_transport_assignments(prod_id)
    helper_assigns = get_helper_assignments(prod_id)
    guard_assigns = get_guard_camp_assignments(prod_id)
    all_assigns = boat_assigns + pb_assigns + sb_assigns + tr_assigns + helper_assigns + guard_assigns

    total_estimate = sum(r.get("amount_estimate") or 0 for r in all_assigns if r.get("working_days"))
    budget_consumed_pct = round(total_spent / max(total_estimate, 1) * 100, 1)

    return jsonify({
        "burn_data": burn_data,
        "total_spent": round(total_spent, 2),
        "total_estimate": round(total_estimate, 2),
        "budget_consumed_pct": budget_consumed_pct,
        "daily_rate": round(daily_rate, 2),
        "projected_total": round(daily_rate * total_days, 2),
        "days_elapsed": days_elapsed,
        "days_remaining": days_remaining,
    })


# ─── Export PDF Dashboard (P5.4) ─────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/export/dashboard-pdf", methods=["GET"])
def api_export_dashboard_pdf(prod_id):
    """Generate and return a one-page PDF summary of the executive dashboard."""
    from datetime import datetime as dt
    from export_dashboard import generate_dashboard_pdf

    prod = prod_or_404(prod_id)

    # Fetch KPI data by calling endpoint functions directly
    kpis = api_dashboard_kpis(prod_id).get_json()
    alerts_data = api_dashboard_alerts(prod_id).get_json()
    burnrate = api_dashboard_burnrate(prod_id).get_json()

    production_name = prod.get("name", prod.get("title", f"Production #{prod_id}"))
    pdf_bytes = generate_dashboard_pdf(production_name, kpis, alerts_data, burnrate)

    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    safe_name = production_name.replace(' ', '_').replace('/', '-')
    response.headers['Content-Disposition'] = (
        f'attachment; filename="dashboard_{safe_name}_{dt.now().strftime("%Y%m%d")}.pdf"'
    )
    return response


# ─── Conflict Alerts (AXE 7.3) ────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/alerts", methods=["GET"])
def api_alerts(prod_id):
    """Detect scheduling conflicts and return alerts:
    - Boats assigned on Off days
    - Locations in Film without guards
    - Cumulative boat capacity issues
    """
    prod_or_404(prod_id)

    alerts = []

    # ── 1. Boats assigned on Off days ──
    shooting_days = get_shooting_days(prod_id)
    # Build a set of dates that are "off" days (have at least one 'off' event)
    off_dates = set()
    day_date_map = {}  # date -> day info
    for day in shooting_days:
        d = day.get("date", "")
        if not d:
            continue
        day_date_map[d] = day
        for ev in day.get("events", []):
            if (ev.get("event_type") or "").lower() == "off":
                off_dates.add(d)

    if off_dates:
        # Check boat assignments (all contexts: boats, picture_boats, security_boats)
        for context_label, getter in [
            ("Boats", lambda: get_boat_assignments(prod_id, context='boats')),
            ("Picture Boats", lambda: get_picture_boat_assignments(prod_id)),
            ("Security Boats", lambda: get_security_boat_assignments(prod_id)),
        ]:
            assignments = getter()
            for asgn in assignments:
                start = (asgn.get("start_date") or "")[:10]
                end = (asgn.get("end_date") or "")[:10]
                if not start or not end:
                    continue
                overrides = json.loads(asgn.get("day_overrides") or "{}")
                boat_name = asgn.get("boat_name") or asgn.get("boat_name_override") or "Unknown"
                func_name = asgn.get("function_name") or ""
                for off_d in off_dates:
                    if off_d < start or off_d > end:
                        # Check if override explicitly activates this date
                        ov = overrides.get(off_d, "")
                        if ov and ov != "empty":
                            alerts.append({
                                "type": "boat_on_off",
                                "severity": "warning",
                                "module": context_label.lower().replace(" ", "_"),
                                "date": off_d,
                                "msg": f"{context_label}: '{boat_name}' ({func_name}) assigned on Off day {off_d}",
                                "entity": boat_name,
                                "function": func_name,
                            })
                        continue
                    # Date is in range - check if override disables it
                    ov = overrides.get(off_d, "")
                    if ov == "empty":
                        continue  # Explicitly disabled, no conflict
                    alerts.append({
                        "type": "boat_on_off",
                        "severity": "warning",
                        "module": context_label.lower().replace(" ", "_"),
                        "date": off_d,
                        "msg": f"{context_label}: '{boat_name}' ({func_name}) assigned on Off day {off_d}",
                        "entity": boat_name,
                        "function": func_name,
                    })

    # ── 2. Locations in Film without guards ──
    loc_schedules = get_location_schedules(prod_id)
    guard_loc_schedules = get_guard_location_schedules(prod_id)

    # Build guard lookup: {(location_name, date): nb_guards}
    guard_lookup = {}
    for gs in guard_loc_schedules:
        key = (gs["location_name"], gs["date"])
        guard_lookup[key] = gs.get("nb_guards") or 0

    for ls in loc_schedules:
        if ls.get("status") == "F":  # Film day
            loc_name = ls["location_name"]
            date = ls["date"]
            nb_guards = guard_lookup.get((loc_name, date), 0)
            if nb_guards == 0:
                alerts.append({
                    "type": "film_no_guards",
                    "severity": "danger",
                    "module": "guards",
                    "date": date,
                    "msg": f"Location '{loc_name}' in Film on {date} with no guards assigned",
                    "entity": loc_name,
                })

    # ── 3. Cumulative boat capacity vs function needs ──
    # For each date, sum up the capacity of all assigned boats per function group
    # Flag if any function has boats with capacity "?" or total seems low
    boat_assignments = get_boat_assignments(prod_id, context='boats')
    boats_data = get_boats(prod_id)
    boat_cap_map = {}  # boat_id -> numeric capacity (None if "EQ" or "?")
    for b in boats_data:
        cap = b.get("capacity", "")
        try:
            boat_cap_map[b["id"]] = int(cap)
        except (ValueError, TypeError):
            boat_cap_map[b["id"]] = None

    # Group assignments by function_group + date
    from collections import defaultdict
    func_group_daily_cap = defaultdict(lambda: {"total_cap": 0, "unknown": 0, "boats": []})

    for asgn in boat_assignments:
        start = (asgn.get("start_date") or "")[:10]
        end = (asgn.get("end_date") or "")[:10]
        if not start or not end:
            continue
        overrides = json.loads(asgn.get("day_overrides") or "{}")
        fg = asgn.get("function_group") or "Other"
        boat_id = asgn.get("boat_id")
        cap = boat_cap_map.get(boat_id)
        boat_name = asgn.get("boat_name") or asgn.get("boat_name_override") or "Unknown"

        # Iterate through all PDT dates
        for day in shooting_days:
            d = day.get("date", "")
            if not d or d in off_dates:
                continue
            # Check if assignment is active on this date
            if d in overrides:
                if overrides[d] == "empty":
                    continue
            elif d < start or d > end:
                continue

            key = (fg, d)
            if cap is not None:
                func_group_daily_cap[key]["total_cap"] += cap
            else:
                func_group_daily_cap[key]["unknown"] += 1
            func_group_daily_cap[key]["boats"].append(boat_name)

    # Check for function groups with only unknown-capacity boats
    seen_cap_alerts = set()
    for (fg, d), info in func_group_daily_cap.items():
        if info["unknown"] > 0 and info["total_cap"] == 0 and len(info["boats"]) > 0:
            alert_key = f"{fg}"
            if alert_key not in seen_cap_alerts:
                seen_cap_alerts.add(alert_key)
                alerts.append({
                    "type": "capacity_unknown",
                    "severity": "info",
                    "module": "boats",
                    "date": d,
                    "msg": f"Boats ({fg}): all boats have unknown capacity - verify manually",
                    "entity": fg,
                })

    # Sort alerts: danger first, then warning, then info
    severity_order = {"danger": 0, "warning": 1, "info": 2}
    alerts.sort(key=lambda a: (severity_order.get(a.get("severity"), 3), a.get("date", "")))

    return jsonify({"alerts": alerts, "count": len(alerts)})


# ─── Budget Daily (AXE 6.2) ──────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/budget/daily", methods=["GET"])
def api_budget_daily(prod_id):
    """Return cost breakdown per shooting day."""
    prod_or_404(prod_id)
    return jsonify_cached(get_daily_budget(prod_id))


# ─── Budget Variance (P5.5) ───────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/budget/variance", methods=["GET"])
def api_budget_variance(prod_id):
    """Return estimated vs actual variance per department with drill-down rows."""
    prod_or_404(prod_id)
    budget = get_budget(prod_id)
    by_dept = budget.get("by_department", {})

    departments = []
    for dept_name, ddata in by_dept.items():
        estimated = ddata.get("total_estimate", 0) or 0
        actual = ddata.get("total_actual", 0) or 0
        variance = actual - estimated
        variance_pct = round((variance / estimated) * 100, 1) if estimated else 0

        # Drill-down: individual lines with variance
        lines = []
        for row in ddata.get("lines", []):
            r_est = row.get("amount_estimate", 0) or 0
            r_act = row.get("amount_actual", 0) or 0
            r_var = r_act - r_est
            r_pct = round((r_var / r_est) * 100, 1) if r_est else 0
            lines.append({
                "name": row.get("name", ""),
                "detail": row.get("boat") or row.get("detail") or "",
                "estimated": round(r_est, 2),
                "actual": round(r_act, 2),
                "variance": round(r_var, 2),
                "variance_percent": r_pct,
            })

        departments.append({
            "department_name": dept_name,
            "estimated_total": round(estimated, 2),
            "actual_total": round(actual, 2),
            "variance": round(variance, 2),
            "variance_percent": variance_pct,
            "lines": lines,
        })

    # Sort by absolute variance descending
    departments.sort(key=lambda d: abs(d["variance"]), reverse=True)

    grand_est = budget.get("grand_total_estimate", 0)
    grand_act = budget.get("grand_total_actual", 0)
    grand_var = grand_act - grand_est
    grand_pct = round((grand_var / grand_est) * 100, 1) if grand_est else 0

    return jsonify({
        "departments": departments,
        "grand_estimated": round(grand_est, 2),
        "grand_actual": round(grand_act, 2),
        "grand_variance": round(grand_var, 2),
        "grand_variance_percent": grand_pct,
    })


# ─── Budget Snapshots & Price Log (AXE 6.3) ──────────────────────────────────

@app.route("/api/productions/<int:prod_id>/budget/snapshots", methods=["GET"])
def api_budget_snapshots(prod_id):
    """List budget snapshots for a production."""
    prod_or_404(prod_id)
    limit = int(request.args.get("limit", 50))
    return jsonify(get_budget_snapshots(prod_id, limit))


@app.route("/api/productions/<int:prod_id>/budget/snapshots", methods=["POST"])
def api_create_budget_snapshot(prod_id):
    """Manually create a budget snapshot."""
    prod_or_404(prod_id)
    data = request.json or {}
    note = data.get('note', '')
    snap_id = create_budget_snapshot(
        prod_id, trigger_type='manual',
        trigger_detail=note or 'Manual snapshot',
        user_id=getattr(g, 'user_id', None),
        user_nickname=getattr(g, 'nickname', None)
    )
    return jsonify({"ok": True, "snapshot_id": snap_id}), 201


@app.route("/api/productions/<int:prod_id>/budget/snapshots/<int:snap_id>", methods=["GET"])
def api_budget_snapshot_detail(prod_id, snap_id):
    """Get a single snapshot with full budget data."""
    prod_or_404(prod_id)
    snap = get_budget_snapshot(snap_id)
    if not snap or snap['production_id'] != prod_id:
        abort(404)
    return jsonify(snap)


@app.route("/api/productions/<int:prod_id>/budget/snapshots/<int:snap_id>", methods=["DELETE"])
def api_delete_budget_snapshot(prod_id, snap_id):
    """Delete a budget snapshot."""
    prod_or_404(prod_id)
    snap = get_budget_snapshot(snap_id)
    if not snap or snap['production_id'] != prod_id:
        abort(404)
    delete_budget_snapshot(snap_id)
    return jsonify({"ok": True})


@app.route("/api/productions/<int:prod_id>/budget/snapshots/compare", methods=["GET"])
def api_compare_budget_snapshots(prod_id):
    """Compare two budget snapshots. Query params: a=<id>&b=<id>"""
    prod_or_404(prod_id)
    snap_a = request.args.get("a")
    snap_b = request.args.get("b")
    if not snap_a or not snap_b:
        return jsonify({"error": "Query params a and b required"}), 400
    result = compare_budget_snapshots(int(snap_a), int(snap_b))
    if not result:
        return jsonify({"error": "One or both snapshots not found"}), 404
    return jsonify(result)


@app.route("/api/productions/<int:prod_id>/budget/price-log", methods=["GET"])
def api_price_change_log(prod_id):
    """Get price change log entries."""
    prod_or_404(prod_id)
    limit = int(request.args.get("limit", 100))
    entity_type = request.args.get("entity_type")
    entity_id = request.args.get("entity_id")
    if entity_id:
        entity_id = int(entity_id)
    return jsonify(get_price_change_log(prod_id, limit, entity_type, entity_id))


# ─── Comments (AXE 9.1) ──────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/comments", methods=["GET"])
def api_get_comments(prod_id):
    prod_or_404(prod_id)
    entity_type = request.args.get("entity_type")
    entity_id = request.args.get("entity_id")
    if not entity_type or not entity_id:
        return jsonify({"error": "entity_type and entity_id required"}), 400
    entity_id = int(entity_id)
    limit = int(request.args.get("limit", 50))
    return jsonify(get_comments(prod_id, entity_type, entity_id, limit))


@app.route("/api/productions/<int:prod_id>/comments/counts", methods=["GET"])
def api_comment_counts(prod_id):
    prod_or_404(prod_id)
    entity_type = request.args.get("entity_type")
    ids_raw = request.args.get("entity_ids", "")
    if not entity_type or not ids_raw:
        return jsonify({})
    try:
        entity_ids = [int(x) for x in ids_raw.split(",") if x.strip()]
    except ValueError:
        return jsonify({"error": "invalid entity_ids"}), 400
    counts = get_comment_counts(prod_id, entity_type, entity_ids)
    return jsonify(counts)


@app.route("/api/productions/<int:prod_id>/comments", methods=["POST"])
def api_create_comment(prod_id):
    prod_or_404(prod_id)
    data = request.json or {}
    entity_type = data.get("entity_type")
    entity_id = data.get("entity_id")
    body = (data.get("body") or "").strip()
    if not entity_type or not entity_id or not body:
        return jsonify({"error": "entity_type, entity_id, and body required"}), 400
    comment = create_comment(prod_id, entity_type, int(entity_id), body)
    # Notify other project members about the new comment
    try:
        nickname = getattr(g, 'nickname', 'Someone')
        create_notifications_for_production(
            prod_id, 'comment_added',
            f'{nickname} commented on {entity_type}',
            body[:100],
            entity_type=entity_type, entity_id=int(entity_id),
            exclude_user_id=getattr(g, 'user_id', None)
        )
    except Exception:
        pass  # notification failure should not block comment creation
    return jsonify(comment), 201


@app.route("/api/comments/<int:comment_id>", methods=["DELETE"])
def api_delete_comment(comment_id):
    result = delete_comment(comment_id)
    if not result:
        abort(404)
    return jsonify({"deleted": comment_id})


# ─── Notifications (AXE 9.2) ────────────────────────────────────────────────

@app.route("/api/notifications", methods=["GET"])
def api_get_notifications():
    user_id = getattr(g, 'user_id', None)
    if not user_id:
        return jsonify({"error": "auth required"}), 401
    prod_id = request.args.get("production_id")
    if prod_id:
        prod_id = int(prod_id)
    unread_only = request.args.get("unread_only") == "1"
    limit = int(request.args.get("limit", 50))
    return jsonify(get_notifications(user_id, prod_id, unread_only, limit))


@app.route("/api/notifications/count", methods=["GET"])
def api_notification_count():
    user_id = getattr(g, 'user_id', None)
    if not user_id:
        return jsonify({"error": "auth required"}), 401
    prod_id = request.args.get("production_id")
    if prod_id:
        prod_id = int(prod_id)
    count = get_unread_notification_count(user_id, prod_id)
    return jsonify({"count": count})


@app.route("/api/notifications/<int:notif_id>/read", methods=["POST"])
def api_mark_notification_read(notif_id):
    user_id = getattr(g, 'user_id', None)
    if not user_id:
        return jsonify({"error": "auth required"}), 401
    mark_notification_read(notif_id, user_id)
    return jsonify({"ok": True})


@app.route("/api/notifications/read-all", methods=["POST"])
def api_mark_all_read():
    user_id = getattr(g, 'user_id', None)
    if not user_id:
        return jsonify({"error": "auth required"}), 401
    prod_id = request.args.get("production_id")
    if prod_id:
        prod_id = int(prod_id)
    mark_all_notifications_read(user_id, prod_id)
    return jsonify({"ok": True})


# ─── Notification triggers (AXE 9.2) ────────────────────────────────────────
# Helper to send notifications on key mutations. Called from within endpoints.

def _notify_assignment_change(prod_id, action, entity_type, entity_name):
    """Send notification when an assignment is created/updated/deleted."""
    try:
        nickname = getattr(g, 'nickname', 'Someone')
        verb = {'create': 'created', 'update': 'updated', 'delete': 'deleted'}.get(action, action)
        create_notifications_for_production(
            prod_id, f'assignment_{verb}',
            f'{nickname} {verb} {entity_type}',
            entity_name,
            exclude_user_id=getattr(g, 'user_id', None)
        )
    except Exception:
        pass


def _notify_vessel_breakdown(prod_id, vessel_name, module_label):
    """Send notification when a vessel/vehicle assignment status changes to breakdown."""
    try:
        nickname = getattr(g, 'nickname', 'Someone')
        create_notifications_for_production(
            prod_id, 'vessel_breakdown',
            f'{vessel_name} is in breakdown',
            f'{nickname} marked {vessel_name} ({module_label}) as breakdown',
            exclude_user_id=getattr(g, 'user_id', None)
        )
    except Exception:
        pass


def _notify_budget_overrun(prod_id, dept_label, overrun_pct):
    """Send notification when budget exceeds +10% for a department."""
    try:
        create_notifications_for_production(
            prod_id, 'budget_overrun',
            f'{dept_label}: +{overrun_pct}% over budget',
            f'Department {dept_label} has exceeded its estimated budget by {overrun_pct}%',
        )
    except Exception:
        pass


def _notify_pdt_change(prod_id, action, day_info):
    """Send notification when PDT is modified."""
    try:
        nickname = getattr(g, 'nickname', 'Someone')
        verb = {'create': 'added', 'update': 'modified', 'delete': 'removed'}.get(action, action)
        create_notifications_for_production(
            prod_id, 'pdt_modified',
            f'{nickname} {verb} a shooting day',
            day_info,
            exclude_user_id=getattr(g, 'user_id', None)
        )
    except Exception:
        pass


# ─── CSV Import & Templates (AXE 10.4) ────────────────────────────────────────

_CSV_TEMPLATES = {
    "boats": {
        "header": "name,boat_nr,capacity,wave_rating,captain,vendor,group_name,daily_rate_estimate,night_ok,notes",
        "example": "Panga 1,1,12,Waves,Carlos,QS Marine,Shared,80,0,Fast boat",
        "create_fn": "create_boat",
        "fields": {
            "name": {"type": "str", "required": True},
            "boat_nr": {"type": "int"},
            "capacity": {"type": "str"},
            "wave_rating": {"type": "str", "default": "Waves"},
            "captain": {"type": "str"},
            "vendor": {"type": "str"},
            "group_name": {"type": "str", "default": "Shared"},
            "daily_rate_estimate": {"type": "float", "default": 0},
            "night_ok": {"type": "int", "default": 0},
            "notes": {"type": "str"},
        }
    },
    "picture_boats": {
        "header": "name,boat_nr,capacity,wave_rating,captain,vendor,group_name,daily_rate_estimate,night_ok,notes",
        "example": "Camera Boat A,1,8,Waves,Pedro,QS Marine,Custom,120,0,Stabilized",
        "create_fn": "create_picture_boat",
        "fields": {
            "name": {"type": "str", "required": True},
            "boat_nr": {"type": "int"},
            "capacity": {"type": "str"},
            "wave_rating": {"type": "str", "default": "Waves"},
            "captain": {"type": "str"},
            "vendor": {"type": "str"},
            "group_name": {"type": "str", "default": "Custom"},
            "daily_rate_estimate": {"type": "float", "default": 0},
            "night_ok": {"type": "int", "default": 0},
            "notes": {"type": "str"},
        }
    },
    "security_boats": {
        "header": "name,boat_nr,capacity,wave_rating,captain,vendor,group_name,daily_rate_estimate,night_ok,notes",
        "example": "Safety 1,1,6,Big Waves,Jose,Local,SAFETY,60,1,Night capable",
        "create_fn": "create_security_boat",
        "fields": {
            "name": {"type": "str", "required": True},
            "boat_nr": {"type": "int"},
            "capacity": {"type": "str"},
            "wave_rating": {"type": "str", "default": "Waves"},
            "captain": {"type": "str"},
            "vendor": {"type": "str"},
            "group_name": {"type": "str", "default": "SAFETY"},
            "daily_rate_estimate": {"type": "float", "default": 0},
            "night_ok": {"type": "int", "default": 0},
            "notes": {"type": "str"},
        }
    },
    "transport": {
        "header": "name,vehicle_nr,type,driver,vendor,group_name,daily_rate_estimate,notes",
        "example": "SUV-01,1,SUV,Miguel,Rent-a-Car,UNIT,120,Air conditioned",
        "create_fn": "create_transport_vehicle",
        "fields": {
            "name": {"type": "str", "required": True},
            "vehicle_nr": {"type": "int"},
            "type": {"type": "str", "default": "SUV"},
            "driver": {"type": "str"},
            "vendor": {"type": "str"},
            "group_name": {"type": "str", "default": "UNIT"},
            "daily_rate_estimate": {"type": "float", "default": 0},
            "notes": {"type": "str"},
        }
    },
    "locations": {
        "header": "name,type,location_type,lat,lng,access_note,price_p,price_f,price_w,global_deal",
        "example": "Mogo Mogo,ile,game,8.35,-78.92,By boat only,100,200,100,",
        "create_fn": "create_location_site",
        "fields": {
            "name": {"type": "str", "required": True},
            "type": {"type": "str", "default": "ile"},
            "location_type": {"type": "str", "default": "game"},
            "lat": {"type": "float"},
            "lng": {"type": "float"},
            "access_note": {"type": "str"},
            "price_p": {"type": "float"},
            "price_f": {"type": "float"},
            "price_w": {"type": "float"},
            "global_deal": {"type": "float"},
        }
    },
}

# Map create function names to actual functions
_CSV_CREATE_FNS = {
    "create_boat": create_boat,
    "create_picture_boat": create_picture_boat,
    "create_security_boat": create_security_boat,
    "create_transport_vehicle": create_transport_vehicle,
    "create_location_site": create_location_site,
}


@app.route("/api/productions/<int:prod_id>/import-csv/<string:module>", methods=["POST"])
def api_import_csv(prod_id, module):
    """Generic CSV import. Module: boats, picture_boats, security_boats, transport, locations."""
    prod_or_404(prod_id)
    tpl = _CSV_TEMPLATES.get(module)
    if not tpl:
        return jsonify({"error": f"Unknown module: {module}. Available: {', '.join(_CSV_TEMPLATES.keys())}"}), 400

    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400

    content = f.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))

    create_fn = _CSV_CREATE_FNS[tpl["create_fn"]]
    fields = tpl["fields"]
    created = []
    errors = []
    line_num = 1

    for line_num, row in enumerate(reader, start=2):  # CSV header is line 1
        # Validate required fields
        rec = {"production_id": prod_id}
        row_errors = []

        for field_name, field_spec in fields.items():
            raw = (row.get(field_name) or "").strip()
            if field_spec.get("required") and not raw:
                row_errors.append(f"Missing required field '{field_name}'")
                continue

            if not raw:
                if "default" in field_spec:
                    rec[field_name] = field_spec["default"]
                else:
                    rec[field_name] = None
                continue

            ftype = field_spec["type"]
            try:
                if ftype == "int":
                    rec[field_name] = int(raw)
                elif ftype == "float":
                    rec[field_name] = float(raw)
                else:
                    rec[field_name] = raw
            except (ValueError, TypeError):
                row_errors.append(f"Invalid {ftype} for '{field_name}': '{raw}'")
                rec[field_name] = field_spec.get("default")

        if row_errors:
            errors.append({"line": line_num, "errors": row_errors})
            # Still try to create if name is present
            if not rec.get("name"):
                continue

        try:
            result = create_fn(rec)
            eid = result["id"] if isinstance(result, dict) else result
            created.append(eid)
        except Exception as e:
            errors.append({"line": line_num, "errors": [str(e)]})

    return jsonify({
        "created": len(created),
        "ids": created,
        "errors": errors,
        "total_rows": line_num - 1,
    }), 201


@app.route("/api/csv-template/<string:module>", methods=["GET"])
def api_csv_template(module):
    """Download a CSV template for a module."""
    tpl = _CSV_TEMPLATES.get(module)
    if not tpl:
        # Check helpers/guards which have different templates
        if module == "helpers":
            content = "name,role,group,rate,notes\nJohn Doe,Setup,GENERAL,45,\nJane Smith,Runner,GENERAL,50,Experienced"
        elif module == "guard_camp":
            content = "name,role,group,rate,notes\nGuard 1,Security,GENERAL,45,\nGuard 2,Patrol,GENERAL,50,"
        else:
            return jsonify({"error": f"Unknown module: {module}"}), 404
    else:
        content = tpl["header"] + "\n" + tpl["example"]

    return Response(
        content,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={module}_template.csv"}
    )


# ─── Soft Delete: Restore endpoints (P2.4) ───────────────────────────────────

_RESTORE_ROUTES = {
    "boats":              ("boats",              "/api/boats/<int:entity_id>/restore"),
    "picture_boats":      ("picture_boats",      "/api/picture-boats/<int:entity_id>/restore"),
    "security_boats":     ("security_boats",     "/api/security-boats/<int:entity_id>/restore"),
    "transport_vehicles": ("transport_vehicles",  "/api/transport-vehicles/<int:entity_id>/restore"),
    "helpers":            ("helpers",             "/api/helpers/<int:entity_id>/restore"),
    "guard_camp_workers": ("guard_camp_workers",  "/api/guard-camp-workers/<int:entity_id>/restore"),
    "locations":          ("locations",           "/api/locations/<int:entity_id>/restore"),
    "boat_functions":     ("boat_functions",      "/api/boat-functions/<int:entity_id>/restore"),
    "physical_vessels":   ("physical_vessels",    "/api/physical-vessels/<int:entity_id>/restore"),
    "fuel_machinery":     ("fuel_machinery",      "/api/fuel-machinery/<int:entity_id>/restore"),
    "guard_posts":        ("guard_posts",         "/api/guard-posts/<int:entity_id>/restore"),
    "fnb_items":          ("fnb_items",           "/api/fnb-items/<int:entity_id>/restore"),
}

for _key, (_table, _route) in _RESTORE_ROUTES.items():
    def _make_restore(table):
        def _restore(entity_id):
            restore_entity(table, entity_id)
            return jsonify({"restored": entity_id})
        _restore.__name__ = f"api_restore_{table}"
        return _restore
    app.add_url_rule(_route, view_func=_make_restore(_table), methods=["POST"])


@app.route("/api/fnb-categories/<int:cat_id>/restore", methods=["POST"])
def api_restore_fnb_category(cat_id):
    restore_fnb_category(cat_id)
    return jsonify({"restored": cat_id})


# ─── Documents & Versioning (P5.7) ────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/documents", methods=["GET"])
def api_get_documents(prod_id):
    prod_or_404(prod_id)
    db = get_db()
    rows = db.execute(
        "SELECT * FROM documents WHERE production_id = ? ORDER BY uploaded_at DESC",
        (prod_id,)
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/productions/<int:prod_id>/documents", methods=["POST"])
def api_create_document(prod_id):
    prod_or_404(prod_id)
    db = get_db()
    data = request.form if request.files else request.get_json(force=True)
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "name is required"}), 400
    doc_type = data.get("doc_type", "")
    department_id = data.get("department_id") or None
    status = data.get("status", "draft")
    user_id = getattr(g, 'user_id', None)
    nickname = getattr(g, 'nickname', None)

    file_path = None
    fmt = data.get("format", "")
    if "file" in request.files:
        f = request.files["file"]
        if f.filename:
            import werkzeug.utils
            upload_dir = os.path.join(os.path.dirname(__file__), 'data', 'documents', str(prod_id))
            os.makedirs(upload_dir, exist_ok=True)
            safe_name = werkzeug.utils.secure_filename(f.filename)
            dest = os.path.join(upload_dir, safe_name)
            f.save(dest)
            file_path = f"data/documents/{prod_id}/{safe_name}"
            fmt = fmt or os.path.splitext(safe_name)[1].lstrip('.')

    cur = db.execute(
        """INSERT INTO documents (production_id, department_id, name, doc_type, format, file_path, uploaded_by, status, current_version)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)""",
        (prod_id, department_id, name, doc_type, fmt, file_path, user_id, status)
    )
    doc_id = cur.lastrowid

    # Create initial version entry
    db.execute(
        """INSERT INTO document_versions (document_id, version_number, file_path, uploaded_by, upload_nickname)
           VALUES (?, 1, ?, ?, ?)""",
        (doc_id, file_path, user_id, nickname)
    )
    db.commit()
    return jsonify({"id": doc_id}), 201


@app.route("/api/productions/<int:prod_id>/documents/<int:doc_id>", methods=["PUT"])
def api_update_document(prod_id, doc_id):
    prod_or_404(prod_id)
    db = get_db()
    row = db.execute("SELECT * FROM documents WHERE id = ? AND production_id = ?", (doc_id, prod_id)).fetchone()
    if not row:
        abort(404, description="Document not found")
    data = request.get_json(force=True)
    name = data.get("name", row["name"])
    doc_type = data.get("doc_type", row["doc_type"])
    department_id = data.get("department_id", row["department_id"])
    db.execute(
        "UPDATE documents SET name = ?, doc_type = ?, department_id = ? WHERE id = ?",
        (name, doc_type, department_id, doc_id)
    )
    db.commit()
    return jsonify({"updated": doc_id})


@app.route("/api/productions/<int:prod_id>/documents/<int:doc_id>", methods=["DELETE"])
def api_delete_document(prod_id, doc_id):
    prod_or_404(prod_id)
    db = get_db()
    db.execute("DELETE FROM document_versions WHERE document_id = ?", (doc_id,))
    db.execute("DELETE FROM documents WHERE id = ? AND production_id = ?", (doc_id, prod_id))
    db.commit()
    return jsonify({"deleted": doc_id})


@app.route("/api/productions/<int:prod_id>/documents/<int:doc_id>/status", methods=["PUT"])
def api_update_document_status(prod_id, doc_id):
    prod_or_404(prod_id)
    db = get_db()
    row = db.execute("SELECT * FROM documents WHERE id = ? AND production_id = ?", (doc_id, prod_id)).fetchone()
    if not row:
        abort(404, description="Document not found")
    data = request.get_json(force=True)
    new_status = data.get("status", "").strip()
    valid = ('draft', 'under_review', 'approved', 'archived')
    if new_status not in valid:
        return jsonify({"error": f"status must be one of {valid}"}), 400
    db.execute("UPDATE documents SET status = ? WHERE id = ?", (new_status, doc_id))
    db.commit()
    return jsonify({"id": doc_id, "status": new_status})


@app.route("/api/productions/<int:prod_id>/documents/<int:doc_id>/versions", methods=["GET"])
def api_get_document_versions(prod_id, doc_id):
    prod_or_404(prod_id)
    db = get_db()
    row = db.execute("SELECT id FROM documents WHERE id = ? AND production_id = ?", (doc_id, prod_id)).fetchone()
    if not row:
        abort(404, description="Document not found")
    versions = db.execute(
        "SELECT * FROM document_versions WHERE document_id = ? ORDER BY version_number DESC",
        (doc_id,)
    ).fetchall()
    return jsonify([dict(v) for v in versions])


@app.route("/api/productions/<int:prod_id>/documents/<int:doc_id>/versions", methods=["POST"])
def api_upload_document_version(prod_id, doc_id):
    prod_or_404(prod_id)
    db = get_db()
    row = db.execute("SELECT * FROM documents WHERE id = ? AND production_id = ?", (doc_id, prod_id)).fetchone()
    if not row:
        abort(404, description="Document not found")

    user_id = getattr(g, 'user_id', None)
    nickname = getattr(g, 'nickname', None)
    current_ver = row["current_version"] or 1
    new_ver = current_ver + 1

    file_path = None
    if "file" in request.files:
        f = request.files["file"]
        if f.filename:
            import werkzeug.utils
            upload_dir = os.path.join(os.path.dirname(__file__), 'data', 'documents', str(prod_id))
            os.makedirs(upload_dir, exist_ok=True)
            safe_name = werkzeug.utils.secure_filename(f.filename)
            base, ext = os.path.splitext(safe_name)
            versioned_name = f"{base}_v{new_ver}{ext}"
            dest = os.path.join(upload_dir, versioned_name)
            f.save(dest)
            file_path = f"data/documents/{prod_id}/{versioned_name}"

    db.execute(
        """INSERT INTO document_versions (document_id, version_number, file_path, uploaded_by, upload_nickname)
           VALUES (?, ?, ?, ?, ?)""",
        (doc_id, new_ver, file_path, user_id, nickname)
    )
    db.execute(
        "UPDATE documents SET current_version = ?, file_path = ? WHERE id = ?",
        (new_ver, file_path or row["file_path"], doc_id)
    )
    db.commit()
    return jsonify({"id": doc_id, "version": new_ver}), 201


@app.route("/api/documents/download/<path:filepath>", methods=["GET"])
def api_download_document(filepath):
    """Serve a document file from the data/documents directory."""
    full = os.path.join(os.path.dirname(__file__), filepath)
    if not os.path.isfile(full):
        abort(404, description="File not found")
    from flask import send_file
    return send_file(full, as_attachment=True)


# ─── Daily Production Report (P5.9) ──────────────────────────────────────────

def _assignment_active_on_date(asgn, target_date):
    """Check if an assignment is active on a specific date (respects day_overrides)."""
    start = (asgn.get("start_date") or "")[:10]
    end = (asgn.get("end_date") or "")[:10]
    if not start or not end:
        return False
    if target_date < start or target_date > end:
        # Check if override explicitly activates this date
        overrides = json.loads(asgn.get("day_overrides") or "{}")
        ov = overrides.get(target_date, "")
        return bool(ov and ov != "empty")
    # Date is in range - check if override disables it
    overrides = json.loads(asgn.get("day_overrides") or "{}")
    return overrides.get(target_date, "") != "empty"


def _daily_rate(asgn, rate_actual_key="boat_daily_rate_actual",
                rate_estimate_key="boat_daily_rate_estimate"):
    """Get effective daily rate (price_override > actual > estimate)."""
    if asgn.get("price_override") is not None:
        return asgn["price_override"]
    return asgn.get(rate_actual_key) or asgn.get(rate_estimate_key) or 0


@app.route("/api/productions/<int:prod_id>/reports/daily", methods=["GET"])
def api_daily_report(prod_id):
    """Generate a PDF daily production report for a given date."""
    from datetime import datetime as dt
    from daily_report import generate_daily_report

    prod = prod_or_404(prod_id)
    target_date = request.args.get("date")
    if not target_date:
        target_date = dt.now().strftime("%Y-%m-%d")

    production_name = prod.get("name", prod.get("title", f"Production #{prod_id}"))

    # ── PDT info for the day ──
    shooting_days = get_shooting_days(prod_id)
    day_info = {"day_number": "?", "location": "N/A", "status": "", "events": []}
    for day in shooting_days:
        if (day.get("date") or "")[:10] == target_date:
            day_info = day
            break

    # ── Boats active on this date ──
    boat_assigns = get_boat_assignments(prod_id, context='boats')
    boats = []
    for a in boat_assigns:
        if _assignment_active_on_date(a, target_date):
            boats.append({
                "name": a.get("boat_name") or a.get("boat_name_override") or "",
                "function": a.get("function_name") or "",
                "group_name": a.get("function_group") or "",
                "status": a.get("assignment_status") or "",
                "rate": _daily_rate(a, "boat_daily_rate_actual", "boat_daily_rate_estimate"),
            })

    # ── Picture Boats ──
    pb_assigns = get_picture_boat_assignments(prod_id)
    picture_boats = []
    for a in pb_assigns:
        if _assignment_active_on_date(a, target_date):
            picture_boats.append({
                "name": a.get("boat_name") or a.get("boat_name_override") or "",
                "function": a.get("function_name") or "",
                "group_name": a.get("function_group") or "",
                "status": a.get("assignment_status") or "",
                "rate": _daily_rate(a, "boat_daily_rate_actual", "boat_daily_rate_estimate"),
            })

    # ── Security Boats ──
    sb_assigns = get_security_boat_assignments(prod_id)
    security_boats = []
    for a in sb_assigns:
        if _assignment_active_on_date(a, target_date):
            security_boats.append({
                "name": a.get("boat_name") or a.get("boat_name_override") or "",
                "function": a.get("function_name") or "",
                "group_name": a.get("function_group") or "",
                "status": a.get("assignment_status") or "",
                "rate": _daily_rate(a, "boat_daily_rate_actual", "boat_daily_rate_estimate"),
            })

    # ── Transport ──
    tr_assigns = get_transport_assignments(prod_id)
    vehicles = []
    for a in tr_assigns:
        if _assignment_active_on_date(a, target_date):
            vehicles.append({
                "name": a.get("vehicle_name") or "",
                "vehicle_type": a.get("vehicle_type") or "",
                "group_name": a.get("function_group") or "",
                "rate": _daily_rate(a, "vehicle_daily_rate_actual", "vehicle_daily_rate_estimate"),
            })

    # ── Personnel / Labour ──
    helper_assigns = get_helper_assignments(prod_id)
    personnel = []
    for a in helper_assigns:
        if _assignment_active_on_date(a, target_date):
            personnel.append({
                "name": a.get("helper_name") or "",
                "function": a.get("function_name") or "",
                "group_name": a.get("function_group") or a.get("helper_group") or "",
                "rate": _daily_rate(a, "helper_daily_rate_actual", "helper_daily_rate_estimate"),
            })

    # ── Guards (Camp) ──
    guard_assigns = get_guard_camp_assignments(prod_id)
    guards = []
    for a in guard_assigns:
        if _assignment_active_on_date(a, target_date):
            guards.append({
                "name": a.get("helper_name") or "",
                "post": a.get("function_name") or "",
                "shift": a.get("function_group") or "",
                "rate": _daily_rate(a, "helper_daily_rate_actual", "helper_daily_rate_estimate"),
            })

    # ── Fuel entries for this date ──
    all_fuel = get_fuel_entries(prod_id)
    fuel = [f for f in all_fuel if (f.get("date") or "")[:10] == target_date]

    # ── Alerts for this date ──
    alerts_resp = api_alerts(prod_id)
    all_alerts = alerts_resp.get_json() if hasattr(alerts_resp, 'get_json') else []
    day_alerts = [a for a in all_alerts if a.get("date", "") == target_date]

    # ── Generate PDF ──
    pdf_bytes = generate_daily_report(
        production_name=production_name,
        report_date=target_date,
        day_info=day_info,
        boats=boats,
        picture_boats=picture_boats,
        security_boats=security_boats,
        vehicles=vehicles,
        personnel=personnel,
        fuel_entries=fuel,
        alerts=day_alerts,
        guards=guards,
    )

    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    safe_name = production_name.replace(' ', '_').replace('/', '-')
    response.headers['Content-Disposition'] = (
        f'attachment; filename="daily_report_{safe_name}_{target_date}.pdf"'
    )
    return response


@app.route("/api/productions/<int:prod_id>/reports/daily/data", methods=["GET"])
def api_daily_report_data(prod_id):
    """Return daily report data as JSON (for dashboard preview)."""
    from datetime import datetime as dt

    prod = prod_or_404(prod_id)
    target_date = request.args.get("date")
    if not target_date:
        target_date = dt.now().strftime("%Y-%m-%d")

    shooting_days = get_shooting_days(prod_id)
    day_info = None
    for day in shooting_days:
        if (day.get("date") or "")[:10] == target_date:
            day_info = day
            break

    # Count active resources
    boats_count = sum(1 for a in get_boat_assignments(prod_id, context='boats')
                      if _assignment_active_on_date(a, target_date))
    pb_count = sum(1 for a in get_picture_boat_assignments(prod_id)
                   if _assignment_active_on_date(a, target_date))
    sb_count = sum(1 for a in get_security_boat_assignments(prod_id)
                   if _assignment_active_on_date(a, target_date))
    tr_count = sum(1 for a in get_transport_assignments(prod_id)
                   if _assignment_active_on_date(a, target_date))
    pers_count = sum(1 for a in get_helper_assignments(prod_id)
                     if _assignment_active_on_date(a, target_date))
    guard_count = sum(1 for a in get_guard_camp_assignments(prod_id)
                      if _assignment_active_on_date(a, target_date))
    fuel_count = sum(1 for f in get_fuel_entries(prod_id)
                     if (f.get("date") or "")[:10] == target_date)

    return jsonify({
        "date": target_date,
        "day_number": day_info.get("day_number") if day_info else None,
        "location": day_info.get("location") if day_info else None,
        "has_shooting_day": day_info is not None,
        "resources": {
            "boats": boats_count,
            "picture_boats": pb_count,
            "security_boats": sb_count,
            "transport": tr_count,
            "personnel": pers_count,
            "guards": guard_count,
            "fuel_entries": fuel_count,
        },
        "total_resources": boats_count + pb_count + sb_count + tr_count + pers_count + guard_count,
    })


# TODO: Add a cron job (e.g. via APScheduler or system cron) to auto-generate
# the daily report at 06:00 each morning for the current shooting day:
#   0 6 * * * curl -H "Authorization: Bearer $TOKEN" \
#     "https://shootlogix.app/api/productions/<id>/reports/daily?date=$(date +%Y-%m-%d)" \
#     -o /data/reports/daily_$(date +%Y-%m-%d).pdf


# ─── Holidays API (P5.10) ─────────────────────────────────────────────────────

@app.route("/api/holidays", methods=["GET"])
def api_holidays():
    """Return all holidays (optionally filtered by country)."""
    country = request.args.get("country")
    with get_db() as conn:
        if country:
            rows = conn.execute("SELECT id, date, name, country FROM holidays WHERE country=?", (country,)).fetchall()
        else:
            rows = conn.execute("SELECT id, date, name, country FROM holidays").fetchall()
        return jsonify([dict(r) for r in rows])


# ─── Timeline (Gantt) ─────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/timeline", methods=["GET"])
def api_timeline(prod_id):
    """Aggregate all resources and assignments for the Gantt timeline view."""
    prod = prod_or_404(prod_id)
    with get_db() as conn:
        # Production date range
        start_date = prod.get('start_date') or prod.get('shooting_start')
        end_date = prod.get('end_date') or prod.get('shooting_end')

        # Shooting days
        days = [dict(r) for r in conn.execute(
            "SELECT id, date, day_number, location, game_name, status FROM shooting_days WHERE production_id=? ORDER BY date",
            (prod_id,)
        ).fetchall()]

        resources = []

        # --- Boats ---
        boats = conn.execute("SELECT id, name, group_name FROM boats WHERE production_id=?", (prod_id,)).fetchall()
        for b in boats:
            assignments = conn.execute(
                "SELECT id, start_date, end_date, assignment_status, day_overrides, boat_function_id FROM boat_assignments WHERE boat_id=?",
                (b['id'],)
            ).fetchall()
            resources.append({
                'id': f"boat-{b['id']}", 'name': b['name'], 'type': 'boat', 'group': 'Boats',
                'subgroup': b['group_name'] or 'Boats',
                'assignments': [_timeline_assign(a) for a in assignments]
            })

        # --- Picture boats ---
        pboats = conn.execute("SELECT id, name, group_name FROM picture_boats WHERE production_id=?", (prod_id,)).fetchall()
        for b in pboats:
            assignments = conn.execute(
                "SELECT id, start_date, end_date, assignment_status, day_overrides, boat_function_id FROM picture_boat_assignments WHERE picture_boat_id=?",
                (b['id'],)
            ).fetchall()
            resources.append({
                'id': f"pboat-{b['id']}", 'name': b['name'], 'type': 'picture_boat', 'group': 'Boats',
                'subgroup': b['group_name'] or 'Picture Boats',
                'assignments': [_timeline_assign(a) for a in assignments]
            })

        # --- Security boats ---
        sboats = conn.execute("SELECT id, name, group_name FROM security_boats WHERE production_id=?", (prod_id,)).fetchall()
        for b in sboats:
            assignments = conn.execute(
                "SELECT id, start_date, end_date, assignment_status, day_overrides, boat_function_id FROM security_boat_assignments WHERE security_boat_id=?",
                (b['id'],)
            ).fetchall()
            resources.append({
                'id': f"sboat-{b['id']}", 'name': b['name'], 'type': 'security_boat', 'group': 'Boats',
                'subgroup': b['group_name'] or 'Security Boats',
                'assignments': [_timeline_assign(a) for a in assignments]
            })

        # --- Vehicles ---
        vehicles = conn.execute("SELECT id, name, type FROM transport_vehicles WHERE production_id=?", (prod_id,)).fetchall()
        for v in vehicles:
            assignments = conn.execute(
                "SELECT id, start_date, end_date, assignment_status, day_overrides, boat_function_id FROM transport_assignments WHERE vehicle_id=?",
                (v['id'],)
            ).fetchall()
            resources.append({
                'id': f"vehicle-{v['id']}", 'name': v['name'], 'type': 'vehicle', 'group': 'Vehicles',
                'subgroup': v['type'] or 'Vehicle',
                'assignments': [_timeline_assign(a) for a in assignments]
            })

        # --- Labour (helpers) ---
        helpers = conn.execute("SELECT id, name, role, group_name FROM helpers WHERE production_id=?", (prod_id,)).fetchall()
        for h in helpers:
            assignments = conn.execute(
                "SELECT id, start_date, end_date, assignment_status, day_overrides, boat_function_id FROM helper_assignments WHERE helper_id=?",
                (h['id'],)
            ).fetchall()
            resources.append({
                'id': f"helper-{h['id']}", 'name': h['name'], 'type': 'labour', 'group': 'Crew',
                'subgroup': h['group_name'] or h['role'] or 'Labour',
                'assignments': [_timeline_assign(a) for a in assignments]
            })

        # --- Guards (camp workers) ---
        guards = conn.execute("SELECT id, name, role FROM guard_camp_workers WHERE production_id=?", (prod_id,)).fetchall()
        for g in guards:
            assignments = conn.execute(
                "SELECT id, start_date, end_date, assignment_status, day_overrides, boat_function_id FROM guard_camp_assignments WHERE worker_id=?",
                (g['id'],)
            ).fetchall()
            resources.append({
                'id': f"guard-{g['id']}", 'name': g['name'], 'type': 'guard', 'group': 'Crew',
                'subgroup': g['role'] or 'Guards',
                'assignments': [_timeline_assign(a) for a in assignments]
            })

        # --- Locations ---
        locations = conn.execute("SELECT id, name, site FROM locations WHERE production_id=?", (prod_id,)).fetchall()
        for loc in locations:
            schedules = conn.execute(
                "SELECT id, date, prep, filming, wrap FROM location_schedules WHERE location_id=?",
                (loc['id'],)
            ).fetchall()
            loc_assignments = []
            for s in schedules:
                phases = []
                if s['prep']: phases.append('P')
                if s['filming']: phases.append('F')
                if s['wrap']: phases.append('W')
                if phases:
                    loc_assignments.append({
                        'id': s['id'], 'start_date': s['date'], 'end_date': s['date'],
                        'status': 'confirmed', 'phases': '/'.join(phases)
                    })
            resources.append({
                'id': f"loc-{loc['id']}", 'name': loc['name'], 'type': 'location', 'group': 'Locations',
                'subgroup': loc['site'] or 'Location',
                'assignments': loc_assignments
            })

        # Boat functions for label lookups
        functions = [dict(f) for f in conn.execute(
            "SELECT id, name, context FROM boat_functions WHERE production_id=?", (prod_id,)
        ).fetchall()]

    return jsonify({
        'start_date': start_date,
        'end_date': end_date,
        'shooting_days': days,
        'resources': resources,
        'functions': functions,
    })


def _timeline_assign(row):
    """Convert an assignment row to a timeline-friendly dict."""
    d = {
        'id': row['id'],
        'start_date': row['start_date'],
        'end_date': row['end_date'],
        'status': row['assignment_status'] or 'confirmed',
        'function_id': row['boat_function_id'],
    }
    if row['day_overrides']:
        try:
            d['day_overrides'] = json.loads(row['day_overrides']) if isinstance(row['day_overrides'], str) else row['day_overrides']
        except (json.JSONDecodeError, TypeError):
            pass
    return d


# ─── Daily Checklists ─────────────────────────────────────────────────────────

@app.route("/api/productions/<int:prod_id>/checklists/generate", methods=["POST"])
def api_generate_checklist(prod_id):
    date = request.args.get("date") or (request.json or {}).get("date")
    if not date:
        return jsonify({"error": "date parameter required"}), 400
    checklist = generate_daily_checklist(prod_id, date)
    return jsonify(checklist), 201


@app.route("/api/productions/<int:prod_id>/checklists", methods=["GET"])
def api_get_checklist(prod_id):
    date = request.args.get("date")
    if not date:
        return jsonify({"error": "date parameter required"}), 400
    checklist = get_daily_checklist(prod_id, date)
    if not checklist:
        return jsonify(None), 200
    return jsonify(checklist)


@app.route("/api/productions/<int:prod_id>/checklists/items/<int:item_id>/check", methods=["PUT"])
def api_check_checklist_item(prod_id, item_id):
    data = request.json or {}
    checked = data.get("checked", True)
    user_id = data.get("user_id")
    item = check_checklist_item(item_id, checked, user_id)
    if not item:
        return jsonify({"error": "Item not found"}), 404
    return jsonify(item)


# ─── Bootstrap ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import os
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    init_db()
    # Auth: run auth schema migrations (adds nickname, is_admin, project_memberships, etc.)
    from auth.models import migrate_auth_tables
    migrate_auth_tables()
    from data_loader import bootstrap
    bootstrap()
    # Auth: seed default users and project memberships
    from auth.seed import seed_auth_data
    seed_auth_data()
    # Auto-match boat photos from BATEAUX source images and uploads folder
    _ensure_boat_images_symlink()
    n = _auto_match_boat_photos()
    if n:
        print(f"Auto-matched {n} boat photo(s)")
    app.run(host="0.0.0.0", port=5002, debug=True)
